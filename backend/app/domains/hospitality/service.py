import json
import logging
import mimetypes
import re
from collections import defaultdict
from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from email import policy
from email.parser import BytesParser
from io import BytesIO
from pathlib import Path
from uuid import uuid4
from xml.etree import ElementTree
from zipfile import ZipFile

from fastapi import HTTPException, UploadFile, status
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.core.config import LOCAL_ENVS, get_settings
from app.core.mongo_utils import utcnow
from app.domains.hospitality.repository import HospitalityRepository
from app.domains.hospitality.schemas import (
    AIPricingPersistRequest,
    AlertResolveRequest,
    ContractPathIngestionRequest,
    ReconciliationImportPersistRequest,
    PromotionPathIngestionRequest,
    RuleGenerationRequest,
    SyncRequest,
    ValidationBatchRequest,
)

BOARD_TYPE_LOOKUP = {
    "ro": "RO",
    "room only": "RO",
    "bb": "BB",
    "bed & breakfast": "BB",
    "bed and breakfast": "BB",
    "hb": "HB",
    "half board": "HB",
    "fb": "FB",
    "full board": "FB",
    "ai": "AI",
    "all inclusive": "AI",
}

ROOM_MATCH_CANONICAL_REPLACEMENTS: tuple[tuple[re.Pattern[str], str], ...] = (
    (
        re.compile(
            r"\bseafront\b|\bsea\s*front\b|\bfront\s+sea(?:\s+view)?\b|\bfrontal\s+sea(?:\s+view)?\b|\bsea\s+facing\b",
            flags=re.IGNORECASE,
        ),
        "sea view",
    ),
    (
        re.compile(r"\bside\s+sea(?:\s+view)?\b|\bsea\s+side(?:\s+view)?\b", flags=re.IGNORECASE),
        "side sea view",
    ),
    (
        re.compile(r"\binland(?:\s+view)?\b|\bland\s+view\b", flags=re.IGNORECASE),
        "land view",
    ),
)

ROOM_MATCH_STOPWORDS = {
    "room",
    "with",
    "the",
    "and",
    "type",
    "category",
    "accommodation",
}

ROOM_TYPE_PATTERNS = [
    r"standard(?:\s+double|\s+twin)?\s+room",
    r"superior(?:\s+double|\s+twin)?\s+room",
    r"deluxe(?:\s+double|\s+twin)?\s+room",
    r"family\s+room",
    r"junior\s+suite",
    r"suite",
    r"double\s+room",
    r"twin\s+room",
    r"single\s+room",
]

PROMOTION_KEYWORDS = {
    "early bird": "Early Bird",
    "valentine": "Valentine Offer",
    "long stay": "Long Stay",
    "flash": "Flash Offer",
    "spo": "Special Offer",
}

RECONCILIATION_HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "reservation_id": ("reservation", "booking", "confirmation", "reference", "res id"),
    "hotel_code": ("hotel", "property", "hotel code"),
    "operator_code": ("operator", "tour operator", "to code", "operator code"),
    "room_type": ("room type", "room", "accommodation"),
    "board_type": ("board type", "board", "meal"),
    "booking_date": (
        "booking date",
        "booking datetime",
        "reservation date",
        "reservation created",
        "booked on",
        "book date",
        "date booked",
        "booking created",
        "created at",
        "created on",
        "issue date",
        "issued at",
    ),
    "check_in_date": ("check in", "check-in", "arrival date", "arrival", "start date"),
    "check_out_date": ("check out", "check-out", "departure date", "departure", "end date"),
    "stay_date": ("stay date", "arrival", "check in", "check-in", "date"),
    "nights": ("nights", "los", "length of stay"),
    "pax_adults": ("adults", "adult pax", "adult"),
    "pax_children": ("children", "child", "kids"),
    "actual_price": ("actual", "fidelio", "charged", "actual price", "amount", "price"),
    "contract_rate": ("contract rate", "contract price", "expected rate", "base rate"),
    "promo_code": ("promo", "offer", "promotion", "coupon"),
    "status": (
        "status",
        "booking status",
        "reservation status",
        "res status",
        "record status",
        "booking state",
        "reservation state",
        "status code",
        "is cancelled",
        "is canceled",
        "cancelled flag",
        "canceled flag",
        "cxl status",
        "void status",
        "no show status",
    ),
}

RECONCILIATION_STATUS_COLUMN_HINTS = (
    "status",
    "state",
    "cxl",
    "cancel",
    "void",
    "no show",
    "noshow",
)

RECONCILIATION_CANCELLED_STATUS_TOKENS = {
    "cancelled",
    "canceled",
    "cancellation",
    "cancelation",
    "cancel",
    "cxl",
    "cxlled",
    "void",
    "voided",
    "annulled",
    "annulated",
    "no show",
    "noshow",
    "did not arrive",
    "didnt arrive",
}

RECONCILIATION_CANCELLED_STATUS_COMPACT_TOKENS = {
    token.replace(" ", "") for token in RECONCILIATION_CANCELLED_STATUS_TOKENS
}

DEFAULT_AI_PRICING_SCHEMA: dict = {
    "type": "object",
    "properties": {
        "hotel_name": {"type": "string"},
        "tour_operator": {"type": "string"},
        "season_label": {"type": "string"},
        "room_types": {
            "type": "array",
            "items": {
                "anyOf": [
                    {"type": "string"},
                    {"type": "object", "properties": {"name": {"type": "string"}}, "required": ["name"]},
                ]
            },
        },
        "seasonal_pricing_periods": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "label": {"type": "string"},
                    "start_date": {"type": "string"},
                    "end_date": {"type": "string"},
                },
            },
        },
        "board_types": {
            "type": "array",
            "items": {
                "anyOf": [
                    {"type": "string"},
                    {
                        "type": "object",
                        "properties": {
                            "code": {"type": "string"},
                            "name": {"type": "string"},
                        },
                    },
                ]
            },
        },
        "pricing_lines": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "room_type": {"type": "string"},
                    "board_type": {"type": "string"},
                    "period_label": {"type": "string"},
                    "adult_price": {"type": "number"},
                    "currency": {"type": "string"},
                },
            },
        },
        "extra_guest_rules": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "description": {"type": "string"},
                    "condition": {"type": "string"},
                    "guest_type": {"type": "string"},
                    "guest_position": {"type": "number"},
                    "age_min": {"type": "number"},
                    "age_max": {"type": "number"},
                    "percent_of_adult": {"type": "number"},
                },
            },
        },
        "discounts": {
            "type": "array",
            "items": {
                "anyOf": [
                    {"type": "string"},
                    {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "description": {"type": "string"},
                            "percent": {"type": "number"},
                        },
                    },
                ]
            },
        },
        "supplements": {
            "type": "array",
            "items": {
                "anyOf": [
                    {"type": "string"},
                    {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "description": {"type": "string"},
                            "amount": {"type": "number"},
                            "currency": {"type": "string"},
                        },
                    },
                ]
            },
        },
        "marketing_contributions": {
            "type": "array",
            "items": {
                "anyOf": [
                    {"type": "string"},
                    {
                        "type": "object",
                        "properties": {
                            "description": {"type": "string"},
                            "amount": {"type": "number"},
                            "currency": {"type": "string"},
                        },
                    },
                ]
            },
        },
        "promotional_offers": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "description": {"type": "string"},
                    "discount_percent": {"type": "number"},
                    "start_date": {"type": "string"},
                    "end_date": {"type": "string"},
                    "booking_start_date": {"type": ["string", "null"]},
                    "booking_end_date": {"type": ["string", "null"]},
                    "arrival_start_date": {"type": ["string", "null"]},
                    "arrival_end_date": {"type": ["string", "null"]},
                    "non_cumulative": {"type": ["boolean", "null"]},
                    "scope": {"type": ["string", "null"]},
                    "applicable_room_types": {"type": "array", "items": {"type": "string"}},
                    "applicable_board_types": {"type": "array", "items": {"type": "string"}},
                },
            },
        },
        "notes": {"type": "string"},
    },
    "additionalProperties": True,
}

AI_CONTENT_RECOMMENDATION_SCHEMA: dict = {
    "type": "object",
    "properties": {
        "content_summary": {"type": "string"},
        "confidence": {"type": "string", "enum": ["low", "medium", "high"]},
        "detected_signals": {"type": "object"},
        "coverage_feedback": {"type": "array", "items": {"type": "string"}},
        "recommended_data": {"type": "object"},
        "suggested_schema": {"type": "object"},
        "suggested_schema_rationale": {"type": "string"},
        "suggested_mapping_instructions": {"type": "string"},
        "database_mapping": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "target_entity": {"type": "string"},
                    "source_path": {"type": "string"},
                    "destination_field": {"type": "string"},
                    "transform": {"type": "string"},
                    "required": {"type": "boolean"},
                    "note": {"type": "string"},
                },
                "required": [
                    "target_entity",
                    "source_path",
                    "destination_field",
                    "transform",
                    "required",
                ],
                "additionalProperties": True,
            },
        },
    },
    "required": [
        "content_summary",
        "confidence",
        "coverage_feedback",
        "recommended_data",
        "suggested_schema",
        "suggested_schema_rationale",
        "suggested_mapping_instructions",
        "database_mapping",
    ],
    "additionalProperties": True,
}

AI_PROMOTION_EXTRACTION_SCHEMA: dict = {
    "type": "object",
    "properties": {
        "offer_name": {"type": "string"},
        "description": {"type": "string"},
        "discount_percent": {"type": ["number", "null"]},
        "booking_start_date": {"type": ["string", "null"]},
        "booking_end_date": {"type": ["string", "null"]},
        "arrival_start_date": {"type": ["string", "null"]},
        "arrival_end_date": {"type": ["string", "null"]},
        "start_date": {"type": ["string", "null"]},
        "end_date": {"type": ["string", "null"]},
        "scope": {"type": "string"},
        "non_cumulative": {"type": "boolean"},
        "combinability_note": {"type": ["string", "null"]},
        "applicable_room_types": {"type": "array", "items": {"type": "string"}},
        "applicable_board_types": {"type": "array", "items": {"type": "string"}},
        "confidence": {"type": "string", "enum": ["low", "medium", "high"]},
    },
    "required": [
        "offer_name",
        "description",
        "discount_percent",
        "booking_start_date",
        "booking_end_date",
        "arrival_start_date",
        "arrival_end_date",
        "scope",
        "non_cumulative",
        "applicable_room_types",
        "applicable_board_types",
        "confidence",
    ],
    "additionalProperties": True,
}

AI_RECONCILIATION_MAPPING_SCHEMA: dict = {
    "type": "object",
    "properties": {
        "mapping_summary": {"type": "string"},
        "source_system": {"type": ["string", "null"]},
        "header_mapping": {
            "type": "object",
            "additionalProperties": {"type": "string"},
        },
        "lines": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "reservation_id": {"type": "string"},
                    "hotel_code": {"type": "string"},
                    "operator_code": {"type": "string"},
                    "room_type": {"type": "string"},
                    "board_type": {"type": "string"},
                    "status": {"type": ["string", "null"]},
                    "reservation_status": {"type": ["string", "null"]},
                    "booking_status": {"type": ["string", "null"]},
                    "is_cancelled": {"type": ["boolean", "null"]},
                    "booking_date": {"type": ["string", "null"]},
                    "check_in_date": {"type": ["string", "null"]},
                    "check_out_date": {"type": ["string", "null"]},
                    "stay_date": {"type": "string"},
                    "nights": {"type": ["number", "integer"]},
                    "pax_adults": {"type": ["number", "integer"]},
                    "pax_children": {"type": ["number", "integer"]},
                    "actual_price": {"type": "number"},
                    "contract_rate": {"type": ["number", "null"]},
                    "promo_code": {"type": ["string", "null"]},
                },
                "required": [
                    "reservation_id",
                    "room_type",
                    "board_type",
                    "stay_date",
                    "nights",
                    "pax_adults",
                    "pax_children",
                    "actual_price",
                ],
                "additionalProperties": True,
            },
        },
    },
    "required": ["mapping_summary", "header_mapping", "lines"],
    "additionalProperties": True,
}

CORE_EXTRACT_ARRAY_KEYS: tuple[str, ...] = (
    "room_types",
    "seasonal_pricing_periods",
    "board_types",
    "pricing_lines",
    "extra_guest_rules",
    "discounts",
    "supplements",
    "marketing_contributions",
    "promotional_offers",
)

CORE_EXTRACT_REQUIRED_KEYS: tuple[str, ...] = (
    "room_types",
    "seasonal_pricing_periods",
    "board_types",
    "pricing_lines",
    "extra_guest_rules",
)

EXTRACTION_ALIAS_MAP: dict[str, tuple[str, ...]] = {
    "room_types": ("room_types", "rooms", "room_categories", "room_type_list"),
    "seasonal_pricing_periods": ("seasonal_pricing_periods", "seasonal_periods", "periods", "pricing_periods", "date_ranges"),
    "board_types": ("board_types", "boards", "meal_plans", "board_basis", "board_types_list"),
    "pricing_lines": ("pricing_lines", "price_lines", "rate_lines", "rates", "prices", "price_list", "matrix_entries"),
    "extra_guest_rules": (
        "extra_guest_rules",
        "guest_rules",
        "occupancy_rules",
        "occupancy_adjustments",
        "child_rules",
        "adult_rules",
    ),
    "discounts": ("discounts", "discount_rules"),
    "supplements": ("supplements", "surcharges", "supplement_rules"),
    "marketing_contributions": ("marketing_contributions", "marketing", "marketing_fees"),
    "promotional_offers": ("promotional_offers", "promotions", "offers", "special_offers"),
}

CANONICAL_MAPPING_REQUIREMENTS = (
    "Output JSON MUST include these top-level keys: room_types, seasonal_pricing_periods, board_types, "
    "pricing_lines, extra_guest_rules, discounts, supplements, marketing_contributions, promotional_offers, notes. "
    "Use arrays for these entities; if a value is unknown use an empty array. For pricing_lines use adult_price "
    "numeric values and for extra_guest_rules use percent_of_adult numeric values whenever available. "
    "For promotional_offers include discount_percent and date windows. For early-booking offers, capture both "
    "booking_start_date/booking_end_date and arrival_start_date/arrival_end_date; if only one range is explicit in "
    "the source, mirror it to both windows. Available board types are HB, BB, FB."
)

CANONICAL_RECONCILIATION_MAPPING_REQUIREMENTS = (
    "Map by semantic meaning, not by exact column names or fixed column order. "
    "The sheet structure can vary (different headers, merged cells, reordered columns, optional columns). "
    "Produce a header_mapping that links the canonical fields to the best matching source columns when possible. "
    "Canonical fields are: reservation_id, room_type, board_type, booking_date, stay_date, nights, pax_adults, "
    "pax_children, actual_price, contract_rate, promo_code, status, hotel_code, operator_code. "
    "Map booking_date when a booking/reservation creation date exists (needed for early-booking eligibility). "
    "If a status column exists, map it to status and EXCLUDE rows where status indicates cancelled/canceled/cxl/void/no-show. "
    "If a dedicated nights/LOS column exists, map it to nights. "
    "If nights is not explicitly available, map check_in_date and check_out_date and derive nights from date difference. "
    "Treat meal plan / board basis / catering basis as board_type. "
    "Normalize board_type to common codes when possible (RO, BB, HB, FB, AI). "
    "Normalize stay_date to ISO format (YYYY-MM-DD) using date-like columns even when they are Excel serial values "
    "or locale-formatted text dates. "
    "Use contract context defaults for hotel_code and operator_code when sheet values are absent. "
    "Never invent prices or reservation IDs; if actual_price is missing or not trustworthy, exclude that row."
)

MAX_CONTRACT_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_PROMOTION_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_RECONCILIATION_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_PRICING_AI_UPLOAD_BYTES = 20 * 1024 * 1024
UPLOAD_LIMIT_MB_TO_BYTES = 1024 * 1024
UPLOAD_LIMIT_DEFAULTS_MB = {
    "contract": MAX_CONTRACT_UPLOAD_BYTES // UPLOAD_LIMIT_MB_TO_BYTES,
    "promotion": MAX_PROMOTION_UPLOAD_BYTES // UPLOAD_LIMIT_MB_TO_BYTES,
    "reconciliation": MAX_RECONCILIATION_UPLOAD_BYTES // UPLOAD_LIMIT_MB_TO_BYTES,
    "pricing_ai": MAX_PRICING_AI_UPLOAD_BYTES // UPLOAD_LIMIT_MB_TO_BYTES,
}

ALLOWED_CONTRACT_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xlsm", ".xls", ".txt", ".csv"}
ALLOWED_PROMOTION_EXTENSIONS = ALLOWED_CONTRACT_EXTENSIONS | {".eml", ".msg"}
ALLOWED_RECONCILIATION_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
SAFE_INLINE_RESPONSE_CONTENT_TYPES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "image/webp",
    "image/gif",
}
BLOCKED_RESPONSE_CONTENT_TYPES = {
    "text/html",
    "application/xhtml+xml",
    "application/javascript",
    "text/javascript",
    "application/ecmascript",
    "text/ecmascript",
    "image/svg+xml",
}

logger = logging.getLogger(__name__)


class HospitalityService:
    def __init__(self, db: AsyncIOMotorDatabase):
        self.repository = HospitalityRepository(db)
        self.settings = get_settings()

    async def ensure_indexes(self) -> None:
        await self.repository.create_indexes()

    @staticmethod
    def _coerce_upload_limit_mb(raw_value: object, *, fallback: int) -> int:
        if isinstance(raw_value, bool):
            return fallback
        if isinstance(raw_value, (int, float)):
            parsed = int(raw_value)
            if 1 <= parsed <= 100:
                return parsed
        return fallback

    async def get_upload_limits(self) -> dict:
        stored = await self.repository.get_upload_limits()
        contract_mb = self._coerce_upload_limit_mb(
            (stored or {}).get("contract_mb"),
            fallback=UPLOAD_LIMIT_DEFAULTS_MB["contract"],
        )
        promotion_mb = self._coerce_upload_limit_mb(
            (stored or {}).get("promotion_mb"),
            fallback=UPLOAD_LIMIT_DEFAULTS_MB["promotion"],
        )
        reconciliation_mb = self._coerce_upload_limit_mb(
            (stored or {}).get("reconciliation_mb"),
            fallback=UPLOAD_LIMIT_DEFAULTS_MB["reconciliation"],
        )
        pricing_ai_mb = self._coerce_upload_limit_mb(
            (stored or {}).get("pricing_ai_mb"),
            fallback=UPLOAD_LIMIT_DEFAULTS_MB["pricing_ai"],
        )
        return {
            "contract_mb": contract_mb,
            "promotion_mb": promotion_mb,
            "reconciliation_mb": reconciliation_mb,
            "pricing_ai_mb": pricing_ai_mb,
            "updated_at": (stored or {}).get("updated_at"),
            "updated_by_user_id": (stored or {}).get("updated_by_user_id"),
        }

    async def update_upload_limits(
        self,
        *,
        contract_mb: int,
        promotion_mb: int,
        reconciliation_mb: int,
        pricing_ai_mb: int,
        updated_by_user_id: str,
    ) -> dict:
        now = utcnow()
        stored = await self.repository.upsert_upload_limits(
            contract_mb=contract_mb,
            promotion_mb=promotion_mb,
            reconciliation_mb=reconciliation_mb,
            pricing_ai_mb=pricing_ai_mb,
            updated_at=now,
            updated_by_user_id=updated_by_user_id,
        )
        return {
            "contract_mb": int(stored.get("contract_mb") or contract_mb),
            "promotion_mb": int(stored.get("promotion_mb") or promotion_mb),
            "reconciliation_mb": int(stored.get("reconciliation_mb") or reconciliation_mb),
            "pricing_ai_mb": int(stored.get("pricing_ai_mb") or pricing_ai_mb),
            "updated_at": stored.get("updated_at"),
            "updated_by_user_id": stored.get("updated_by_user_id"),
        }

    async def cleanup_contract_related_data(self, *, deleted_by_user_id: str) -> dict:
        summary = await self.repository.delete_all_contract_related_data()
        return {
            "scope": "hospitality_contract_domain",
            "deleted_collections": summary.get("deleted_collections", {}),
            "total_deleted_documents": int(summary.get("total_deleted_documents") or 0),
            "preserved_collections": [
                "users",
                "invitations",
                "password_resets",
                "hospitality_settings",
            ],
            "deleted_by_user_id": deleted_by_user_id,
            "deleted_at": utcnow(),
        }

    async def _get_effective_upload_limits_bytes(self) -> dict[str, int]:
        limits = await self.get_upload_limits()
        return {
            "contract": int(limits["contract_mb"]) * UPLOAD_LIMIT_MB_TO_BYTES,
            "promotion": int(limits["promotion_mb"]) * UPLOAD_LIMIT_MB_TO_BYTES,
            "reconciliation": int(limits["reconciliation_mb"]) * UPLOAD_LIMIT_MB_TO_BYTES,
            "pricing_ai": int(limits["pricing_ai_mb"]) * UPLOAD_LIMIT_MB_TO_BYTES,
        }

    async def _resolve_hotel_identity(self, *, hotel_id: str | None, hotel_code: str | None) -> tuple[str, str]:
        normalized_hotel_id = (hotel_id or "").strip()
        normalized_hotel_code = (hotel_code or "").strip().upper()

        if normalized_hotel_id:
            hotel = await self.repository.get_hotel_by_id(normalized_hotel_id)
            if not hotel:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Hotel not found for id '{normalized_hotel_id}'.",
                )
            resolved_code = str(hotel.get("code") or "").strip().upper()
            if normalized_hotel_code and normalized_hotel_code != resolved_code:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=(
                        f"hotel_code '{normalized_hotel_code}' does not match selected hotel "
                        f"code '{resolved_code}'."
                    ),
                )
            return str(hotel["id"]), resolved_code

        if normalized_hotel_code:
            hotel = await self.repository.get_hotel_by_code(normalized_hotel_code)
            if not hotel:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=(
                        f"Hotel code '{normalized_hotel_code}' does not exist. "
                        "Create it from Hotel Management first."
                    ),
                )
            return str(hotel["id"]), str(hotel.get("code") or normalized_hotel_code).strip().upper()

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="hotel_id or hotel_code is required.",
        )

    def _log_internal_error(
        self,
        *,
        event: str,
        exc: Exception | None = None,
        context: dict[str, object] | None = None,
    ) -> str:
        reference_id = uuid4().hex[:12]
        if exc is not None:
            logger.exception(
                "hospitality_error event=%s ref=%s context=%s",
                event,
                reference_id,
                context or {},
            )
        else:
            logger.error(
                "hospitality_error event=%s ref=%s context=%s",
                event,
                reference_id,
                context or {},
            )
        return reference_id

    def _is_local_environment(self) -> bool:
        settings = getattr(self, "settings", None)
        if settings is None:
            return False
        app_env = str(getattr(settings, "app_env", "") or "").strip().lower()
        return app_env in LOCAL_ENVS

    @staticmethod
    def _sanitize_upstream_text(value: str, *, max_len: int = 120) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"sk-[A-Za-z0-9_-]{12,}", "sk-***", text)
        text = re.sub(r"(?i)(api[_-]?key\s*[:=]\s*)[^\s,;]+", r"\1***", text)
        text = re.sub(r"(?i)(authorization\s*[:=]\s*bearer\s+)[^\s,;]+", r"\1***", text)
        if len(text) > max_len:
            text = f"{text[: max_len - 3]}..."
        return text

    def _build_upstream_error_detail(self, exc: Exception) -> str:
        segments: list[str] = []

        exc_type = str(exc.__class__.__name__ or "ProviderError").strip()
        if exc_type:
            segments.append(exc_type)

        status_code = getattr(exc, "status_code", None)
        if isinstance(status_code, int):
            segments.append(f"status={status_code}")

        request_id = self._sanitize_upstream_text(str(getattr(exc, "request_id", "") or ""), max_len=36)
        if request_id:
            segments.append(f"request_id={request_id}")

        message = getattr(exc, "message", None)
        if not isinstance(message, str) or not message.strip():
            message = str(exc)
        safe_message = self._sanitize_upstream_text(message, max_len=100)
        if safe_message:
            segments.append(f"message={safe_message}")

        body = getattr(exc, "body", None)
        body_text = ""
        if body is not None:
            try:
                if isinstance(body, (dict, list)):
                    body_text = json.dumps(body, ensure_ascii=False, separators=(",", ":"))
                else:
                    body_text = str(body)
            except Exception:
                body_text = str(body)
        safe_body = self._sanitize_upstream_text(body_text, max_len=80)
        if safe_body:
            segments.append(f"body={safe_body}")

        detail = "; ".join(segment for segment in segments if segment)
        return self._sanitize_upstream_text(detail, max_len=170)

    def _raise_sanitized_http_error(
        self,
        *,
        status_code: int,
        user_message: str,
        event: str,
        exc: Exception | None = None,
        context: dict[str, object] | None = None,
    ) -> None:
        reference_id = self._log_internal_error(event=event, exc=exc, context=context)
        detail = f"{user_message} Reference: {reference_id}."
        if self._is_local_environment() and event.startswith("openai.") and exc is not None:
            upstream_detail = self._build_upstream_error_detail(exc)
            if upstream_detail:
                detail = f"{detail} Provider response: {upstream_detail}"
        raise HTTPException(
            status_code=status_code,
            detail=detail,
        ) from exc

    def _resolve_seed_file_path(self, *, path_value: str, label: str) -> Path:
        normalized_input = str(path_value or "").strip()
        if not normalized_input:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"{label} seed path is required.",
            )

        app_env = self.settings.app_env.strip().lower()
        root_value = str(self.settings.seed_ingestion_root or "").strip()
        if app_env not in LOCAL_ENVS:
            if not self.settings.seed_path_ingestion_enabled or not root_value:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Path-based ingestion is disabled in this environment.",
                )

        candidate = Path(normalized_input).expanduser()
        try:
            resolved = candidate.resolve()
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Invalid seed file path.",
                event="seed_path.invalid",
                exc=exc,
                context={"path_label": label},
            )

        if root_value:
            try:
                allowed_root = Path(root_value).expanduser().resolve()
            except Exception as exc:
                self._raise_sanitized_http_error(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    user_message="Seed path ingestion is not configured correctly.",
                    event="seed_path.root_invalid",
                    exc=exc,
                )

            if not allowed_root.exists() or not allowed_root.is_dir():
                self._raise_sanitized_http_error(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    user_message="Seed path ingestion is not configured correctly.",
                    event="seed_path.root_missing",
                    context={"root": str(allowed_root)},
                )

            try:
                resolved.relative_to(allowed_root)
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Requested seed file path is outside the allowed ingestion root.",
                )

        if not resolved.exists() or not resolved.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"{label} seed file was not found.",
            )

        return resolved

    def _ensure_file_payload(
        self,
        *,
        file_name: str,
        content: bytes,
        allowed_extensions: set[str],
        max_bytes: int,
        label: str,
        empty_message: str,
    ) -> str:
        normalized_name = Path(file_name or "upload").name.strip() or "upload"
        if not content:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=empty_message)

        content_size = len(content)
        if content_size > max_bytes:
            max_mb = max_bytes // (1024 * 1024)
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"{label} file is too large. Maximum allowed size is {max_mb} MB.",
            )

        suffix = Path(normalized_name).suffix.lower()
        if suffix not in allowed_extensions:
            allowed = ", ".join(sorted(allowed_extensions))
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Unsupported {label} file type '{suffix or '<none>'}'. Allowed types: {allowed}.",
            )

        return normalized_name

    async def ingest_contract(
        self,
        file_name: str,
        content: bytes,
        content_type: str | None,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        season_label: str | None,
        uploaded_by_user_id: str,
        source: str = "upload",
    ) -> dict:
        upload_limits = await self._get_effective_upload_limits_bytes()
        normalized_file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_CONTRACT_EXTENSIONS,
            max_bytes=upload_limits["contract"],
            label="contract",
            empty_message="Contract file is empty",
        )

        resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
            hotel_id=hotel_id,
            hotel_code=hotel_code,
        )
        normalized_operator = operator_code.strip().upper()

        extracted_text = self._extract_text_from_bytes(file_name=normalized_file_name, content=content)
        extraction = self._extract_contract_terms(extracted_text)

        now = utcnow()
        contract = {
            "file_name": normalized_file_name,
            "file_type": Path(normalized_file_name).suffix.lower() or "unknown",
            "file_size": len(content),
            "has_uploaded_file": False,
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": normalized_operator,
            "season_label": season_label.strip() if season_label else None,
            "source": source,
            "uploaded_by_user_id": uploaded_by_user_id,
            "extraction": extraction,
            "parsed_text_preview": extracted_text,
            "created_at": now,
            "updated_at": now,
        }
        created_contract = await self.repository.create_contract(contract)
        stored_content_type = self._guess_content_type(file_name=normalized_file_name, fallback=content_type)
        await self.repository.upsert_contract_file(
            contract_id=created_contract["id"],
            file_name=normalized_file_name,
            file_type=contract["file_type"],
            content_type=stored_content_type,
            file_size=len(content),
            content=content,
            uploaded_by_user_id=uploaded_by_user_id,
            updated_at=now,
        )
        updated_contract = await self.repository.set_contract_uploaded_file_flag(
            contract_id=created_contract["id"],
            has_uploaded_file=True,
            updated_at=now,
        )
        if updated_contract:
            return updated_contract
        created_contract["has_uploaded_file"] = True
        return created_contract

    async def ingest_contract_upload(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        season_label: str | None,
        uploaded_by_user_id: str,
    ) -> dict:
        file_name = upload_file.filename or "contract-upload"
        content = await upload_file.read()
        return await self.ingest_contract(
            file_name=file_name,
            content=content,
            content_type=upload_file.content_type,
            hotel_id=hotel_id,
            hotel_code=hotel_code,
            operator_code=operator_code,
            season_label=season_label,
            uploaded_by_user_id=uploaded_by_user_id,
            source="upload",
        )

    async def ingest_contracts_from_paths(self, payload: ContractPathIngestionRequest, uploaded_by_user_id: str) -> list[dict]:
        results: list[dict] = []
        for path_value in payload.paths:
            path = self._resolve_seed_file_path(path_value=path_value, label="Contract")
            content = path.read_bytes()
            results.append(
                await self.ingest_contract(
                    file_name=path.name,
                    content=content,
                    content_type=None,
                    hotel_id=payload.hotel_id,
                    hotel_code=payload.hotel_code,
                    operator_code=payload.operator_code,
                    season_label=payload.season_label,
                    uploaded_by_user_id=uploaded_by_user_id,
                    source="seed",
                )
            )
        return results

    async def list_contracts(
        self,
        search_text: str | None = None,
        hotel_id: str | None = None,
        hotel_code: str | None = None,
        operator_code: str | None = None,
        season_label: str | None = None,
        source: str | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
        limit: int = 300,
    ) -> list[dict]:
        return await self.repository.list_contracts(
            search_text=search_text,
            hotel_id=hotel_id,
            hotel_code=hotel_code,
            operator_code=operator_code,
            season_label=season_label,
            source=source,
            sort_by=sort_by,
            sort_order=sort_order,
            limit=limit,
        )

    async def get_contract(self, contract_id: str) -> dict:
        contract = await self.repository.get_contract(contract_id)
        if not contract:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contract not found")
        return contract

    async def get_contract_uploaded_file(self, contract_id: str) -> dict:
        contract = await self.get_contract(contract_id)
        file_record = await self.repository.get_contract_file(contract["id"])
        if not file_record:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Uploaded contract file was not found for this record.",
            )

        content_bytes = self._coerce_binary_content(file_record.get("content"))
        if not content_bytes:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Uploaded contract file content is missing.",
            )

        stored_name = str(file_record.get("file_name") or contract.get("file_name") or f"contract-{contract['id']}")
        stored_type = str(file_record.get("file_type") or contract.get("file_type") or "")
        content_type = self._guess_content_type(
            file_name=stored_name,
            fallback=str(file_record.get("content_type") or ""),
        )
        if content_type in BLOCKED_RESPONSE_CONTENT_TYPES:
            content_type = "application/octet-stream"

        return {
            "contract_id": contract["id"],
            "file_name": stored_name,
            "file_type": stored_type,
            "content_type": content_type,
            "inline_safe": content_type in SAFE_INLINE_RESPONSE_CONTENT_TYPES,
            "file_size": int(file_record.get("file_size") or len(content_bytes)),
            "content": content_bytes,
        }

    async def get_contract_price_matrix(
        self,
        contract_id: str,
        include_promotions: bool = False,
        promotion_ids: list[str] | None = None,
        booking_date: date | None = None,
    ) -> dict:
        contract = await self.get_contract(contract_id)
        rules = await self.repository.list_rules(contract_id=contract["id"])
        ai_extracted_data = contract.get("ai_extracted_data")
        if not isinstance(ai_extracted_data, dict):
            ai_extracted_data = {}

        period_ranges = self._extract_price_period_ranges(ai_extracted_data)
        period_lookup = {period["label"].strip().lower(): period for period in period_ranges if period.get("label")}

        base_entries = self._extract_base_price_entries(
            rules=rules,
            ai_extracted_data=ai_extracted_data,
            period_lookup=period_lookup,
        )
        adjustments = self._extract_age_adjustments(rules=rules, ai_extracted_data=ai_extracted_data)
        derived_entries = self._derive_adjusted_price_entries(base_entries=base_entries, adjustments=adjustments)

        entries = [*base_entries, *derived_entries]
        selected_promotion_ids = sorted(
            {
                item.strip()
                for item in (promotion_ids or [])
                if isinstance(item, str) and item.strip()
            }
        )

        applied_promotion_ids: list[str] = []
        applied_promotion_names: list[str] = []
        if include_promotions:
            promotions = await self.repository.list_promotions_for_contract(
                operator_code=contract["operator_code"],
                hotel_code=contract["hotel_code"],
                hotel_id=contract.get("hotel_id"),
                contract_id=contract["id"],
            )
            if selected_promotion_ids:
                selected_set = set(selected_promotion_ids)
                promotions = [promo for promo in promotions if promo.get("id") in selected_set]

            entries, used_promotion_ids, used_promotion_names = self._apply_promotions_to_matrix_entries(
                entries=entries,
                promotions=promotions,
                booking_date=booking_date,
            )
            applied_promotion_ids = sorted(used_promotion_ids)
            applied_promotion_names = sorted(used_promotion_names)

        entries.sort(
            key=lambda row: (
                str(row.get("room_type") or ""),
                str(row.get("board_type") or ""),
                str(row.get("age_bucket") or ""),
                str(row.get("period_label") or ""),
                str(row.get("start_date") or ""),
                float(row.get("price") or 0.0),
            )
        )

        room_types = sorted({str(item.get("room_type") or "").strip() for item in entries if item.get("room_type")})
        board_types = sorted({str(item.get("board_type") or "").strip() for item in entries if item.get("board_type")})
        age_buckets = sorted({str(item.get("age_bucket") or "").strip() for item in entries if item.get("age_bucket")})
        currencies = sorted({str(item.get("currency") or "").strip() for item in entries if item.get("currency")})

        return {
            "contract_id": contract["id"],
            "contract_file_name": contract.get("file_name", ""),
            "hotel_id": contract.get("hotel_id"),
            "hotel_code": contract.get("hotel_code", ""),
            "operator_code": contract.get("operator_code", ""),
            "season_label": contract.get("season_label"),
            "period_ranges": period_ranges,
            "room_types": room_types,
            "board_types": board_types,
            "age_buckets": age_buckets,
            "currencies": currencies,
            "include_promotions": include_promotions,
            "booking_date": booking_date.isoformat() if booking_date else None,
            "selected_promotion_ids": selected_promotion_ids,
            "applied_promotion_ids": applied_promotion_ids,
            "applied_promotion_names": applied_promotion_names,
            "entries": entries,
        }

    async def ingest_promotion(
        self,
        file_name: str,
        content: bytes,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
    ) -> dict:
        upload_limits = await self._get_effective_upload_limits_bytes()
        normalized_file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_PROMOTION_EXTENSIONS,
            max_bytes=upload_limits["promotion"],
            label="promotion",
            empty_message="Promotion file is empty",
        )

        resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
            hotel_id=hotel_id,
            hotel_code=hotel_code,
        )
        extracted_text = self._extract_text_from_bytes(file_name=normalized_file_name, content=content)
        parsed = self._extract_promotion_offer(extracted_text, fallback_name=normalized_file_name)

        now = utcnow()
        promotion = {
            "file_name": normalized_file_name,
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": operator_code.strip().upper(),
            "offer_name": parsed["offer_name"],
            "description": parsed["description"],
            "discount_percent": parsed["discount_percent"],
            "start_date": parsed["start_date"],
            "end_date": parsed["end_date"],
            "booking_start_date": parsed.get("booking_start_date"),
            "booking_end_date": parsed.get("booking_end_date"),
            "arrival_start_date": parsed.get("arrival_start_date"),
            "arrival_end_date": parsed.get("arrival_end_date"),
            "non_cumulative": parsed.get("non_cumulative"),
            "combinability_note": parsed.get("combinability_note"),
            "promotion_category": parsed.get("promotion_category"),
            "applicable_room_types": parsed.get("applicable_room_types", []),
            "applicable_board_types": parsed.get("applicable_board_types", []),
            "affected_contract_ids": [],
            "ingestion_mode": "heuristic_promotion_ingestion",
            "scope": parsed["scope"],
            "created_at": now,
            "updated_at": now,
        }
        return await self.repository.create_promotion(self._to_bson_compatible(promotion))

    async def ingest_promotion_upload(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
    ) -> dict:
        file_name = upload_file.filename or "promotion-upload"
        content = await upload_file.read()
        return await self.ingest_promotion(
            file_name=file_name,
            content=content,
            hotel_id=hotel_id,
            hotel_code=hotel_code,
            operator_code=operator_code,
        )

    async def ingest_promotions_from_paths(self, payload: PromotionPathIngestionRequest) -> list[dict]:
        results: list[dict] = []
        for path_value in payload.paths:
            path = self._resolve_seed_file_path(path_value=path_value, label="Promotion")
            content = path.read_bytes()
            results.append(
                await self.ingest_promotion(
                    file_name=path.name,
                    content=content,
                    hotel_id=payload.hotel_id,
                    hotel_code=payload.hotel_code,
                    operator_code=payload.operator_code,
                )
            )
        return results

    async def list_promotions(
        self,
        hotel_id: str | None = None,
        contract_id: str | None = None,
        operator_code: str | None = None,
    ) -> list[dict]:
        if contract_id and contract_id.strip():
            contract = await self.get_contract(contract_id.strip())
            return await self.repository.list_promotions_for_contract(
                operator_code=contract.get("operator_code", ""),
                hotel_code=contract.get("hotel_code", ""),
                hotel_id=contract.get("hotel_id"),
                contract_id=contract["id"],
            )
        return await self.repository.list_promotions(
            hotel_id=hotel_id,
            operator_code=operator_code,
            contract_id=None,
        )

    async def ai_ingest_promotion_upload(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        contract_ids: list[str] | None,
        created_by_user_id: str,
    ) -> dict:
        file_name = upload_file.filename or "promotion-email.txt"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_PROMOTION_EXTENSIONS,
            max_bytes=upload_limits["promotion"],
            label="promotion",
            empty_message="Promotion file is empty",
        )

        resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
            hotel_id=hotel_id,
            hotel_code=hotel_code,
        )
        normalized_operator = operator_code.strip().upper()
        extracted_text = self._extract_text_from_bytes(file_name=file_name, content=content)
        if not extracted_text.strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Could not extract text from the uploaded promotion file.",
            )

        target_contracts = await self._resolve_promotion_target_contracts(
            contract_ids=contract_ids or [],
            hotel_id=resolved_hotel_id,
            hotel_code=resolved_hotel_code,
            operator_code=normalized_operator,
        )

        heuristic = self._extract_promotion_offer(extracted_text, fallback_name=file_name)
        ai_result, usage, model = await self._run_openai_promotion_extraction(
            text=extracted_text,
            hotel_code=resolved_hotel_code,
            operator_code=normalized_operator,
            fallback_offer=heuristic,
        )
        merged_offer = self._merge_promotion_offer_data(
            ai_offer=ai_result,
            fallback_offer=heuristic,
            fallback_name=file_name,
        )

        now = utcnow()
        affected_contract_ids = [contract["id"] for contract in target_contracts]
        promotion = {
            "file_name": file_name,
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": normalized_operator,
            "offer_name": merged_offer["offer_name"],
            "description": merged_offer["description"],
            "discount_percent": merged_offer["discount_percent"],
            "start_date": merged_offer["start_date"],
            "end_date": merged_offer["end_date"],
            "booking_start_date": merged_offer["booking_start_date"],
            "booking_end_date": merged_offer["booking_end_date"],
            "arrival_start_date": merged_offer["arrival_start_date"],
            "arrival_end_date": merged_offer["arrival_end_date"],
            "non_cumulative": merged_offer["non_cumulative"],
            "combinability_note": merged_offer["combinability_note"],
            "promotion_category": merged_offer["promotion_category"],
            "applicable_room_types": merged_offer["applicable_room_types"],
            "applicable_board_types": merged_offer["applicable_board_types"],
            "affected_contract_ids": affected_contract_ids,
            "scope": merged_offer["scope"],
            "ingestion_mode": "ai_email_ingestion",
            "analysis_provider": "openai",
            "analysis_model": model,
            "analysis_usage": usage,
            "ai_extracted_data": ai_result,
            "parsed_text_preview": extracted_text[:12000],
            "created_by_user_id": created_by_user_id,
            "created_at": now,
            "updated_at": now,
        }
        created_promotion = await self.repository.create_promotion(self._to_bson_compatible(promotion))
        contract_rule_updates = await self._append_promotion_rule_to_contracts(
            promotion=created_promotion,
            contracts=target_contracts,
        )
        category_label = str(created_promotion.get("promotion_category") or "").strip().lower()
        if not category_label:
            category_label = "early_booking" if self._is_early_booking_promotion(created_promotion) else "general"
        booking_window = (
            f"{created_promotion.get('booking_start_date') or '?'} to {created_promotion.get('booking_end_date') or '?'}"
        )
        arrival_window = (
            f"{created_promotion.get('arrival_start_date') or created_promotion.get('start_date') or '?'} "
            f"to {created_promotion.get('arrival_end_date') or created_promotion.get('end_date') or '?'}"
        )

        return {
            "promotion": created_promotion,
            "analysis_summary": (
                f"AI extracted promotion '{created_promotion.get('offer_name')}' "
                f"({created_promotion.get('discount_percent') or 0}% discount) "
                f"[{category_label}] booking window {booking_window}; arrival window {arrival_window}; "
                f"for {len(affected_contract_ids)} contract(s)."
            ),
            "impacted_contract_ids": affected_contract_ids,
            "contract_rule_updates": contract_rule_updates,
        }

    async def _resolve_promotion_target_contracts(
        self,
        *,
        contract_ids: list[str],
        hotel_id: str,
        hotel_code: str,
        operator_code: str,
    ) -> list[dict]:
        normalized_ids = sorted({item.strip() for item in contract_ids if isinstance(item, str) and item.strip()})
        if normalized_ids:
            contracts = await self.repository.get_contracts_by_ids(normalized_ids)
            found_ids = {contract["id"] for contract in contracts}
            missing_ids = [item for item in normalized_ids if item not in found_ids]
            if missing_ids:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Some selected contracts do not exist: {', '.join(missing_ids)}",
                )
            invalid_targets = [
                contract
                for contract in contracts
                if str(contract.get("hotel_id") or "") != hotel_id
                or str(contract.get("operator_code") or "").strip().upper() != operator_code
            ]
            if invalid_targets:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Selected contracts must belong to the same hotel and operator as the promotion.",
                )
            return sorted(contracts, key=lambda item: str(item.get("file_name") or ""))

        contracts = await self.repository.list_contracts(
            hotel_id=hotel_id,
            hotel_code=hotel_code,
            operator_code=operator_code,
            limit=1000,
            sort_by="updated_at",
            sort_order="desc",
        )
        return contracts

    async def _append_promotion_rule_to_contracts(self, promotion: dict, contracts: list[dict]) -> list[dict]:
        updates: list[dict] = []
        promotion_id = str(promotion.get("id") or "")
        discount = self._coerce_float(promotion.get("discount_percent"))
        if not promotion_id or discount is None or discount <= 0:
            for contract in contracts:
                existing_rules = await self.repository.list_rules(contract_id=contract["id"])
                updates.append(
                    {
                        "contract_id": contract["id"],
                        "contract_file_name": str(contract.get("file_name") or ""),
                        "promotion_rule_added": False,
                        "total_rules_after_update": len(existing_rules),
                    }
                )
            return updates

        for contract in contracts:
            existing_rules = await self.repository.list_rules(contract_id=contract["id"])
            already_exists = any(
                isinstance(rule.get("metadata"), dict) and str(rule["metadata"].get("promotion_id") or "") == promotion_id
                for rule in existing_rules
                if isinstance(rule, dict)
            )
            if already_exists:
                updates.append(
                    {
                        "contract_id": contract["id"],
                        "contract_file_name": str(contract.get("file_name") or ""),
                        "promotion_rule_added": False,
                        "total_rules_after_update": len(existing_rules),
                    }
                )
                continue

            now = utcnow()
            persisted_rules = [self._materialize_rule_for_replace(rule, fallback_contract_id=contract["id"], now=now) for rule in existing_rules]
            promotion_rule = self._build_promotion_rule_document(
                contract=contract,
                promotion=promotion,
                created_at=now,
            )
            persisted_rules.append(promotion_rule)
            updated_rules = await self.repository.replace_rules(contract_id=contract["id"], rules=persisted_rules)
            updates.append(
                {
                    "contract_id": contract["id"],
                    "contract_file_name": str(contract.get("file_name") or ""),
                    "promotion_rule_added": True,
                    "total_rules_after_update": len(updated_rules),
                }
            )

        return updates

    def _materialize_rule_for_replace(self, rule: dict, fallback_contract_id: str, now: datetime) -> dict:
        metadata = rule.get("metadata")
        created_at = rule.get("created_at")
        updated_at = rule.get("updated_at")
        return {
            "contract_id": str(rule.get("contract_id") or fallback_contract_id),
            "hotel_id": rule.get("hotel_id"),
            "name": str(rule.get("name") or "Rule"),
            "rule_type": str(rule.get("rule_type") or "base_rate"),
            "expression": str(rule.get("expression") or ""),
            "priority": int(rule.get("priority") or 100),
            "is_active": bool(rule.get("is_active", True)),
            "metadata": metadata if isinstance(metadata, dict) else {},
            "created_at": created_at if isinstance(created_at, datetime) else now,
            "updated_at": updated_at if isinstance(updated_at, datetime) else now,
        }

    def _build_promotion_rule_document(self, contract: dict, promotion: dict, created_at: datetime) -> dict:
        discount = self._coerce_float(promotion.get("discount_percent")) or 0.0
        start_date = promotion.get("arrival_start_date") or promotion.get("start_date")
        end_date = promotion.get("arrival_end_date") or promotion.get("end_date")
        offer_name = str(promotion.get("offer_name") or "Promotion").strip() or "Promotion"
        is_early_booking = self._is_early_booking_promotion(promotion)
        expression = (
            f"if booking_date and stay_date are within offer window then discount {discount:.2f}%"
            if is_early_booking
            else f"if promotion active then discount {discount:.2f}%"
        )
        return {
            "contract_id": contract["id"],
            "hotel_id": contract.get("hotel_id"),
            "name": offer_name,
            "rule_type": "promotion",
            "expression": expression,
            "priority": 35,
            "is_active": True,
            "metadata": {
                "promotion_id": promotion.get("id"),
                "offer_name": offer_name,
                "discount_percent": round(discount, 2),
                "start_date": start_date,
                "end_date": end_date,
                "scope": promotion.get("scope", "all"),
                "non_cumulative": bool(promotion.get("non_cumulative", False)),
                "booking_start_date": promotion.get("booking_start_date"),
                "booking_end_date": promotion.get("booking_end_date"),
                "arrival_start_date": promotion.get("arrival_start_date"),
                "arrival_end_date": promotion.get("arrival_end_date"),
                "applicable_room_types": promotion.get("applicable_room_types", []),
                "applicable_board_types": promotion.get("applicable_board_types", []),
                "promotion_category": "early_booking" if is_early_booking else "general",
            },
            "created_at": created_at,
            "updated_at": created_at,
        }

    async def ai_recommend_pricing_content(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        season_label: str | None,
    ) -> dict:
        file_name = upload_file.filename or "pricing-contract.pdf"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_CONTRACT_EXTENSIONS,
            max_bytes=upload_limits["pricing_ai"],
            label="contract",
            empty_message="Contract file is empty",
        )

        text = self._extract_text_from_bytes(file_name=file_name, content=content)
        if not text.strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Could not extract text from the uploaded contract file.",
            )

        resolved_hotel_id: str | None = None
        resolved_hotel_code: str | None = None
        if (hotel_id and hotel_id.strip()) or (hotel_code and hotel_code.strip()):
            resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
                hotel_id=hotel_id,
                hotel_code=hotel_code,
            )

        heuristic_recommendation = self._recommend_content_for_mapping(text=text)
        ai_recommendation, usage, analysis_model = await self._run_openai_content_recommendation(
            text=text,
            hotel_code=(resolved_hotel_code or hotel_code or "").strip().upper(),
            operator_code=operator_code.strip().upper(),
            season_label=season_label,
            baseline_recommendation=heuristic_recommendation,
        )
        recommendation = self._merge_content_recommendation(
            ai_recommendation=ai_recommendation,
            fallback=heuristic_recommendation,
        )
        return {
            "file_name": file_name,
            "file_type": Path(file_name).suffix.lower() or "unknown",
            "file_size": len(content),
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": operator_code.strip().upper(),
            "season_label": season_label.strip() if season_label else None,
            "analysis_provider": "openai",
            "analysis_model": analysis_model,
            "analysis_usage": usage,
            "content_summary": recommendation["content_summary"],
            "confidence": recommendation["confidence"],
            "detected_signals": recommendation["detected_signals"],
            "coverage_feedback": recommendation["coverage_feedback"],
            "recommended_data": recommendation["recommended_data"],
            "suggested_schema": recommendation["suggested_schema"],
            "suggested_schema_rationale": recommendation["suggested_schema_rationale"],
            "suggested_mapping_instructions": recommendation["suggested_mapping_instructions"],
            "database_mapping": recommendation["database_mapping"],
        }

    async def ai_recommend_pricing_model(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        season_label: str | None,
    ) -> dict:
        # Backward-compatible alias. Recommendation now focuses on content/data mapping.
        return await self.ai_recommend_pricing_content(
            upload_file=upload_file,
            hotel_id=hotel_id,
            hotel_code=hotel_code,
            operator_code=operator_code,
            season_label=season_label,
        )

    async def ai_extract_pricing_contract(
        self,
        upload_file: UploadFile,
        hotel_id: str | None,
        hotel_code: str | None,
        operator_code: str,
        season_label: str | None,
        model: str | None,
        schema_override: dict | None,
        mapping_instructions: str | None,
        created_by_user_id: str,
    ) -> dict:
        file_name = upload_file.filename or "pricing-contract.pdf"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_CONTRACT_EXTENSIONS,
            max_bytes=upload_limits["pricing_ai"],
            label="contract",
            empty_message="Contract file is empty",
        )

        text = self._extract_text_from_bytes(file_name=file_name, content=content)
        if not text.strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Could not extract text from the uploaded contract file.",
            )

        resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
            hotel_id=hotel_id,
            hotel_code=hotel_code,
        )
        schema_used = self._build_effective_extraction_schema(schema_override)
        effective_mapping_instructions = self._build_effective_mapping_instructions(mapping_instructions)
        extracted_data_raw, usage, selected_model = await self._run_openai_structured_extraction(
            text=text,
            model=model,
            schema=schema_used,
            hotel_code=resolved_hotel_code,
            operator_code=operator_code,
            season_label=season_label,
            mapping_instructions=effective_mapping_instructions,
        )
        extracted_data = self._stabilize_extracted_data_shape(extracted_data_raw)
        if not extracted_data.get("pricing_lines") and not extracted_data.get("extra_guest_rules"):
            fallback_recommendation = self._recommend_content_for_mapping(text=text)
            fallback_data = fallback_recommendation.get("recommended_data")
            if isinstance(fallback_data, dict):
                extracted_data = self._merge_extracted_with_fallback(extracted_data, fallback_data)

        normalized_extraction = self._normalize_ai_extraction(extracted_data)
        suggested_rules = self._build_rule_drafts_from_ai_data(extracted_data)

        now = utcnow()
        run = {
            "file_name": file_name,
            "file_size": len(content),
            "file_type": Path(file_name).suffix.lower() or "unknown",
            "uploaded_file_name": file_name,
            "uploaded_file_content_type": self._guess_content_type(file_name=file_name, fallback=upload_file.content_type),
            "uploaded_file_content": content,
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": operator_code.strip().upper(),
            "season_label": season_label.strip() if season_label else None,
            "model": selected_model,
            "parsed_text_preview": text,
            "schema_used": schema_used,
            "extracted_data": extracted_data,
            "raw_extracted_data": extracted_data_raw,
            "normalized_extraction": normalized_extraction,
            "suggested_rules": suggested_rules,
            "usage": usage,
            "created_by_user_id": created_by_user_id,
            "created_at": now,
            "persisted_contract_id": None,
            "persisted_at": None,
            "persisted_by_user_id": None,
        }
        return await self.repository.create_ai_extraction(run)

    async def persist_ai_extraction(self, payload: AIPricingPersistRequest, persisted_by_user_id: str) -> dict:
        extraction_run = await self.repository.get_ai_extraction(payload.extraction_run_id)
        if not extraction_run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI extraction run not found")

        raw_effective_data = payload.reviewed_data if payload.reviewed_data is not None else extraction_run.get("extracted_data", {})
        effective_data = self._stabilize_extracted_data_shape(raw_effective_data if isinstance(raw_effective_data, dict) else {})
        if not effective_data.get("pricing_lines") and not effective_data.get("extra_guest_rules"):
            parsed_text = str(extraction_run.get("parsed_text_preview") or "").strip()
            if parsed_text:
                fallback_recommendation = self._recommend_content_for_mapping(text=parsed_text)
                fallback_data = fallback_recommendation.get("recommended_data")
                if isinstance(fallback_data, dict):
                    effective_data = self._merge_extracted_with_fallback(effective_data, fallback_data)
        normalized_extraction = self._normalize_ai_extraction(effective_data)
        resolved_hotel_id, resolved_hotel_code = await self._resolve_hotel_identity(
            hotel_id=payload.hotel_id or extraction_run.get("hotel_id"),
            hotel_code=payload.hotel_code or extraction_run.get("hotel_code"),
        )

        now = utcnow()
        contract_doc = {
            "file_name": extraction_run.get("file_name", "ai-pricing-contract"),
            "file_type": extraction_run.get("file_type", "unknown"),
            "file_size": int(extraction_run.get("file_size") or 0),
            "has_uploaded_file": False,
            "hotel_id": resolved_hotel_id,
            "hotel_code": resolved_hotel_code,
            "operator_code": payload.operator_code.strip().upper(),
            "season_label": payload.season_label.strip() if payload.season_label else extraction_run.get("season_label"),
            "source": "upload",
            "uploaded_by_user_id": persisted_by_user_id,
            "extraction": normalized_extraction,
            "parsed_text_preview": extraction_run.get("parsed_text_preview", ""),
            "ai_extracted_data": effective_data,
            "created_at": now,
            "updated_at": now,
            "ingestion_mode": "ai_pricing_ingestion",
            "ai_extraction_run_id": extraction_run["id"],
        }
        created_contract = await self.repository.create_contract(contract_doc)

        stored_content = self._coerce_binary_content(extraction_run.get("uploaded_file_content"))
        if stored_content:
            stored_file_name = str(extraction_run.get("uploaded_file_name") or contract_doc["file_name"] or "ai-pricing-contract")
            stored_file_type = str(extraction_run.get("file_type") or contract_doc["file_type"] or "unknown")
            stored_content_type = self._guess_content_type(
                file_name=stored_file_name,
                fallback=str(extraction_run.get("uploaded_file_content_type") or ""),
            )
            await self.repository.upsert_contract_file(
                contract_id=created_contract["id"],
                file_name=stored_file_name,
                file_type=stored_file_type,
                content_type=stored_content_type,
                file_size=int(extraction_run.get("file_size") or len(stored_content)),
                content=stored_content,
                uploaded_by_user_id=persisted_by_user_id,
                updated_at=now,
            )
            updated_contract = await self.repository.set_contract_uploaded_file_flag(
                contract_id=created_contract["id"],
                has_uploaded_file=True,
                updated_at=now,
            )
            if updated_contract:
                created_contract = updated_contract
            else:
                created_contract["has_uploaded_file"] = True

        rules_payload = self._build_rules_from_ai_data(
            extracted_data=effective_data,
            contract_id=created_contract["id"],
            hotel_id=resolved_hotel_id,
            created_at=now,
        )
        created_rules = await self.repository.replace_rules(contract_id=created_contract["id"], rules=rules_payload)

        await self.repository.mark_ai_extraction_persisted(
            extraction_id=extraction_run["id"],
            contract_id=created_contract["id"],
            persisted_at=now,
            persisted_by_user_id=persisted_by_user_id,
        )

        return {
            "contract": created_contract,
            "created_rules": created_rules,
            "extraction_run_id": extraction_run["id"],
            "model": extraction_run.get("model", ""),
        }

    async def generate_rules(self, payload: RuleGenerationRequest) -> list[dict]:
        contract = await self.get_contract(payload.contract_id)
        extraction = contract.get("extraction", {})

        promotions: list[dict] = []
        if payload.include_promotions:
            promotions = await self.repository.list_promotions_for_contract(
                operator_code=contract["operator_code"],
                hotel_code=contract["hotel_code"],
                hotel_id=contract.get("hotel_id"),
                contract_id=contract["id"],
            )

        rates = self._extract_rate_candidates(contract.get("parsed_text_preview", ""))
        room_types = extraction.get("room_types") or ["Standard Room", "Family Room", "Suite"]
        board_types = extraction.get("board_types") or ["RO", "BB", "HB"]

        now = utcnow()
        rules: list[dict] = []
        index = 0
        for room_type in room_types:
            for board_type in board_types:
                rate_value = rates[index % len(rates)] if rates else 120.0 + 15 * (index % 4)
                rules.append(
                    {
                        "contract_id": contract["id"],
                        "hotel_id": contract.get("hotel_id"),
                        "name": f"{room_type} / {board_type} base rate",
                        "rule_type": "base_rate",
                        "expression": f"expected = {rate_value:.2f} * nights",
                        "priority": 100 + index,
                        "is_active": True,
                        "metadata": {
                            "room_type": room_type,
                            "board_type": board_type,
                            "base_rate": float(round(rate_value, 2)),
                        },
                        "created_at": now,
                        "updated_at": now,
                    }
                )
                index += 1

        child_discount_percent = self._extract_child_discount_percent(extraction)
        rules.append(
            {
                "contract_id": contract["id"],
                "hotel_id": contract.get("hotel_id"),
                "name": "2nd child discount",
                "rule_type": "child_discount",
                "expression": f"if pax_children >= 2 then discount {child_discount_percent}% on expected",
                "priority": 20,
                "is_active": True,
                "metadata": {
                    "min_children": 2,
                    "discount_percent": child_discount_percent,
                },
                "created_at": now,
                "updated_at": now,
            }
        )

        for promo in promotions:
            discount = float(promo.get("discount_percent") or 0)
            if discount <= 0:
                continue
            rules.append(
                {
                    "contract_id": contract["id"],
                    "hotel_id": contract.get("hotel_id"),
                    "name": promo.get("offer_name", "Promotion"),
                    "rule_type": "promotion",
                    "expression": f"if promotion active then discount {discount}%",
                    "priority": 40,
                    "is_active": True,
                    "metadata": {
                        "promotion_id": promo.get("id"),
                        "offer_name": promo.get("offer_name", "Promotion"),
                        "discount_percent": discount,
                        "start_date": promo.get("start_date"),
                        "end_date": promo.get("end_date"),
                        "booking_start_date": promo.get("booking_start_date"),
                        "booking_end_date": promo.get("booking_end_date"),
                        "arrival_start_date": promo.get("arrival_start_date"),
                        "arrival_end_date": promo.get("arrival_end_date"),
                        "scope": promo.get("scope", "all"),
                        "non_cumulative": bool(promo.get("non_cumulative", False)),
                        "applicable_room_types": promo.get("applicable_room_types", []),
                        "applicable_board_types": promo.get("applicable_board_types", []),
                    },
                    "created_at": now,
                    "updated_at": now,
                }
            )

        return await self.repository.replace_rules(contract_id=contract["id"], rules=rules)

    async def list_rules(self, contract_id: str | None = None) -> list[dict]:
        return await self.repository.list_rules(contract_id)

    async def sync_configuration(self, payload: SyncRequest, created_by_user_id: str) -> dict:
        contract = await self.get_contract(payload.contract_id)
        rules = await self.repository.list_rules(contract_id=contract["id"])
        if not rules:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No pricing rules found. Generate rules before sync.",
            )

        raw_promotions = await self.repository.list_promotions_for_contract(
            operator_code=contract["operator_code"],
            hotel_code=contract["hotel_code"],
            hotel_id=contract.get("hotel_id"),
            contract_id=contract["id"],
        )
        promotions = self._collect_effective_promotions(promotions=raw_promotions, rules=rules)

        now = utcnow()
        sync_payload = {
            "contract": {
                "id": contract["id"],
                "hotel_id": contract.get("hotel_id"),
                "operator_code": contract["operator_code"],
                "hotel_code": contract["hotel_code"],
                "season_label": contract.get("season_label"),
            },
            "rules": [
                {
                    "name": item["name"],
                    "type": item["rule_type"],
                    "expression": item["expression"],
                    "metadata": item.get("metadata", {}),
                }
                for item in rules
                if item.get("is_active", True)
            ],
            "promotions": [
                {
                    "offer_name": item.get("offer_name"),
                    "discount_percent": item.get("discount_percent"),
                    "start_date": item.get("start_date"),
                    "end_date": item.get("end_date"),
                }
                for item in promotions
            ],
        }

        status_value = "dry_run" if payload.dry_run else "synced"
        details = {
            "mode": "simulated",
            "target": payload.target_system,
            "message": "Configuration prepared for Fidelio/third-party integration adapter.",
            "records_sent": len(sync_payload["rules"]),
        }

        run = {
            "contract_id": contract["id"],
            "hotel_id": contract.get("hotel_id"),
            "target_system": payload.target_system,
            "status": status_value,
            "payload": sync_payload,
            "details": details,
            "created_by_user_id": created_by_user_id,
            "created_at": now,
            "updated_at": now,
        }
        return await self.repository.create_sync_run(run)

    async def list_sync_runs(self, contract_id: str | None = None, hotel_id: str | None = None) -> list[dict]:
        return await self.repository.list_sync_runs(contract_id=contract_id, hotel_id=hotel_id)

    async def persist_reconciliation_import(self, payload: ReconciliationImportPersistRequest, created_by_user_id: str) -> dict:
        contract = await self.get_contract(payload.contract_id)
        now = utcnow()
        normalized_source_system = str(payload.source_system or "").strip().lower() or None
        normalized_reservation_id_column = str(payload.reservation_id_column or "").strip() or None

        reservations_by_unique_key: dict[str, dict] = {}
        for index, line in enumerate(payload.lines, start=1):
            line_data = line.model_dump()
            reservation_id = str(line_data.get("reservation_id") or "").strip() or f"AUTO-{payload.sheet_name}-{index}"
            room_type = str(line_data.get("room_type") or "").strip() or "Standard Room"
            board_type = self._normalize_board_type(str(line_data.get("board_type") or "").strip())
            booking_date = self._coerce_optional_date(line_data.get("booking_date"))
            stay_date = self._coerce_date(line_data.get("stay_date"))
            nights = max(1, min(60, self._coerce_int(line_data.get("nights"), default=1)))
            pax_adults = max(1, min(10, self._coerce_int(line_data.get("pax_adults"), default=2)))
            pax_children = max(0, min(10, self._coerce_int(line_data.get("pax_children"), default=0)))
            actual_price = self._coerce_float(line_data.get("actual_price"))
            if actual_price is None:
                continue
            contract_rate = self._coerce_float(line_data.get("contract_rate"))
            promo_code = str(line_data.get("promo_code") or "").strip() or None
            unique_key = self._normalize_reservation_unique_key(reservation_id, payload.sheet_name, index)
            group_key = self._normalize_reservation_group_key(reservation_id, payload.sheet_name, index)

            # Keep the last occurrence when duplicate IDs exist in one upload.
            reservations_by_unique_key[unique_key] = {
                "contract_id": contract["id"],
                "hotel_id": contract.get("hotel_id"),
                "hotel_code": contract["hotel_code"],
                "operator_code": contract["operator_code"],
                "file_name": payload.file_name,
                "sheet_name": payload.sheet_name,
                "source_system": normalized_source_system,
                "reservation_id_column": normalized_reservation_id_column,
                "reservation_id": reservation_id,
                "reservation_group_key": group_key,
                "reservation_unique_key": unique_key,
                "room_type": self._to_title_case(room_type),
                "board_type": board_type,
                "booking_date": booking_date,
                "stay_date": stay_date,
                "nights": nights,
                "pax_adults": pax_adults,
                "pax_children": pax_children,
                "actual_price": round(actual_price, 2),
                "contract_rate": round(contract_rate, 2) if contract_rate is not None else None,
                "promo_code": promo_code,
                "created_at": now,
                "updated_at": now,
            }

        reservations = list(reservations_by_unique_key.values())

        if not reservations:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No valid reservation lines were provided for persistence.",
            )

        import_doc = {
            "contract_id": contract["id"],
            "hotel_id": contract.get("hotel_id"),
            "hotel_code": contract["hotel_code"],
            "operator_code": contract["operator_code"],
            "file_name": payload.file_name,
            "sheet_name": payload.sheet_name,
            "source_system": normalized_source_system,
            "reservation_id_column": normalized_reservation_id_column,
            "mapping_summary": payload.mapping_summary,
            "analysis_provider": payload.analysis_provider,
            "analysis_model": payload.analysis_model,
            "analysis_usage": payload.analysis_usage,
            "ingestion_mode": "v1_replace",
            "line_count": len(reservations),
            "created_by_user_id": created_by_user_id,
            "created_at": now,
        }

        return await self.repository.create_reconciliation_import(import_doc, reservations)

    async def persist_reconciliation_import_v2(self, payload: ReconciliationImportPersistRequest, created_by_user_id: str) -> dict:
        contract = await self.get_contract(payload.contract_id)
        now = utcnow()
        normalized_source_system = str(payload.source_system or "").strip().lower() or None
        normalized_reservation_id_column = str(payload.reservation_id_column or "").strip() or None

        reservations: list[dict] = []
        for index, line in enumerate(payload.lines, start=1):
            line_data = line.model_dump()
            reservation_id = str(line_data.get("reservation_id") or "").strip() or f"AUTO-{payload.sheet_name}-{index}"
            room_type = str(line_data.get("room_type") or "").strip() or "Standard Room"
            board_type = self._normalize_board_type(str(line_data.get("board_type") or "").strip())
            booking_date = self._coerce_optional_date(line_data.get("booking_date"))
            stay_date = self._coerce_date(line_data.get("stay_date"))
            nights = max(1, min(60, self._coerce_int(line_data.get("nights"), default=1)))
            pax_adults = max(1, min(10, self._coerce_int(line_data.get("pax_adults"), default=2)))
            pax_children = max(0, min(10, self._coerce_int(line_data.get("pax_children"), default=0)))
            actual_price = self._coerce_float(line_data.get("actual_price"))
            if actual_price is None:
                continue

            contract_rate = self._coerce_float(line_data.get("contract_rate"))
            promo_code = str(line_data.get("promo_code") or "").strip() or None
            unique_key = self._normalize_reservation_unique_key(reservation_id, payload.sheet_name, index)
            group_key = self._normalize_reservation_group_key(reservation_id, payload.sheet_name, index)

            reservations.append(
                {
                    "contract_id": contract["id"],
                    "hotel_id": contract.get("hotel_id"),
                    "hotel_code": contract["hotel_code"],
                    "operator_code": contract["operator_code"],
                    "file_name": payload.file_name,
                    "sheet_name": payload.sheet_name,
                    "source_system": normalized_source_system,
                    "reservation_id_column": normalized_reservation_id_column,
                    "reservation_id": reservation_id,
                    "reservation_group_key": group_key,
                    "reservation_unique_key": unique_key,
                    "room_type": self._to_title_case(room_type),
                    "board_type": board_type,
                    "booking_date": booking_date,
                    "stay_date": stay_date,
                    "nights": nights,
                    "pax_adults": pax_adults,
                    "pax_children": pax_children,
                    "actual_price": round(actual_price, 2),
                    "contract_rate": round(contract_rate, 2) if contract_rate is not None else None,
                    "promo_code": promo_code,
                    "created_at": now,
                    "updated_at": now,
                }
            )

        if not reservations:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No valid reservation lines were provided for persistence.",
            )

        import_doc = {
            "contract_id": contract["id"],
            "hotel_id": contract.get("hotel_id"),
            "hotel_code": contract["hotel_code"],
            "operator_code": contract["operator_code"],
            "file_name": payload.file_name,
            "sheet_name": payload.sheet_name,
            "source_system": normalized_source_system,
            "reservation_id_column": normalized_reservation_id_column,
            "mapping_summary": payload.mapping_summary,
            "analysis_provider": payload.analysis_provider,
            "analysis_model": payload.analysis_model,
            "analysis_usage": payload.analysis_usage,
            "ingestion_mode": "v2_append",
            "line_count": len(reservations),
            "created_by_user_id": created_by_user_id,
            "created_at": now,
        }

        return await self.repository.create_reconciliation_import_append(import_doc, reservations)

    async def list_reconciliation_imports(
        self,
        *,
        contract_id: str | None = None,
        hotel_id: str | None = None,
        source_system: str | None = None,
        limit: int = 500,
    ) -> list[dict]:
        normalized_source = str(source_system or "").strip().lower() or None
        return await self.repository.list_reconciliation_imports(
            contract_id=contract_id,
            hotel_id=hotel_id,
            source_system=normalized_source,
            limit=limit,
        )

    async def list_reconciliation_reservations(
        self,
        *,
        contract_id: str | None = None,
        hotel_id: str | None = None,
        import_id: str | None = None,
        room_type: str | None = None,
        board_type: str | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        source_system: str | None = None,
        search_text: str | None = None,
        sort_by: str = "stay_date",
        sort_order: str = "desc",
        limit: int = 500,
        offset: int = 0,
    ) -> dict:
        return await self.repository.list_reconciliation_reservations(
            contract_id=contract_id,
            hotel_id=hotel_id,
            import_id=import_id,
            room_type=room_type,
            board_type=board_type,
            start_date=start_date,
            end_date=end_date,
            source_system=source_system,
            search_text=search_text,
            sort_by=sort_by,
            sort_order=sort_order,
            limit=limit,
            offset=offset,
        )

    async def delete_reconciliation_reservation(self, reservation_row_id: str) -> dict:
        deleted = await self.repository.delete_reconciliation_reservation(reservation_row_id=reservation_row_id)
        if not deleted or not deleted.get("deleted"):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Persisted reconciliation reservation not found.",
            )
        return deleted

    async def delete_reconciliation_reservations(
        self,
        *,
        contract_id: str,
        hotel_id: str | None = None,
        import_id: str | None = None,
        room_type: str | None = None,
        board_type: str | None = None,
        source_system: str | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        search_text: str | None = None,
    ) -> dict:
        await self.get_contract(contract_id)
        return await self.repository.delete_reconciliation_reservations(
            contract_id=contract_id,
            hotel_id=hotel_id,
            import_id=import_id,
            room_type=room_type,
            board_type=board_type,
            source_system=source_system,
            start_date=start_date,
            end_date=end_date,
            search_text=search_text,
        )

    async def preview_reconciliation_upload(
        self,
        upload_file: UploadFile,
        contract_id: str,
        sheet_name: str | None = None,
        sample_limit: int = 150,
    ) -> dict:
        file_name = upload_file.filename or "reconciliation-upload"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_RECONCILIATION_EXTENSIONS,
            max_bytes=upload_limits["reconciliation"],
            label="reconciliation",
            empty_message="Reconciliation file is empty.",
        )
        lines = await self._extract_reconciliation_lines(
            file_name=file_name,
            content=content,
            contract_id=contract_id,
            sheet_name=sheet_name,
            max_lines=max(1, min(sample_limit, 1000)),
        )
        return {
            "file_name": file_name,
            "contract_id": contract_id,
            "line_count": len(lines),
            "lines": lines,
        }

    async def preview_reconciliation_workbook(
        self,
        upload_file: UploadFile,
        contract_id: str,
        sample_rows: int = 6,
    ) -> dict:
        file_name = upload_file.filename or "reconciliation-upload"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_RECONCILIATION_EXTENSIONS,
            max_bytes=upload_limits["reconciliation"],
            label="reconciliation",
            empty_message="Reconciliation file is empty.",
        )

        # Validate contract id and hotel/operator context before preview.
        await self.get_contract(contract_id)

        workbook = self._load_reconciliation_workbook(file_name=file_name, content=content)
        preview_rows = max(2, min(sample_rows, 12))
        sheet_previews: list[dict] = []
        try:
            for sheet in workbook.worksheets[:24]:
                sheet_previews.append(self._build_reconciliation_sheet_preview(sheet=sheet, sample_rows=preview_rows))
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        suggested_sheet_name: str | None = None
        if sheet_previews:
            ranked = sorted(
                sheet_previews,
                key=lambda item: (
                    float(item.get("confidence") or 0.0),
                    int(item.get("non_empty_rows") or 0),
                ),
                reverse=True,
            )
            suggested_sheet_name = str(ranked[0].get("sheet_name") or "")
            if not suggested_sheet_name:
                suggested_sheet_name = None

        return {
            "file_name": file_name,
            "contract_id": contract_id,
            "sheet_count": len(sheet_previews),
            "suggested_sheet_name": suggested_sheet_name,
            "sheets": sheet_previews,
        }

    async def ai_map_reconciliation_upload(
        self,
        upload_file: UploadFile,
        contract_id: str,
        sheet_name: str,
        source_system: str | None = None,
        reservation_id_column: str | None = None,
        sample_limit: int = 250,
        model: str | None = None,
        mapping_instructions: str | None = None,
    ) -> dict:
        file_name = upload_file.filename or "reconciliation-upload"
        content = await upload_file.read()
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_RECONCILIATION_EXTENSIONS,
            max_bytes=upload_limits["reconciliation"],
            label="reconciliation",
            empty_message="Reconciliation file is empty.",
        )
        if not sheet_name or not sheet_name.strip():
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="sheet_name is required for AI reconciliation mapping.",
            )

        contract = await self.get_contract(contract_id)
        rules = await self.repository.list_rules(contract_id=contract["id"])
        raw_promotions = await self.repository.list_promotions_for_contract(
            operator_code=contract["operator_code"],
            hotel_code=contract["hotel_code"],
            hotel_id=contract.get("hotel_id"),
            contract_id=contract["id"],
        )
        promotions = self._collect_effective_promotions(promotions=raw_promotions, rules=rules)

        ai_payload = self._build_reconciliation_ai_payload(
            file_name=file_name,
            content=content,
            sheet_name=sheet_name.strip(),
            max_rows=max(25, min(sample_limit, 1200)),
        )
        if not ai_payload["rows"]:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Selected sheet did not contain non-empty table rows.",
            )

        resolved_reservation_id_column = self._resolve_reconciliation_id_column_key(
            ai_payload=ai_payload,
            selected_column=reservation_id_column,
        )
        reservation_id_values = self._build_reservation_id_value_lookup(
            ai_payload=ai_payload,
            reservation_id_column_key=resolved_reservation_id_column,
        )
        effective_mapping_instructions = self._build_effective_reconciliation_mapping_instructions(mapping_instructions)
        ai_sample_payload = dict(ai_payload)
        ai_sample_payload["rows"] = list(ai_payload.get("rows") or [])[:3]

        ai_result, usage, selected_model = await self._run_openai_reconciliation_mapping(
            model=model,
            source_system=source_system,
            contract=contract,
            rules=rules,
            promotions=promotions,
            sheet_payload=ai_sample_payload,
            reservation_id_column_key=resolved_reservation_id_column,
            mapping_instructions=effective_mapping_instructions,
        )
        resolved_header_mapping = self._resolve_reconciliation_header_mapping(
            ai_payload=ai_payload,
            raw_header_mapping=ai_result.get("header_mapping"),
            reservation_id_column_key=resolved_reservation_id_column,
        )
        mapped_lines = self._build_reconciliation_lines_from_header_mapping(
            sheet_payload=ai_payload,
            resolved_header_mapping=resolved_header_mapping,
            contract_id=contract["id"],
            default_hotel=contract["hotel_code"],
            default_operator=contract["operator_code"],
            sheet_title=sheet_name.strip(),
            max_lines=max(1, min(sample_limit, 1200)),
            reservation_id_values=reservation_id_values,
        )
        if not mapped_lines:
            mapped_lines = self._normalize_ai_reconciliation_lines(
                ai_lines=ai_result.get("lines"),
                contract_id=contract["id"],
                default_hotel=contract["hotel_code"],
                default_operator=contract["operator_code"],
                sheet_title=sheet_name.strip(),
                max_lines=max(1, min(sample_limit, 1200)),
                reservation_id_values=reservation_id_values,
            )
        if not mapped_lines:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="AI mapping returned no valid reconciliation rows. Try another sheet or update mapping instructions.",
            )

        mapping_summary = str(ai_result.get("mapping_summary") or "").strip() or "AI mapping completed."
        mapping_summary += f" Programmatic mapping applied to {len(mapped_lines)} rows."
        column_keys = [str(item).strip() for item in ai_payload.get("columns", []) if str(item).strip()]
        column_labels_raw = ai_payload.get("column_labels")
        column_labels = (
            [str(item).strip() for item in column_labels_raw]
            if isinstance(column_labels_raw, list)
            else []
        )
        column_label_by_key = {
            key: (column_labels[index] if index < len(column_labels) and column_labels[index] else key)
            for index, key in enumerate(column_keys)
        }
        header_mapping: dict[str, str] = {}
        for target_field, source_column in resolved_header_mapping.items():
            header_mapping[target_field] = column_label_by_key.get(source_column, source_column)

        normalized_source_system = str(source_system or "").strip() or None
        ai_source_system = ai_result.get("source_system")
        if not normalized_source_system and isinstance(ai_source_system, str) and ai_source_system.strip():
            normalized_source_system = ai_source_system.strip()

        return {
            "file_name": file_name,
            "contract_id": contract_id,
            "sheet_name": sheet_name.strip(),
            "source_system": normalized_source_system,
            "reservation_id_column": resolved_reservation_id_column,
            "analysis_provider": "openai",
            "analysis_model": selected_model,
            "analysis_usage": usage,
            "mapping_summary": mapping_summary,
            "header_mapping": header_mapping,
            "line_count": len(mapped_lines),
            "lines": mapped_lines,
        }

    async def validate_reconciliation_upload(
        self,
        upload_file: UploadFile,
        contract_id: str,
        sheet_name: str | None,
        source_system: str | None,
        reservation_id_column: str | None,
        use_ai_mapping: bool,
        model: str | None,
        mapping_instructions: str | None,
        run_label: str | None,
        tolerance_amount: float,
        tolerance_percent: float,
        created_by_user_id: str,
    ) -> dict:
        file_name = upload_file.filename or "reconciliation-upload"
        lines: list[dict]
        if use_ai_mapping:
            mapped = await self.ai_map_reconciliation_upload(
                upload_file=upload_file,
                contract_id=contract_id,
                sheet_name=(sheet_name or "").strip(),
                source_system=source_system,
                reservation_id_column=reservation_id_column,
                sample_limit=1000,
                model=model,
                mapping_instructions=mapping_instructions,
            )
            lines = mapped["lines"]
        else:
            content = await upload_file.read()
            lines = await self._extract_reconciliation_lines(
                file_name=file_name,
                content=content,
                contract_id=contract_id,
                sheet_name=sheet_name,
                max_lines=1000,
            )
        if not lines:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No reservation lines were extracted from the reconciliation file.",
            )

        payload = ValidationBatchRequest(
            contract_id=contract_id,
            run_label=run_label or f"Reconciliation import · {file_name}",
            tolerance_amount=tolerance_amount,
            tolerance_percent=tolerance_percent,
            lines=lines,
        )
        return await self.validate_batch(payload=payload, created_by_user_id=created_by_user_id)

    async def validate_batch(self, payload: ValidationBatchRequest, created_by_user_id: str) -> dict:
        contract = await self.get_contract(payload.contract_id)
        rules = await self.repository.list_rules(contract_id=contract["id"])
        raw_promotions = await self.repository.list_promotions_for_contract(
            operator_code=contract["operator_code"],
            hotel_code=contract["hotel_code"],
            hotel_id=contract.get("hotel_id"),
            contract_id=contract["id"],
        )
        promotions = self._collect_effective_promotions(promotions=raw_promotions, rules=rules)

        if not rules:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No rules available for validation. Generate rules first.",
            )

        base_rate_lookup, child_discount_percent = self._build_rule_lookup(rules)
        pricing_context = self._build_validation_pricing_context(
            contract=contract,
            rules=rules,
            base_rate_lookup=base_rate_lookup,
        )

        line_results: list[dict] = []
        mismatch_count = 0
        match_count = 0
        now = utcnow()

        for line in payload.lines:
            result = self._validate_single_line(
                line_data=line.model_dump(),
                base_rate_lookup=base_rate_lookup,
                pricing_context=pricing_context,
                child_discount_percent=child_discount_percent,
                promotions=promotions,
                tolerance_amount=payload.tolerance_amount,
                tolerance_percent=payload.tolerance_percent,
            )
            line_results.append(result)
            if result["status"] == "mismatch":
                mismatch_count += 1
            else:
                match_count += 1

        total_lines = len(line_results)
        mismatch_rate = (mismatch_count / total_lines * 100.0) if total_lines else 0.0

        validation_run = {
            "contract_id": contract["id"],
            "hotel_id": contract.get("hotel_id"),
            "run_label": payload.run_label,
            "total_lines": total_lines,
            "match_count": match_count,
            "mismatch_count": mismatch_count,
            "mismatch_rate": round(mismatch_rate, 2),
            "created_by_user_id": created_by_user_id,
            "created_at": now,
            "results": line_results,
        }
        created_run = await self.repository.create_validation_run(validation_run)

        for row in line_results:
            if row["status"] == "match":
                continue
            await self.repository.create_alert(
                {
                    "validation_run_id": created_run["id"],
                    "contract_id": contract["id"],
                    "hotel_id": contract.get("hotel_id"),
                    "reservation_id": row["reservation_id"],
                    "hotel_code": row["hotel_code"],
                    "operator_code": row["operator_code"],
                    "severity": "high" if abs(float(row["variance_percent"])) >= 10 else "medium",
                    "status": "open",
                    "message": row["reason"],
                    "details": row,
                    "resolution_note": None,
                    "created_at": now,
                    "resolved_at": None,
                    "resolved_by_user_id": None,
                }
            )

        return created_run

    async def list_alerts(self, status_value: str | None = None, hotel_id: str | None = None) -> list[dict]:
        return await self.repository.list_alerts(status=status_value, hotel_id=hotel_id)

    async def resolve_alert(self, alert_id: str, payload: AlertResolveRequest, resolved_by_user_id: str) -> dict:
        updated = await self.repository.resolve_alert(
            alert_id=alert_id,
            resolved_at=utcnow(),
            resolved_by_user_id=resolved_by_user_id,
            resolution_note=payload.resolution_note,
        )
        if not updated:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alert not found")
        return updated

    async def get_overview(self, hotel_id: str | None = None) -> dict:
        return await self.repository.stats_overview(hotel_id=hotel_id)

    async def discrepancy_report(self, hotel_id: str | None = None) -> list[dict]:
        runs = await self.repository.list_validation_runs(limit=2000, hotel_id=hotel_id)
        grouped: dict[tuple[str, str], dict] = {}

        for run in runs:
            if hotel_id and run.get("hotel_id") != hotel_id:
                continue
            created_at = run.get("created_at")
            for row in run.get("results", []):
                key = (row.get("operator_code", "UNKNOWN"), row.get("hotel_code", "UNKNOWN"))
                item = grouped.setdefault(
                    key,
                    {
                        "operator_code": key[0],
                        "hotel_code": key[1],
                        "total_validations": 0,
                        "mismatches": 0,
                        "total_variance_amount": 0.0,
                        "latest_mismatch_at": None,
                    },
                )
                item["total_validations"] += 1
                if row.get("status") == "mismatch":
                    item["mismatches"] += 1
                    item["total_variance_amount"] += float(row.get("variance_amount") or 0)
                    if created_at and (
                        item["latest_mismatch_at"] is None or created_at > item["latest_mismatch_at"]
                    ):
                        item["latest_mismatch_at"] = created_at

        output: list[dict] = []
        for item in grouped.values():
            total = int(item["total_validations"])
            mismatches = int(item["mismatches"])
            output.append(
                {
                    **item,
                    "mismatch_rate": round((mismatches / total * 100.0), 2) if total else 0.0,
                    "total_variance_amount": round(float(item["total_variance_amount"]), 2),
                }
            )
        output.sort(key=lambda row: (row["mismatches"], row["total_variance_amount"]), reverse=True)
        return output

    async def contract_performance_report(self, hotel_id: str | None = None) -> list[dict]:
        runs = await self.repository.list_validation_runs(limit=2000, hotel_id=hotel_id)
        contracts = await self.repository.list_contracts(hotel_id=hotel_id)
        contract_map = {item["id"]: item for item in contracts}

        grouped: dict[str, dict] = {}
        for run in runs:
            if hotel_id and run.get("hotel_id") != hotel_id:
                continue
            contract_id = run.get("contract_id")
            if not contract_id:
                continue
            contract_info = contract_map.get(contract_id, {})
            item = grouped.setdefault(
                contract_id,
                {
                    "contract_id": contract_id,
                    "operator_code": contract_info.get("operator_code", "UNKNOWN"),
                    "hotel_code": contract_info.get("hotel_code", "UNKNOWN"),
                    "total_runs": 0,
                    "total_lines": 0,
                    "total_mismatches": 0,
                    "last_run_at": None,
                },
            )
            item["total_runs"] += 1
            item["total_lines"] += int(run.get("total_lines", 0))
            item["total_mismatches"] += int(run.get("mismatch_count", 0))
            created_at = run.get("created_at")
            if created_at and (item["last_run_at"] is None or created_at > item["last_run_at"]):
                item["last_run_at"] = created_at

        output: list[dict] = []
        for item in grouped.values():
            total_lines = int(item["total_lines"])
            output.append(
                {
                    "contract_id": item["contract_id"],
                    "operator_code": item["operator_code"],
                    "hotel_code": item["hotel_code"],
                    "total_runs": int(item["total_runs"]),
                    "total_lines": total_lines,
                    "mismatch_rate": round((item["total_mismatches"] / total_lines * 100.0), 2) if total_lines else 0.0,
                    "last_run_at": item["last_run_at"],
                }
            )
        output.sort(key=lambda row: row["mismatch_rate"], reverse=True)
        return output

    async def operator_report(self, hotel_id: str | None = None) -> list[dict]:
        contract_rows = await self.contract_performance_report(hotel_id=hotel_id)
        discrepancy_rows = await self.discrepancy_report(hotel_id=hotel_id)

        grouped: dict[str, dict] = defaultdict(
            lambda: {
                "operator_code": "",
                "contract_ids": set(),
                "total_runs": 0,
                "total_mismatches": 0,
                "total_variance_amount": 0.0,
            }
        )

        for row in contract_rows:
            operator = row["operator_code"]
            grouped[operator]["operator_code"] = operator
            grouped[operator]["contract_ids"].add(row["contract_id"])
            grouped[operator]["total_runs"] += int(row["total_runs"])
            grouped[operator]["total_mismatches"] += int(round(row["mismatch_rate"] * row["total_lines"] / 100.0))

        for row in discrepancy_rows:
            operator = row["operator_code"]
            grouped[operator]["operator_code"] = operator
            grouped[operator]["total_variance_amount"] += float(row["total_variance_amount"])

        output: list[dict] = []
        for item in grouped.values():
            output.append(
                {
                    "operator_code": item["operator_code"],
                    "total_contracts": len(item["contract_ids"]),
                    "total_runs": int(item["total_runs"]),
                    "total_mismatches": int(item["total_mismatches"]),
                    "total_variance_amount": round(float(item["total_variance_amount"]), 2),
                }
            )
        output.sort(key=lambda row: (row["total_mismatches"], row["total_variance_amount"]), reverse=True)
        return output

    def _extract_text_from_bytes(self, file_name: str, content: bytes) -> str:
        suffix = Path(file_name).suffix.lower()
        if suffix == ".pdf":
            return self._extract_text_from_pdf(content)
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            return self._extract_text_from_excel(content)
        if suffix == ".docx":
            return self._extract_text_from_docx(content)
        if suffix in {".eml", ".msg"}:
            return self._extract_text_from_email(content)

        text = content.decode("utf-8", errors="ignore").strip()
        if not text:
            text = content.decode("latin-1", errors="ignore")
        return text

    def _extract_text_from_pdf(self, content: bytes) -> str:
        try:
            from pypdf import PdfReader

            reader = PdfReader(BytesIO(content), strict=True)
            pages: list[str] = []
            for page in reader.pages:
                pages.append(page.extract_text() or "")
            return "\n".join(pages)
        except Exception:
            return ""

    def _extract_text_from_excel(self, content: bytes) -> str:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True, keep_vba=True)
            lines: list[str] = []
            for sheet in workbook.worksheets[:6]:
                lines.append(f"[SHEET] {sheet.title}")
                for row in sheet.iter_rows(min_row=1, max_row=220, values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        lines.append(" | ".join(values))
            return "\n".join(lines)
        except Exception:
            return ""

    def _extract_text_from_docx(self, content: bytes) -> str:
        try:
            with ZipFile(BytesIO(content)) as archive:
                document_xml = archive.read("word/document.xml")
            tree = ElementTree.fromstring(document_xml)
            namespaces = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            chunks = [node.text for node in tree.findall(".//w:t", namespaces) if node.text]
            return " ".join(chunks)
        except Exception:
            return ""

    def _extract_text_from_email(self, content: bytes) -> str:
        try:
            message = BytesParser(policy=policy.default).parsebytes(content)
            parts: list[str] = []
            subject = message.get("subject")
            if subject:
                parts.append(str(subject))

            if message.is_multipart():
                for piece in message.walk():
                    if piece.get_content_maintype() == "multipart":
                        continue
                    payload = piece.get_payload(decode=True) or b""
                    charset = piece.get_content_charset() or "utf-8"
                    parts.append(payload.decode(charset, errors="ignore"))
            else:
                payload = message.get_payload(decode=True) or b""
                charset = message.get_content_charset() or "utf-8"
                parts.append(payload.decode(charset, errors="ignore"))

            return "\n".join(part.strip() for part in parts if part and part.strip())
        except Exception:
            return content.decode("utf-8", errors="ignore")

    def _extract_contract_terms(self, text: str) -> dict:
        normalized_text = text or ""
        lower = normalized_text.lower()

        room_types = set()
        for pattern in ROOM_TYPE_PATTERNS:
            for match in re.findall(pattern, lower, flags=re.IGNORECASE):
                room_types.add(self._to_title_case(match))

        if not room_types:
            for line in normalized_text.splitlines():
                if "room" in line.lower() and len(line) <= 120:
                    room_types.add(self._to_title_case(line.strip()))
                if len(room_types) >= 8:
                    break

        board_types = {
            code
            for key, code in BOARD_TYPE_LOOKUP.items()
            if re.search(rf"\b{re.escape(key)}\b", lower, flags=re.IGNORECASE)
        }

        seasonal_periods = set(
            re.findall(
                r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s*(?:to|-|–)\s*\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b",
                normalized_text,
                flags=re.IGNORECASE,
            )
        )
        seasonal_periods.update(
            re.findall(
                r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+\d{4}\s*(?:to|-|–)\s*(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+\d{4}\b",
                lower,
                flags=re.IGNORECASE,
            )
        )

        discounts = self._extract_keyword_lines(normalized_text, keywords=["discount", "%", "child", "free night"])[:20]
        supplements = self._extract_keyword_lines(normalized_text, keywords=["supplement", "single", "weekend surcharge"])[:20]
        marketing_contributions = self._extract_keyword_lines(
            normalized_text,
            keywords=["marketing", "contribution", "co-op", "cooperative", "campaign"],
        )[:20]

        raw_highlights = self._extract_keyword_lines(
            normalized_text,
            keywords=["offer", "promotion", "room", "board", "valid", "contract"],
        )[:25]

        return {
            "room_types": sorted(room_types)[:25],
            "seasonal_periods": sorted({item.strip() for item in seasonal_periods if item.strip()})[:25],
            "board_types": sorted(board_types)[:10],
            "discounts": discounts,
            "supplements": supplements,
            "marketing_contributions": marketing_contributions,
            "raw_highlights": raw_highlights,
        }

    def _extract_promotion_offer(self, text: str, fallback_name: str) -> dict:
        lower = text.lower()

        offer_name = "General Promotion"
        for keyword, normalized in PROMOTION_KEYWORDS.items():
            if keyword in lower:
                offer_name = normalized
                break

        percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", text)
        discount_percent = float(percent_match.group(1)) if percent_match else None

        date_values = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", text)
        start_date = self._parse_date_string(date_values[0]) if len(date_values) >= 1 else None
        end_date = self._parse_date_string(date_values[1]) if len(date_values) >= 2 else None
        booking_start_date: date | None = None
        booking_end_date: date | None = None
        arrival_start_date: date | None = None
        arrival_end_date: date | None = None

        booking_match = re.search(
            r"(?:booking\s+dates?|bookings?\s+taken|booking\s+period)[^0-9]{0,25}"
            r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*(?:-|–|to|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            text,
            flags=re.IGNORECASE,
        )
        if booking_match:
            booking_start_date = self._parse_date_string(booking_match.group(1))
            booking_end_date = self._parse_date_string(booking_match.group(2))

        arrival_match = re.search(
            r"(?:arrival\s+dates?|travel\s+dates?|stay\s+dates?|arrival\s+window)[^0-9]{0,25}"
            r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*(?:-|–|to|until)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            text,
            flags=re.IGNORECASE,
        )
        if arrival_match:
            arrival_start_date = self._parse_date_string(arrival_match.group(1))
            arrival_end_date = self._parse_date_string(arrival_match.group(2))

        if not start_date and arrival_start_date:
            start_date = arrival_start_date
        if not end_date and arrival_end_date:
            end_date = arrival_end_date

        description_lines = self._extract_keyword_lines(
            text,
            keywords=["offer", "promotion", "discount", "book", "travel", "valid"],
        )
        description = " ".join(description_lines[:6])
        if not description:
            description = text.strip().replace("\n", " ")[:320]

        scope = "all"
        if "selected room" in lower or "selected rooms" in lower:
            scope = "selected_rooms"
        elif "selected board" in lower:
            scope = "selected_board_types"

        non_cumulative = bool(re.search(r"\bnon[-\s]?cumulative\b|\bnot\s+cumulative\b", lower))
        combinability_note = None
        sentence_match = re.search(r"[^.]*non[-\s]?cumulative[^.]*\.", text, flags=re.IGNORECASE)
        if sentence_match:
            combinability_note = sentence_match.group(0).strip()

        has_window_conjunction = bool(
            (booking_start_date or booking_end_date)
            and (arrival_start_date or arrival_end_date)
        )
        is_early_booking = has_window_conjunction or self._is_early_booking_offer(
            offer_name=offer_name,
            description=description,
        )
        if is_early_booking:
            if not (booking_start_date or booking_end_date) and (arrival_start_date or arrival_end_date):
                booking_start_date = arrival_start_date
                booking_end_date = arrival_end_date
            if not (arrival_start_date or arrival_end_date) and (booking_start_date or booking_end_date):
                arrival_start_date = booking_start_date
                arrival_end_date = booking_end_date

        applicable_room_types = sorted(
            {
                self._to_title_case(match)
                for pattern in ROOM_TYPE_PATTERNS
                for match in re.findall(pattern, lower, flags=re.IGNORECASE)
            }
        )
        applicable_board_types = sorted(
            {
                code
                for key, code in BOARD_TYPE_LOOKUP.items()
                if re.search(rf"\b{re.escape(key)}\b", lower, flags=re.IGNORECASE)
            }
        )

        return {
            "offer_name": offer_name if offer_name else fallback_name,
            "description": description,
            "discount_percent": discount_percent,
            "start_date": start_date,
            "end_date": end_date,
            "booking_start_date": booking_start_date,
            "booking_end_date": booking_end_date,
            "arrival_start_date": arrival_start_date,
            "arrival_end_date": arrival_end_date,
            "non_cumulative": non_cumulative,
            "combinability_note": combinability_note,
            "applicable_room_types": applicable_room_types,
            "applicable_board_types": applicable_board_types,
            "scope": scope,
            "promotion_category": "early_booking" if is_early_booking else "general",
        }

    async def _run_openai_promotion_extraction(
        self,
        text: str,
        hotel_code: str,
        operator_code: str,
        fallback_offer: dict,
    ) -> tuple[dict, dict, str]:
        settings = get_settings()
        if not settings.openai_api_key:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="OPENAI_API_KEY is not configured.",
            )

        try:
            from openai import AsyncOpenAI
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="OpenAI SDK is not installed. Add `openai` to backend requirements.",
            ) from exc

        selected_model = settings.openai_pricing_model
        configured_base_url = (settings.openai_base_url or "").strip()
        client = AsyncOpenAI(
            api_key=settings.openai_api_key,
            base_url=configured_base_url or "https://api.openai.com/v1",
            timeout=None,
        )

        # fallback_offer can contain date objects from heuristic parsing.
        baseline = json.dumps(fallback_offer, ensure_ascii=False, default=str)[:6000]
        prompt = (
            f"Hotel code: {hotel_code}\n"
            f"Operator code: {operator_code}\n\n"
            "Extract structured promotion terms from this hotel contracting email or document.\n"
            "Prioritize these fields: discount_percent, booking_start_date/booking_end_date, "
            "arrival_start_date/arrival_end_date, non_cumulative, and scope.\n"
            "Promotions are additive overlays on top of contract base rates/rules (all other terms remain in force).\n"
            "For early-booking offers, BOTH booking and arrival windows are required; if only one is explicit, "
            "mirror it to both windows.\n"
            "If dates are ambiguous, still return best effort and set confidence accordingly.\n\n"
            f"Heuristic baseline:\n{baseline}\n\n"
            f"Promotion content:\n{text[:40000]}"
        )
        system_prompt = (
            "You are a senior hospitality commercial analyst. "
            "Return strict JSON only. "
            "These promotions are used to update hotel pricing matrices and promotion rules."
        )

        try:
            completion = await client.chat.completions.create(
                model=selected_model,
                temperature=0,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "promotion_extraction",
                        "schema": AI_PROMOTION_EXTRACTION_SCHEMA,
                    },
                },
            )
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider request failed.",
                event="openai.promotion_extraction.request_failed",
                exc=exc,
            )

        if not completion.choices:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI promotion extraction returned no choices.",
            )

        message = completion.choices[0].message
        refusal = getattr(message, "refusal", None)
        if refusal:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Upstream AI provider refused promotion extraction.",
                event="openai.promotion_extraction.refused",
                context={"refusal": str(refusal)[:400]},
            )

        content = message.content or ""
        if not isinstance(content, str) or not content.strip():
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI promotion extraction returned empty content.",
            )

        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider returned invalid JSON.",
                event="openai.promotion_extraction.invalid_json",
                exc=exc,
            )

        usage_obj = getattr(completion, "usage", None)
        usage = {}
        if usage_obj:
            usage = {
                "prompt_tokens": int(getattr(usage_obj, "prompt_tokens", 0) or 0),
                "completion_tokens": int(getattr(usage_obj, "completion_tokens", 0) or 0),
                "total_tokens": int(getattr(usage_obj, "total_tokens", 0) or 0),
            }
        return parsed if isinstance(parsed, dict) else {}, usage, selected_model

    def _merge_promotion_offer_data(self, ai_offer: dict, fallback_offer: dict, fallback_name: str) -> dict:
        def read_text(key: str, default: str = "") -> str:
            value = ai_offer.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
            fallback_value = fallback_offer.get(key)
            if isinstance(fallback_value, str) and fallback_value.strip():
                return fallback_value.strip()
            return default

        def read_date(*keys: str) -> date | None:
            for key in keys:
                ai_value = self._coerce_optional_date(ai_offer.get(key))
                if ai_value:
                    return ai_value
            for key in keys:
                fallback_value = self._coerce_optional_date(fallback_offer.get(key))
                if fallback_value:
                    return fallback_value
            return None

        discount = self._coerce_float(ai_offer.get("discount_percent"))
        if discount is None:
            discount = self._coerce_float(fallback_offer.get("discount_percent"))

        booking_start_date = read_date("booking_start_date")
        booking_end_date = read_date("booking_end_date")
        arrival_start_date = read_date("arrival_start_date", "start_date")
        arrival_end_date = read_date("arrival_end_date", "end_date")
        start_date = arrival_start_date or read_date("start_date")
        end_date = arrival_end_date or read_date("end_date")

        scope = read_text("scope", default=str(fallback_offer.get("scope") or "all"))
        if scope not in {"all", "selected_rooms", "selected_board_types"}:
            scope = "all"

        ai_room_types = self._coerce_mixed_items_to_text(ai_offer.get("applicable_room_types"))
        fallback_room_types = self._coerce_mixed_items_to_text(fallback_offer.get("applicable_room_types"))
        applicable_room_types = sorted({self._to_title_case(item) for item in [*ai_room_types, *fallback_room_types] if item})

        ai_board_types = self._coerce_mixed_items_to_text(ai_offer.get("applicable_board_types"))
        fallback_board_types = self._coerce_mixed_items_to_text(fallback_offer.get("applicable_board_types"))
        applicable_board_types = sorted(
            {
                self._normalize_board_type(item)
                for item in [*ai_board_types, *fallback_board_types]
                if item
            }
        )

        non_cumulative: bool
        if isinstance(ai_offer.get("non_cumulative"), bool):
            non_cumulative = bool(ai_offer.get("non_cumulative"))
        else:
            non_cumulative = bool(fallback_offer.get("non_cumulative", False))

        offer_name = read_text("offer_name", default=fallback_name)
        description = read_text("description", default=fallback_name)
        has_window_conjunction = bool(
            (booking_start_date or booking_end_date)
            and (arrival_start_date or arrival_end_date)
        )
        is_early_booking = has_window_conjunction or self._is_early_booking_offer(
            offer_name=offer_name,
            description=description,
            expression=f"{offer_name} {description}",
        )
        if is_early_booking:
            if not (booking_start_date or booking_end_date) and (arrival_start_date or arrival_end_date):
                booking_start_date = arrival_start_date
                booking_end_date = arrival_end_date
            if not (arrival_start_date or arrival_end_date) and (booking_start_date or booking_end_date):
                arrival_start_date = booking_start_date
                arrival_end_date = booking_end_date

        return {
            "offer_name": offer_name,
            "description": description,
            "discount_percent": round(discount, 2) if discount is not None and discount > 0 else None,
            "start_date": start_date,
            "end_date": end_date,
            "booking_start_date": booking_start_date,
            "booking_end_date": booking_end_date,
            "arrival_start_date": arrival_start_date,
            "arrival_end_date": arrival_end_date,
            "non_cumulative": non_cumulative,
            "combinability_note": read_text("combinability_note"),
            "applicable_room_types": applicable_room_types,
            "applicable_board_types": applicable_board_types,
            "scope": scope,
            "promotion_category": "early_booking" if is_early_booking else "general",
        }

    async def _run_openai_structured_extraction(
        self,
        text: str,
        model: str | None,
        schema: dict,
        hotel_code: str,
        operator_code: str,
        season_label: str | None,
        mapping_instructions: str | None,
    ) -> tuple[dict, dict, str]:
        settings = get_settings()
        if not settings.openai_api_key:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="OPENAI_API_KEY is not configured.",
            )

        try:
            from openai import AsyncOpenAI
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="OpenAI SDK is not installed. Add `openai` to backend requirements.",
            ) from exc

        selected_model = model.strip() if model and model.strip() else settings.openai_pricing_model
        configured_base_url = (settings.openai_base_url or "").strip()
        client_kwargs: dict[str, object] = {
            "api_key": settings.openai_api_key,
            # Always pass an explicit URL because OPENAI_BASE_URL can be present but blank in env files.
            "base_url": configured_base_url or "https://api.openai.com/v1",
            "timeout": None,
        }
        client = AsyncOpenAI(**client_kwargs)

        text_window = text[:120000]
        instructions = mapping_instructions.strip() if mapping_instructions else ""
        prompt = (
            f"Hotel code: {hotel_code.strip().upper()}\n"
            f"Operator code: {operator_code.strip().upper()}\n"
            f"Season label: {(season_label or '').strip()}\n\n"
            "Extract structured pricing contract data from the text below.\n"
            "This is a tour-operator commercial terms contract for hotel pricing.\n"
            "Focus on room types, board types, pricing periods, base prices, extra guest pricing rules, discounts, "
            "supplements, marketing contributions, and promotional offers.\n"
            "For extra guest pricing rules, capture child and adult logic explicitly.\n"
            "Examples to extract without omission: '2nd child 50% of adult', '3rd adult 30% of adult price'.\n"
            "When available, set guest_type (adult/child/infant), guest_position (2nd/3rd...), age_min, age_max, "
            "and percent_of_adult.\n"
            "Capture exact numeric percentages and pricing values whenever present.\n"
            "Available board types are HB, BB, FB.\n"
        )
        if instructions:
            prompt += f"\nAdditional mapping instructions:\n{instructions}\n"
        prompt += f"\nContract text:\n{text_window}"

        system_prompt = (
            "You are a hospitality pricing contract parser. "
            "The source documents are commercial term contracts from tour operators. "
            "Prefer precise, structured extraction for pricing tables and occupancy rules. "
            "Return only data that can be directly supported by the supplied text. "
            "If uncertain, leave the field empty instead of guessing."
        )

        try:
            completion = await client.chat.completions.create(
                model=selected_model,
                temperature=0,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "pricing_contract_extraction",
                        "schema": schema,
                    },
                },
            )
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider request failed.",
                event="openai.pricing_extraction.request_failed",
                exc=exc,
            )

        if not completion.choices:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI extraction returned no choices.",
            )

        message = completion.choices[0].message
        refusal = getattr(message, "refusal", None)
        if refusal:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Upstream AI provider refused pricing extraction.",
                event="openai.pricing_extraction.refused",
                context={"refusal": str(refusal)[:400]},
            )

        content = message.content or ""
        if not isinstance(content, str) or not content.strip():
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI extraction returned empty content.",
            )

        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider returned invalid JSON.",
                event="openai.pricing_extraction.invalid_json",
                exc=exc,
            )

        usage_obj = getattr(completion, "usage", None)
        usage = {}
        if usage_obj:
            usage = {
                "prompt_tokens": int(getattr(usage_obj, "prompt_tokens", 0) or 0),
                "completion_tokens": int(getattr(usage_obj, "completion_tokens", 0) or 0),
                "total_tokens": int(getattr(usage_obj, "total_tokens", 0) or 0),
            }

        return parsed if isinstance(parsed, dict) else {"data": parsed}, usage, selected_model

    async def _run_openai_content_recommendation(
        self,
        text: str,
        hotel_code: str,
        operator_code: str,
        season_label: str | None,
        baseline_recommendation: dict,
    ) -> tuple[dict, dict, str]:
        settings = get_settings()
        if not settings.openai_api_key:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="OPENAI_API_KEY is not configured.",
            )

        try:
            from openai import AsyncOpenAI
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="OpenAI SDK is not installed. Add `openai` to backend requirements.",
            ) from exc

        selected_model = settings.openai_pricing_model
        configured_base_url = (settings.openai_base_url or "").strip()
        client = AsyncOpenAI(
            api_key=settings.openai_api_key,
            base_url=configured_base_url or "https://api.openai.com/v1",
            timeout=None,
        )

        text_window = text[:120000]
        baseline_payload = json.dumps(baseline_recommendation, ensure_ascii=False)[:60000]
        prompt = (
            f"Hotel code: {hotel_code}\n"
            f"Operator code: {operator_code}\n"
            f"Season label: {(season_label or '').strip()}\n\n"
            "Analyze this tour-operator commercial contract and create content-to-database mapping guidance.\n"
            "Your output MUST include:\n"
            "1) recommended_data representing the best structured interpretation of content,\n"
            "2) suggested_schema for extraction payload structure,\n"
            "3) suggested_mapping_instructions for better extraction mapping,\n"
            "4) database_mapping rows that map source paths to persistence entities/fields,\n"
            "5) coverage_feedback explaining missing/weak areas.\n"
            "Use the heuristic baseline only as context, but improve it with your own analysis.\n\n"
            f"Heuristic baseline:\n{baseline_payload}\n\n"
            f"Contract text:\n{text_window}"
        )

        system_prompt = (
            "You are a senior hospitality data-modeling analyst. "
            "Return strict JSON for mapping review workflow. "
            "Focus on content semantics from contracts, especially pricing periods, room/board structures, "
            "occupancy/extra-guest rules, promotions, and persistence mapping quality."
        )

        try:
            completion = await client.chat.completions.create(
                model=selected_model,
                temperature=0,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "pricing_content_recommendation",
                        "schema": AI_CONTENT_RECOMMENDATION_SCHEMA,
                    },
                },
            )
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider request failed.",
                event="openai.content_recommendation.request_failed",
                exc=exc,
            )

        if not completion.choices:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI content recommendation returned no choices.",
            )

        message = completion.choices[0].message
        refusal = getattr(message, "refusal", None)
        if refusal:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Upstream AI provider refused content recommendation.",
                event="openai.content_recommendation.refused",
                context={"refusal": str(refusal)[:400]},
            )

        content = message.content or ""
        if not isinstance(content, str) or not content.strip():
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI content recommendation returned empty content.",
            )

        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider returned invalid JSON.",
                event="openai.content_recommendation.invalid_json",
                exc=exc,
            )

        usage_obj = getattr(completion, "usage", None)
        usage = {}
        if usage_obj:
            usage = {
                "prompt_tokens": int(getattr(usage_obj, "prompt_tokens", 0) or 0),
                "completion_tokens": int(getattr(usage_obj, "completion_tokens", 0) or 0),
                "total_tokens": int(getattr(usage_obj, "total_tokens", 0) or 0),
            }

        return parsed if isinstance(parsed, dict) else {}, usage, selected_model

    def _merge_content_recommendation(self, ai_recommendation: dict, fallback: dict) -> dict:
        def pick_text(key: str, default: str = "") -> str:
            ai_value = ai_recommendation.get(key)
            if isinstance(ai_value, str) and ai_value.strip():
                return ai_value.strip()
            fb_value = fallback.get(key)
            if isinstance(fb_value, str) and fb_value.strip():
                return fb_value.strip()
            return default

        confidence_value = str(ai_recommendation.get("confidence") or fallback.get("confidence") or "medium").strip().lower()
        confidence = confidence_value if confidence_value in {"low", "medium", "high"} else "medium"

        detected_signals = {}
        if isinstance(fallback.get("detected_signals"), dict):
            detected_signals.update(fallback["detected_signals"])
        ai_signals = ai_recommendation.get("detected_signals")
        if isinstance(ai_signals, dict):
            detected_signals.update(ai_signals)

        coverage_feedback = self._coerce_mixed_items_to_text(ai_recommendation.get("coverage_feedback"))
        if not coverage_feedback:
            coverage_feedback = self._coerce_mixed_items_to_text(fallback.get("coverage_feedback"))

        fallback_recommended_data = fallback.get("recommended_data") if isinstance(fallback.get("recommended_data"), dict) else {}
        recommended_data_candidate = ai_recommendation.get("recommended_data")
        if not isinstance(recommended_data_candidate, dict):
            recommended_data_candidate = fallback_recommended_data
        recommended_data = self._stabilize_extracted_data_shape(recommended_data_candidate)
        if fallback_recommended_data:
            recommended_data = self._merge_extracted_with_fallback(recommended_data, fallback_recommended_data)

        suggested_schema_candidate = ai_recommendation.get("suggested_schema")
        if not isinstance(suggested_schema_candidate, dict) or not suggested_schema_candidate:
            suggested_schema_candidate = fallback.get("suggested_schema") if isinstance(fallback.get("suggested_schema"), dict) else {}
        suggested_schema = self._build_effective_extraction_schema(suggested_schema_candidate)

        raw_mapping = ai_recommendation.get("database_mapping")
        if not isinstance(raw_mapping, list):
            raw_mapping = fallback.get("database_mapping") if isinstance(fallback.get("database_mapping"), list) else []
        database_mapping = self._normalize_database_mapping_rows(raw_mapping)
        if not database_mapping:
            database_mapping = self._normalize_database_mapping_rows(self._build_database_mapping_template())

        return {
            "content_summary": pick_text("content_summary", default="Content analysis completed."),
            "confidence": confidence,
            "detected_signals": detected_signals,
            "coverage_feedback": coverage_feedback,
            "recommended_data": recommended_data,
            "suggested_schema": suggested_schema,
            "suggested_schema_rationale": pick_text("suggested_schema_rationale", default="Schema recommendation generated."),
            "suggested_mapping_instructions": pick_text(
                "suggested_mapping_instructions",
                default=(
                    "Extract room_types, board_types, seasonal_pricing_periods, pricing_lines, extra_guest_rules, "
                    "discounts, supplements, marketing_contributions, and promotional_offers. "
                    "Available board types are HB, BB, FB."
                ),
            ),
            "database_mapping": database_mapping,
        }

    def _build_effective_mapping_instructions(self, mapping_instructions: str | None) -> str:
        user_instructions = str(mapping_instructions or "").strip()
        if user_instructions:
            return f"{user_instructions}\n\nNon-negotiable output contract:\n{CANONICAL_MAPPING_REQUIREMENTS}"
        return CANONICAL_MAPPING_REQUIREMENTS

    def _build_effective_reconciliation_mapping_instructions(self, mapping_instructions: str | None) -> str:
        user_instructions = str(mapping_instructions or "").strip()
        if user_instructions:
            return (
                f"{user_instructions}\n\n"
                "Non-negotiable reconciliation mapping contract:\n"
                f"{CANONICAL_RECONCILIATION_MAPPING_REQUIREMENTS}"
            )
        return CANONICAL_RECONCILIATION_MAPPING_REQUIREMENTS

    def _build_effective_extraction_schema(self, schema_override: dict | None) -> dict:
        base_schema = deepcopy(DEFAULT_AI_PRICING_SCHEMA)
        if not isinstance(schema_override, dict) or not schema_override:
            return self._enforce_extraction_schema_contract(base_schema)

        override_schema = deepcopy(schema_override)
        base_properties = base_schema.setdefault("properties", {})
        override_properties = override_schema.get("properties")
        if isinstance(override_properties, dict):
            for key, value in override_properties.items():
                if not isinstance(value, dict):
                    continue
                base_value = base_properties.get(key)
                if isinstance(base_value, dict):
                    base_properties[key] = self._merge_schema_nodes(base_value, value)
                else:
                    base_properties[key] = deepcopy(value)

        for key in ("title", "description"):
            override_value = override_schema.get(key)
            if isinstance(override_value, str) and override_value.strip():
                base_schema[key] = override_value.strip()

        override_required = override_schema.get("required")
        merged_required: list[str] = []
        if isinstance(override_required, list):
            for item in override_required:
                if isinstance(item, str) and item in base_properties and item not in merged_required:
                    merged_required.append(item)
        for item in CORE_EXTRACT_REQUIRED_KEYS:
            if item in base_properties and item not in merged_required:
                merged_required.append(item)
        base_schema["required"] = merged_required
        return self._enforce_extraction_schema_contract(base_schema)

    def _merge_schema_nodes(self, base_node: dict, override_node: dict) -> dict:
        merged = deepcopy(base_node)
        for key, value in override_node.items():
            if key in {"properties", "definitions", "$defs"} and isinstance(value, dict):
                target = merged.get(key)
                if not isinstance(target, dict):
                    target = {}
                for child_key, child_value in value.items():
                    if isinstance(child_value, dict) and isinstance(target.get(child_key), dict):
                        target[child_key] = self._merge_schema_nodes(target[child_key], child_value)
                    else:
                        target[child_key] = deepcopy(child_value)
                merged[key] = target
                continue

            if key == "items" and isinstance(value, dict):
                existing_items = merged.get("items")
                if isinstance(existing_items, dict):
                    merged["items"] = self._merge_schema_nodes(existing_items, value)
                else:
                    merged["items"] = deepcopy(value)
                continue

            if key == "required" and isinstance(value, list):
                existing_required = merged.get("required")
                normalized_required = [item for item in existing_required if isinstance(item, str)] if isinstance(existing_required, list) else []
                for item in value:
                    if isinstance(item, str) and item not in normalized_required:
                        normalized_required.append(item)
                merged["required"] = normalized_required
                continue

            merged[key] = deepcopy(value)
        return merged

    def _enforce_extraction_schema_contract(self, schema: dict) -> dict:
        normalized_schema = deepcopy(schema) if isinstance(schema, dict) else {}
        normalized_schema["type"] = "object"
        normalized_schema["additionalProperties"] = True

        properties = normalized_schema.get("properties")
        if not isinstance(properties, dict):
            properties = {}
            normalized_schema["properties"] = properties

        default_properties = DEFAULT_AI_PRICING_SCHEMA.get("properties", {})
        if isinstance(default_properties, dict):
            for key, value in default_properties.items():
                if key not in properties:
                    properties[key] = deepcopy(value)

        for key in CORE_EXTRACT_ARRAY_KEYS:
            field = properties.get(key)
            if not isinstance(field, dict) or str(field.get("type") or "").strip().lower() != "array":
                replacement = default_properties.get(key) if isinstance(default_properties, dict) else None
                if isinstance(replacement, dict):
                    properties[key] = deepcopy(replacement)
                else:
                    properties[key] = {"type": "array", "items": {"type": "object"}}

        required = normalized_schema.get("required")
        normalized_required = [item for item in required if isinstance(item, str) and item in properties] if isinstance(required, list) else []
        for key in CORE_EXTRACT_REQUIRED_KEYS:
            if key in properties and key not in normalized_required:
                normalized_required.append(key)
        normalized_schema["required"] = normalized_required
        return normalized_schema

    def _stabilize_extracted_data_shape(self, extracted_data: dict) -> dict:
        if not isinstance(extracted_data, dict):
            extracted_data = {}

        normalized = deepcopy(extracted_data)
        for target_key, aliases in EXTRACTION_ALIAS_MAP.items():
            value = normalized.get(target_key)
            if isinstance(value, list) and value:
                continue
            found = self._find_list_by_alias(normalized, aliases=aliases)
            if isinstance(found, list):
                normalized[target_key] = deepcopy(found)

        normalized_room_types = self._coerce_mixed_items_to_text(normalized.get("room_types"))
        normalized["room_types"] = sorted(set(normalized_room_types))[:80]

        board_candidates = self._coerce_mixed_items_to_text(normalized.get("board_types"))
        board_codes = [self._normalize_board_type(item) for item in board_candidates if item]
        normalized["board_types"] = sorted(set(board_codes))[:30]

        normalized["seasonal_pricing_periods"] = self._normalize_period_rows(normalized.get("seasonal_pricing_periods"))
        normalized["pricing_lines"] = self._normalize_pricing_line_rows(normalized.get("pricing_lines"))
        normalized["extra_guest_rules"] = self._normalize_extra_guest_rows(normalized.get("extra_guest_rules"))

        normalized["discounts"] = self._coerce_mixed_items_to_text(normalized.get("discounts"))[:80]
        normalized["supplements"] = self._coerce_mixed_items_to_text(normalized.get("supplements"))[:80]
        normalized["marketing_contributions"] = self._coerce_mixed_items_to_text(normalized.get("marketing_contributions"))[:80]

        promotions: list[dict] = []
        promo_rows = normalized.get("promotional_offers")
        if isinstance(promo_rows, list):
            for item in promo_rows[:60]:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name") or item.get("offer_name") or "Promotion").strip() or "Promotion"
                description = str(item.get("description") or "").strip() or None
                discount = self._coerce_float(item.get("discount_percent"))
                start_date = self._coerce_optional_date(item.get("start_date") or item.get("from_date"))
                end_date = self._coerce_optional_date(item.get("end_date") or item.get("to_date"))
                booking_start_date = self._coerce_optional_date(
                    item.get("booking_start_date")
                    or item.get("booking_from_date")
                    or item.get("book_start_date")
                )
                booking_end_date = self._coerce_optional_date(
                    item.get("booking_end_date")
                    or item.get("booking_to_date")
                    or item.get("book_end_date")
                )
                arrival_start_date = self._coerce_optional_date(
                    item.get("arrival_start_date")
                    or item.get("travel_start_date")
                    or start_date
                )
                arrival_end_date = self._coerce_optional_date(
                    item.get("arrival_end_date")
                    or item.get("travel_end_date")
                    or end_date
                )

                is_early_booking = self._is_early_booking_offer(
                    offer_name=name,
                    description=description,
                    expression=str(item.get("expression") or ""),
                )
                if is_early_booking:
                    if not (booking_start_date or booking_end_date) and (arrival_start_date or arrival_end_date):
                        booking_start_date = arrival_start_date
                        booking_end_date = arrival_end_date
                    if not (arrival_start_date or arrival_end_date) and (booking_start_date or booking_end_date):
                        arrival_start_date = booking_start_date
                        arrival_end_date = booking_end_date

                scope_raw = str(item.get("scope") or "").strip().lower()
                scope = scope_raw if scope_raw in {"all", "selected_rooms", "selected_board_types"} else "all"
                room_filters = sorted(
                    {
                        self._to_title_case(value)
                        for value in self._coerce_mixed_items_to_text(item.get("applicable_room_types"))
                        if value
                    }
                )
                board_filters = sorted(
                    {
                        self._normalize_board_type(value)
                        for value in self._coerce_mixed_items_to_text(item.get("applicable_board_types"))
                        if value
                    }
                )
                promotions.append(
                    {
                        "name": name,
                        "description": description,
                        "discount_percent": round(discount, 2) if discount is not None and discount > 0 else None,
                        "start_date": arrival_start_date.isoformat() if arrival_start_date else (start_date.isoformat() if start_date else None),
                        "end_date": arrival_end_date.isoformat() if arrival_end_date else (end_date.isoformat() if end_date else None),
                        "booking_start_date": booking_start_date.isoformat() if booking_start_date else None,
                        "booking_end_date": booking_end_date.isoformat() if booking_end_date else None,
                        "arrival_start_date": arrival_start_date.isoformat() if arrival_start_date else None,
                        "arrival_end_date": arrival_end_date.isoformat() if arrival_end_date else None,
                        "non_cumulative": self._coerce_optional_bool(item.get("non_cumulative")),
                        "scope": scope,
                        "applicable_room_types": room_filters,
                        "applicable_board_types": board_filters,
                        "promotion_category": "early_booking" if is_early_booking else "general",
                    }
                )
        normalized["promotional_offers"] = promotions

        notes_value = normalized.get("notes")
        if notes_value is None:
            normalized["notes"] = ""
        elif not isinstance(notes_value, str):
            normalized["notes"] = str(notes_value)

        for key in CORE_EXTRACT_ARRAY_KEYS:
            value = normalized.get(key)
            if not isinstance(value, list):
                normalized[key] = []
        return normalized

    def _normalize_period_rows(self, value: object) -> list[dict]:
        if not isinstance(value, list):
            return []
        output: list[dict] = []
        for index, item in enumerate(value[:120], start=1):
            if isinstance(item, dict):
                label = str(
                    item.get("label")
                    or item.get("name")
                    or item.get("period")
                    or item.get("period_label")
                    or ""
                ).strip()
                start_date = self._coerce_optional_date(item.get("start_date") or item.get("from_date") or item.get("start"))
                end_date = self._coerce_optional_date(item.get("end_date") or item.get("to_date") or item.get("end"))
                if start_date and end_date and end_date < start_date:
                    start_date, end_date = end_date, start_date
                if not label:
                    if start_date or end_date:
                        start_text = start_date.isoformat() if start_date else "?"
                        end_text = end_date.isoformat() if end_date else "?"
                        label = f"{start_text} - {end_text}"
                    else:
                        label = f"Period {index}"
                output.append(
                    {
                        "label": label,
                        "start_date": start_date.isoformat() if start_date else None,
                        "end_date": end_date.isoformat() if end_date else None,
                    }
                )
                continue

            if item is None:
                continue
            text_value = str(item).strip()
            if text_value:
                output.append(
                    {
                        "label": text_value,
                        "start_date": None,
                        "end_date": None,
                    }
                )
        return output

    def _normalize_pricing_line_rows(self, value: object) -> list[dict]:
        if not isinstance(value, list):
            return []
        output: list[dict] = []
        for item in value[:300]:
            if not isinstance(item, dict):
                continue
            room_type = str(
                item.get("room_type")
                or item.get("room")
                or item.get("room_name")
                or item.get("accommodation")
                or ""
            ).strip() or "Standard Room"
            board_raw = str(item.get("board_type") or item.get("board") or item.get("meal_plan") or item.get("board_basis") or "").strip()
            board_type = self._normalize_board_type(board_raw) if board_raw else "RO"

            period_label = str(
                item.get("period_label")
                or item.get("period")
                or item.get("season")
                or item.get("season_label")
                or ""
            ).strip() or None
            start_date = self._coerce_optional_date(item.get("start_date") or item.get("from_date") or item.get("start"))
            end_date = self._coerce_optional_date(item.get("end_date") or item.get("to_date") or item.get("end"))
            if start_date and end_date and end_date < start_date:
                start_date, end_date = end_date, start_date

            adult_price = self._coerce_float(item.get("adult_price"))
            if adult_price is None:
                adult_price = self._coerce_float(item.get("price"))
            if adult_price is None:
                adult_price = self._coerce_float(item.get("base_rate"))
            if adult_price is None:
                adult_price = self._coerce_float(item.get("rate"))
            if adult_price is None or adult_price <= 0:
                continue

            currency = str(item.get("currency") or item.get("ccy") or "EUR").strip().upper() or "EUR"
            output.append(
                {
                    "room_type": room_type,
                    "board_type": board_type,
                    "period_label": period_label,
                    "start_date": start_date.isoformat() if start_date else None,
                    "end_date": end_date.isoformat() if end_date else None,
                    "adult_price": round(adult_price, 2),
                    "currency": currency,
                }
            )
        return output

    def _normalize_extra_guest_rows(self, value: object) -> list[dict]:
        if not isinstance(value, list):
            return []
        output: list[dict] = []
        for item in value[:120]:
            if not isinstance(item, dict):
                continue
            description = str(
                item.get("description")
                or item.get("condition")
                or item.get("rule")
                or item.get("text")
                or item.get("name")
                or ""
            ).strip()
            percent = self._coerce_float(item.get("percent_of_adult"))
            if percent is None:
                percent = self._coerce_float(item.get("discount_percent"))
            if percent is None and description:
                percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", description)
                if percent_match:
                    percent = self._coerce_float(percent_match.group(1))
            if percent is None or percent <= 0:
                continue

            guest_type = str(item.get("guest_type") or item.get("type") or "").strip().lower()
            if guest_type in {"children", "kid", "kids"}:
                guest_type = "child"
            elif guest_type in {"adults", "extra adult", "extra_adult", "adult guest"}:
                guest_type = "adult"
            elif guest_type in {"baby"}:
                guest_type = "infant"
            elif guest_type not in {"adult", "child", "infant"}:
                lower = description.lower()
                if "child" in lower or "kid" in lower:
                    guest_type = "child"
                elif "adult" in lower:
                    guest_type = "adult"
                elif "infant" in lower or "baby" in lower:
                    guest_type = "infant"
                else:
                    guest_type = "unknown"

            guest_position = self._coerce_int(item.get("guest_position"), default=0)
            if guest_position <= 0 and description:
                guest_position = self._extract_ordinal_from_text(description.lower())

            age_min = self._coerce_int(item.get("age_min"), default=0) or None
            age_max = self._coerce_int(item.get("age_max"), default=0) or None
            if age_min is None and age_max is None and description:
                age_min, age_max = self._extract_age_range_from_text(description)

            output.append(
                {
                    "description": description or "Additional guest rule",
                    "condition": str(item.get("condition") or description or "").strip() or None,
                    "guest_type": guest_type,
                    "guest_position": guest_position if guest_position > 0 else None,
                    "age_min": age_min,
                    "age_max": age_max,
                    "percent_of_adult": round(percent, 2),
                    "room_type": item.get("room_type"),
                    "board_type": item.get("board_type"),
                }
            )
        return output

    def _find_list_by_alias(self, value: object, aliases: tuple[str, ...], depth: int = 0) -> list | None:
        if depth > 5:
            return None
        alias_keys = {self._normalize_extraction_key(alias) for alias in aliases}
        if isinstance(value, dict):
            for raw_key, child in value.items():
                if self._normalize_extraction_key(raw_key) in alias_keys and isinstance(child, list):
                    return child
            for child in value.values():
                found = self._find_list_by_alias(child, aliases=aliases, depth=depth + 1)
                if isinstance(found, list):
                    return found
            return None
        if isinstance(value, list):
            for child in value:
                found = self._find_list_by_alias(child, aliases=aliases, depth=depth + 1)
                if isinstance(found, list):
                    return found
        return None

    def _normalize_extraction_key(self, key: object) -> str:
        if key is None:
            return ""
        return re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower()).strip("_")

    def _merge_extracted_with_fallback(self, extracted_data: dict, fallback_data: dict) -> dict:
        merged = deepcopy(extracted_data) if isinstance(extracted_data, dict) else {}
        if not isinstance(fallback_data, dict):
            return self._stabilize_extracted_data_shape(merged)

        for key in CORE_EXTRACT_ARRAY_KEYS:
            current = merged.get(key)
            if isinstance(current, list) and current:
                continue
            fallback_value = fallback_data.get(key)
            if isinstance(fallback_value, list) and fallback_value:
                merged[key] = deepcopy(fallback_value)

        if not str(merged.get("notes") or "").strip():
            fallback_notes = fallback_data.get("notes")
            if fallback_notes is not None:
                merged["notes"] = str(fallback_notes)

        return self._stabilize_extracted_data_shape(merged)

    def _normalize_database_mapping_rows(self, rows: list[dict]) -> list[dict]:
        normalized: list[dict] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            target_entity = str(row.get("target_entity") or "").strip()
            source_path = str(row.get("source_path") or "").strip()
            destination_field = str(row.get("destination_field") or "").strip()
            transform = str(row.get("transform") or "").strip()
            if not target_entity or not source_path or not destination_field or not transform:
                continue
            note_value = row.get("note")
            normalized.append(
                {
                    "target_entity": target_entity,
                    "source_path": source_path,
                    "destination_field": destination_field,
                    "transform": transform,
                    "required": bool(row.get("required", True)),
                    "note": str(note_value).strip() if note_value is not None else None,
                }
            )
        return normalized[:80]

    def _normalize_ai_extraction(self, extracted_data: dict) -> dict:
        room_types = self._coerce_mixed_items_to_text(extracted_data.get("room_types"))

        seasonal_periods_raw = extracted_data.get("seasonal_pricing_periods") or extracted_data.get("periods")
        seasonal_periods: list[str] = []
        if isinstance(seasonal_periods_raw, list):
            for item in seasonal_periods_raw:
                if isinstance(item, dict):
                    label = str(item.get("label") or "").strip()
                    start = str(item.get("start_date") or "").strip()
                    end = str(item.get("end_date") or "").strip()
                    period_value = label or f"{start} - {end}".strip(" -")
                    if period_value:
                        seasonal_periods.append(period_value)
                elif item is not None:
                    value = str(item).strip()
                    if value:
                        seasonal_periods.append(value)

        board_raw = extracted_data.get("board_types")
        board_types: list[str] = []
        if isinstance(board_raw, list):
            for entry in board_raw:
                if isinstance(entry, dict):
                    candidate = str(entry.get("code") or entry.get("name") or "").strip()
                else:
                    candidate = str(entry).strip() if entry is not None else ""
                if candidate:
                    board_types.append(self._normalize_board_type(candidate))

        discounts = self._coerce_mixed_items_to_text(extracted_data.get("discounts"))
        supplements = self._coerce_mixed_items_to_text(extracted_data.get("supplements"))
        marketing = self._coerce_mixed_items_to_text(extracted_data.get("marketing_contributions"))

        highlights = self._coerce_mixed_items_to_text(extracted_data.get("extra_guest_rules"))
        highlights.extend(self._coerce_mixed_items_to_text(extracted_data.get("promotional_offers")))
        notes_value = extracted_data.get("notes")
        if notes_value is not None:
            text_note = str(notes_value).strip()
            if text_note:
                highlights.append(text_note)

        return {
            "room_types": sorted(set(room_types))[:25],
            "seasonal_periods": sorted(set(seasonal_periods))[:25],
            "board_types": sorted(set(board_types))[:10],
            "discounts": discounts[:25],
            "supplements": supplements[:25],
            "marketing_contributions": marketing[:25],
            "raw_highlights": highlights[:30],
        }

    def _build_rule_drafts_from_ai_data(self, extracted_data: dict) -> list[dict]:
        now = utcnow()
        return self._build_rules_from_ai_data(
            extracted_data=extracted_data,
            contract_id="draft-contract",
            hotel_id=None,
            created_at=now,
            draft_mode=True,
        )

    def _build_rules_from_ai_data(
        self,
        extracted_data: dict,
        contract_id: str,
        hotel_id: str | None,
        created_at: datetime,
        draft_mode: bool = False,
    ) -> list[dict]:
        rules: list[dict] = []

        pricing_lines = extracted_data.get("pricing_lines")
        if isinstance(pricing_lines, list):
            index = 0
            for line in pricing_lines:
                if not isinstance(line, dict):
                    continue
                room = str(line.get("room_type") or "Standard Room").strip()
                board = self._normalize_board_type(str(line.get("board_type") or "RO"))
                rate = self._coerce_float(line.get("adult_price"))
                if rate is None or rate <= 0:
                    continue
                period_label = str(line.get("period_label") or "").strip()

                payload = {
                    "name": f"{room} / {board} base rate",
                    "rule_type": "base_rate",
                    "expression": f"expected = {rate:.2f} * nights",
                    "priority": 100 + index,
                    "metadata": {
                        "room_type": room,
                        "board_type": board,
                        "base_rate": round(rate, 2),
                        "period_label": period_label or None,
                        "currency": line.get("currency"),
                    },
                }
                if draft_mode:
                    rules.append(payload)
                else:
                    rules.append(
                        {
                            "contract_id": contract_id,
                            "hotel_id": hotel_id,
                            "is_active": True,
                            "created_at": created_at,
                            "updated_at": created_at,
                            **payload,
                        }
                    )
                index += 1

        extra_rules = self._collect_extra_guest_rule_candidates(extracted_data)
        for line in extra_rules:
            if not isinstance(line, dict):
                continue

            description = str(line.get("description") or line.get("condition") or "Additional guest rule").strip()
            if not description:
                description = "Additional guest rule"
            source_text = " ".join(
                [
                    description,
                    str(line.get("condition") or ""),
                    str(line.get("guest_type") or ""),
                ]
            ).strip()

            percent = self._coerce_float(line.get("percent_of_adult"))
            if percent is None or percent <= 0:
                continue

            guest_type_hint = str(line.get("guest_type") or "").strip().lower()
            if guest_type_hint in {"children", "kid", "kids"}:
                guest_type_hint = "child"
            elif guest_type_hint in {"adults", "extra adult", "extra_adult", "adult guest"}:
                guest_type_hint = "adult"

            lower = source_text.lower()
            if not guest_type_hint:
                if "child" in lower or "kid" in lower:
                    guest_type_hint = "child"
                elif "adult" in lower:
                    guest_type_hint = "adult"

            rule_type = "child_discount" if guest_type_hint == "child" else "extra_guest_adjustment"

            ordinal_position = self._coerce_int(line.get("guest_position"), default=0)
            if ordinal_position <= 0:
                ordinal_position = self._extract_ordinal_from_text(lower)

            min_children = ordinal_position if rule_type == "child_discount" else 0
            if rule_type == "child_discount" and min_children <= 0:
                min_children = 2 if "2nd child" in lower or "second child" in lower else 1

            age_category, age_bucket, age_label = self._extract_age_bucket_from_text(
                source_text=source_text,
                min_children=min_children,
                rule_type=rule_type,
                guest_type=guest_type_hint or None,
            )
            age_min_raw = self._coerce_int(line.get("age_min"), default=0)
            age_max_raw = self._coerce_int(line.get("age_max"), default=0)
            age_min = age_min_raw if age_min_raw > 0 else None
            age_max = age_max_raw if age_max_raw > 0 else None

            min_adults = ordinal_position if rule_type == "extra_guest_adjustment" and ordinal_position > 1 else None
            expression = (
                f"if pax_children >= {min_children} then child price = {percent:.2f}% of adult base rate"
                if rule_type == "child_discount"
                else (
                    f"if pax_adults >= {min_adults} then {age_label.lower()} price = {percent:.2f}% of adult base rate"
                    if min_adults
                    else f"apply extra guest adjustment: {percent:.2f}% of adult base rate"
                )
            )

            payload = {
                "name": description,
                "rule_type": rule_type,
                "expression": expression,
                "priority": 30 if rule_type == "child_discount" else 25,
                "metadata": {
                    "description": description,
                    "guest_type": guest_type_hint or None,
                    "guest_position": ordinal_position if ordinal_position > 0 else None,
                    "discount_percent": round(percent, 2),
                    "percent_of_adult": round(percent, 2),
                    "min_children": min_children if rule_type == "child_discount" else None,
                    "min_adults": min_adults,
                    "age_category": age_category,
                    "age_bucket": age_bucket,
                    "age_label": age_label,
                    "age_min": age_min,
                    "age_max": age_max,
                    "room_type": line.get("room_type"),
                    "board_type": self._normalize_board_type(str(line.get("board_type") or "")) if line.get("board_type") else None,
                },
            }
            if draft_mode:
                rules.append(payload)
            else:
                rules.append(
                    {
                        "contract_id": contract_id,
                        "hotel_id": hotel_id,
                        "is_active": True,
                        "created_at": created_at,
                        "updated_at": created_at,
                        **payload,
                    }
                )

        promotions = extracted_data.get("promotional_offers")
        if isinstance(promotions, list):
            for promo in promotions:
                if not isinstance(promo, dict):
                    continue
                discount = self._coerce_float(promo.get("discount_percent"))
                if discount is None or discount <= 0:
                    continue
                name = str(promo.get("name") or "Promotion").strip()
                description = str(promo.get("description") or "").strip() or None

                start_date = self._coerce_optional_date(promo.get("start_date"))
                end_date = self._coerce_optional_date(promo.get("end_date"))
                booking_start_date = self._coerce_optional_date(promo.get("booking_start_date"))
                booking_end_date = self._coerce_optional_date(promo.get("booking_end_date"))
                arrival_start_date = self._coerce_optional_date(promo.get("arrival_start_date") or start_date)
                arrival_end_date = self._coerce_optional_date(promo.get("arrival_end_date") or end_date)

                is_early_booking = self._is_early_booking_promotion(
                    {
                        **promo,
                        "offer_name": name,
                        "description": description,
                    }
                )
                if is_early_booking:
                    if not (booking_start_date or booking_end_date) and (arrival_start_date or arrival_end_date):
                        booking_start_date = arrival_start_date
                        booking_end_date = arrival_end_date
                    if not (arrival_start_date or arrival_end_date) and (booking_start_date or booking_end_date):
                        arrival_start_date = booking_start_date
                        arrival_end_date = booking_end_date

                scope_raw = str(promo.get("scope") or "").strip().lower()
                scope = scope_raw if scope_raw in {"all", "selected_rooms", "selected_board_types"} else "all"
                room_filters = sorted(
                    {
                        self._to_title_case(value)
                        for value in self._coerce_mixed_items_to_text(promo.get("applicable_room_types"))
                        if value
                    }
                )
                board_filters = sorted(
                    {
                        self._normalize_board_type(value)
                        for value in self._coerce_mixed_items_to_text(promo.get("applicable_board_types"))
                        if value
                    }
                )

                expression = (
                    f"if booking_date and stay_date are within offer window then discount {discount:.2f}%"
                    if is_early_booking
                    else f"if promotion active then discount {discount:.2f}%"
                )
                payload = {
                    "name": name,
                    "rule_type": "promotion",
                    "expression": expression,
                    "priority": 40,
                    "metadata": {
                        "offer_name": name,
                        "description": description,
                        "discount_percent": round(discount, 2),
                        "start_date": arrival_start_date.isoformat() if arrival_start_date else (start_date.isoformat() if start_date else None),
                        "end_date": arrival_end_date.isoformat() if arrival_end_date else (end_date.isoformat() if end_date else None),
                        "booking_start_date": booking_start_date.isoformat() if booking_start_date else None,
                        "booking_end_date": booking_end_date.isoformat() if booking_end_date else None,
                        "arrival_start_date": arrival_start_date.isoformat() if arrival_start_date else None,
                        "arrival_end_date": arrival_end_date.isoformat() if arrival_end_date else None,
                        "scope": scope,
                        "non_cumulative": bool(self._coerce_optional_bool(promo.get("non_cumulative"))),
                        "applicable_room_types": room_filters,
                        "applicable_board_types": board_filters,
                        "promotion_category": "early_booking" if is_early_booking else "general",
                    },
                }
                if draft_mode:
                    rules.append(payload)
                else:
                    rules.append(
                        {
                            "contract_id": contract_id,
                            "hotel_id": hotel_id,
                            "is_active": True,
                            "created_at": created_at,
                            "updated_at": created_at,
                            **payload,
                        }
                    )

        return rules

    def _collect_extra_guest_rule_candidates(self, extracted_data: dict) -> list[dict]:
        candidates: list[dict] = []
        seen: set[tuple[str, float]] = set()

        direct_rules = extracted_data.get("extra_guest_rules")
        if isinstance(direct_rules, list):
            for item in direct_rules:
                if not isinstance(item, dict):
                    continue
                description = str(item.get("description") or item.get("condition") or "").strip()
                percent = self._coerce_float(item.get("percent_of_adult"))
                if (percent is None or percent <= 0) and description:
                    percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", description)
                    if percent_match:
                        percent = self._coerce_float(percent_match.group(1))
                if not description or percent is None or percent <= 0:
                    continue
                key = (description.lower(), round(percent, 4))
                if key in seen:
                    continue
                seen.add(key)
                candidates.append(item)

        fallback_lines: list[str] = []
        fallback_lines.extend(self._coerce_mixed_items_to_text(extracted_data.get("discounts")))
        fallback_lines.extend(self._coerce_mixed_items_to_text(extracted_data.get("supplements")))
        fallback_lines.extend(self._coerce_mixed_items_to_text(extracted_data.get("raw_highlights")))

        notes_value = extracted_data.get("notes")
        if notes_value is not None:
            note_text = str(notes_value).strip()
            if note_text:
                fallback_lines.append(note_text)

        for line in fallback_lines:
            lower = line.lower()
            if not any(token in lower for token in ("child", "adult", "infant", "extra guest", "2nd", "second", "3rd", "third", "4th", "fourth")):
                continue
            percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", line)
            if not percent_match:
                continue
            percent = self._coerce_float(percent_match.group(1))
            if percent is None or percent <= 0:
                continue
            key = (line.strip().lower(), round(percent, 4))
            if key in seen:
                continue
            seen.add(key)
            guest_type = "child" if "child" in lower or "kid" in lower else ("adult" if "adult" in lower else None)
            candidates.append(
                {
                    "description": line.strip(),
                    "condition": line.strip(),
                    "guest_type": guest_type,
                    "guest_position": self._extract_ordinal_from_text(lower) or None,
                    "percent_of_adult": percent,
                }
            )

        return candidates

    def _recommend_content_for_mapping(self, text: str) -> dict:
        lines = [line.strip() for line in text.splitlines() if line and line.strip()]
        char_count = len(text)
        line_count = len(lines)
        table_like_lines = sum(
            1
            for line in lines
            if "|" in line
            or "\t" in line
            or line.count(";") >= 2
            or len(re.findall(r"\d+(?:[.,]\d+)?", line)) >= 3
        )
        period_mentions = len(
            re.findall(
                r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|"
                r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\b",
                text,
                flags=re.IGNORECASE,
            )
        )
        price_mentions = len(
            re.findall(
                r"(?:€|\$|£|\bEUR\b|\bUSD\b|\bGBP\b|\bAED\b|\d{2,4}(?:[.,]\d{1,2})?)",
                text,
                flags=re.IGNORECASE,
            )
        )
        room_mentions = len(
            re.findall(
                r"\b(room|suite|studio|villa|family|double|single|twin)\b",
                text,
                flags=re.IGNORECASE,
            )
        )
        board_mentions = len(
            re.findall(
                r"\b(ro|bb|hb|hb\+|fb|ai|bed\s*&\s*breakfast|half\s*board|full\s*board|all\s*inclusive)\b",
                text,
                flags=re.IGNORECASE,
            )
        )
        guest_rule_mentions = len(
            re.findall(
                r"\b(child|children|kid|adult|infant|extra guest|supplement|2nd|second|3rd|third|4th|fourth)\b",
                text,
                flags=re.IGNORECASE,
            )
        )

        complexity_score = 0
        if char_count > 30000:
            complexity_score += 2
        if char_count > 90000:
            complexity_score += 2
        if line_count > 600:
            complexity_score += 2
        if table_like_lines > 120:
            complexity_score += 2
        if price_mentions > 220:
            complexity_score += 1
        if period_mentions > 30:
            complexity_score += 1
        if room_mentions > 40:
            complexity_score += 1
        if board_mentions > 30:
            complexity_score += 1
        if guest_rule_mentions > 15:
            complexity_score += 1

        extraction = self._extract_contract_terms(text)
        period_suggestions = self._build_period_suggestions(extraction.get("seasonal_periods", []))
        rate_candidates = self._extract_rate_candidates(text)
        extra_guest_rules = self._collect_extra_guest_rule_candidates(
            {
                "extra_guest_rules": [],
                "discounts": extraction.get("discounts", []),
                "supplements": extraction.get("supplements", []),
                "raw_highlights": extraction.get("raw_highlights", []),
                "notes": None,
            }
        )
        normalized_extra_rules = self._normalize_extra_guest_rule_suggestions(extra_guest_rules)
        promotion_candidates = self._extract_promotional_offer_candidates(text)
        pricing_lines = self._build_pricing_line_suggestions(
            room_types=extraction.get("room_types", []),
            board_types=extraction.get("board_types", []),
            period_suggestions=period_suggestions,
            rates=rate_candidates,
        )

        detected_signals = {
            "complexity_score": complexity_score,
            "char_count": char_count,
            "line_count": line_count,
            "table_like_lines": table_like_lines,
            "price_mentions": price_mentions,
            "period_mentions": period_mentions,
            "room_mentions": room_mentions,
            "board_mentions": board_mentions,
            "guest_rule_mentions": guest_rule_mentions,
            "room_type_count": len(extraction.get("room_types", [])),
            "board_type_count": len(extraction.get("board_types", [])),
            "seasonal_period_count": len(period_suggestions),
            "pricing_line_suggestion_count": len(pricing_lines),
            "extra_guest_rule_count": len(normalized_extra_rules),
            "promotion_offer_count": len(promotion_candidates),
        }

        coverage_feedback: list[str] = []
        if not extraction.get("room_types"):
            coverage_feedback.append("Room types were not confidently detected. Consider adding room-type hints in mapping instructions.")
        if not extraction.get("board_types"):
            coverage_feedback.append("Board types were not clearly detected. Verify board naming conventions (RO/BB/HB/FB/AI).")
        if not period_suggestions:
            coverage_feedback.append("Seasonal periods were weakly detected. Ensure period labels/date ranges are mapped explicitly.")
        if not pricing_lines:
            coverage_feedback.append("No reliable pricing line suggestions were generated from text values. Keep pricing_lines flexible and review manually.")
        if not normalized_extra_rules:
            coverage_feedback.append("No extra guest rules were confidently extracted. Add explicit instructions for child/adult occupancy logic.")
        if not promotion_candidates:
            coverage_feedback.append("No promotional offers were clearly identified from this content snapshot.")

        if len(coverage_feedback) <= 1:
            confidence = "high"
        elif len(coverage_feedback) <= 3:
            confidence = "medium"
        else:
            confidence = "low"

        content_summary = (
            f"Detected {len(extraction.get('room_types', []))} room types, {len(period_suggestions)} seasonal periods, "
            f"{len(extraction.get('board_types', []))} board types, {len(normalized_extra_rules)} extra-guest rules, "
            f"and {len(promotion_candidates)} promotion hints."
        )

        recommended_data = {
            "room_types": extraction.get("room_types", []),
            "seasonal_pricing_periods": period_suggestions,
            "board_types": extraction.get("board_types", []),
            "pricing_lines": pricing_lines,
            "extra_guest_rules": normalized_extra_rules,
            "discounts": extraction.get("discounts", []),
            "supplements": extraction.get("supplements", []),
            "marketing_contributions": extraction.get("marketing_contributions", []),
            "promotional_offers": promotion_candidates,
            "notes": "Generated as pre-persistence mapping feedback. Review and refine before final extraction.",
        }

        suggested_mapping_instructions = (
            "Treat this input as a tour-operator commercial terms contract. Extract room_types, board_types, "
            "seasonal_pricing_periods, pricing_lines, and extra_guest_rules with emphasis on occupancy logic and "
            "field-level mapping to persistence entities. For each extra_guest_rule include guest_type, guest_position, "
            "age_min/age_max when present, and percent_of_adult numeric values. Available board types are HB, BB, FB."
        )
        suggested_schema, suggested_schema_rationale = self._build_recommended_schema_for_contract(
            complexity_score=complexity_score
        )

        return {
            "content_summary": content_summary,
            "confidence": confidence,
            "detected_signals": detected_signals,
            "coverage_feedback": coverage_feedback,
            "recommended_data": recommended_data,
            "suggested_schema": suggested_schema,
            "suggested_schema_rationale": suggested_schema_rationale,
            "suggested_mapping_instructions": suggested_mapping_instructions,
            "database_mapping": self._build_database_mapping_template(),
        }

    def _build_period_suggestions(self, period_values: list[str]) -> list[dict]:
        suggestions: list[dict] = []
        seen: set[str] = set()

        for raw in period_values[:30]:
            label = str(raw or "").strip()
            if not label:
                continue
            key = label.lower()
            if key in seen:
                continue
            seen.add(key)

            start_date: date | None = None
            end_date: date | None = None
            date_match = re.search(
                r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
                label,
                flags=re.IGNORECASE,
            )
            if date_match:
                start_date = self._parse_date_string(date_match.group(1))
                end_date = self._parse_date_string(date_match.group(2))
                if start_date and end_date and end_date < start_date:
                    start_date, end_date = end_date, start_date

            suggestions.append(
                {
                    "label": label,
                    "start_date": start_date.isoformat() if start_date else None,
                    "end_date": end_date.isoformat() if end_date else None,
                }
            )

        return suggestions

    def _build_pricing_line_suggestions(
        self,
        room_types: list[str],
        board_types: list[str],
        period_suggestions: list[dict],
        rates: list[float],
    ) -> list[dict]:
        if not room_types or not board_types or not rates:
            return []

        suggestions: list[dict] = []
        rate_index = 0
        period_cursor = 0
        max_rows = 24

        for room in room_types[:6]:
            for board in board_types[:4]:
                if len(suggestions) >= max_rows:
                    return suggestions
                period = period_suggestions[period_cursor % len(period_suggestions)] if period_suggestions else None
                suggestions.append(
                    {
                        "room_type": room,
                        "board_type": board,
                        "period_label": period.get("label") if period else None,
                        "start_date": period.get("start_date") if period else None,
                        "end_date": period.get("end_date") if period else None,
                        "adult_price": rates[rate_index % len(rates)],
                        "currency": "EUR",
                    }
                )
                rate_index += 1
                period_cursor += 1

        return suggestions

    def _normalize_extra_guest_rule_suggestions(self, rules: list[dict]) -> list[dict]:
        normalized: list[dict] = []
        for item in rules[:30]:
            if not isinstance(item, dict):
                continue
            description = str(item.get("description") or item.get("condition") or "").strip()
            if not description:
                continue

            percent = self._coerce_float(item.get("percent_of_adult"))
            if percent is None or percent <= 0:
                match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", description)
                if match:
                    percent = self._coerce_float(match.group(1))
            if percent is None or percent <= 0:
                continue

            guest_type = str(item.get("guest_type") or "").strip().lower()
            if guest_type in {"children", "kid", "kids"}:
                guest_type = "child"
            elif guest_type in {"adults", "adult guest", "extra adult", "extra_adult"}:
                guest_type = "adult"
            if guest_type not in {"adult", "child", "infant"}:
                lower = description.lower()
                if "child" in lower or "kid" in lower:
                    guest_type = "child"
                elif "adult" in lower:
                    guest_type = "adult"
                elif "infant" in lower or "baby" in lower:
                    guest_type = "infant"
                else:
                    guest_type = "unknown"

            guest_position = self._coerce_int(item.get("guest_position"), default=0) or self._extract_ordinal_from_text(description.lower()) or None
            age_min = self._coerce_int(item.get("age_min"), default=0) or None
            age_max = self._coerce_int(item.get("age_max"), default=0) or None
            if age_min is None and age_max is None:
                age_min, age_max = self._extract_age_range_from_text(description)

            normalized.append(
                {
                    "description": description,
                    "condition": str(item.get("condition") or description).strip(),
                    "guest_type": guest_type,
                    "guest_position": guest_position,
                    "age_min": age_min,
                    "age_max": age_max,
                    "percent_of_adult": round(percent, 2),
                }
            )

        return normalized

    def _extract_promotional_offer_candidates(self, text: str) -> list[dict]:
        lines = self._extract_keyword_lines(
            text,
            keywords=["promotion", "offer", "early bird", "valentine", "flash", "long stay", "discount"],
        )
        offers: list[dict] = []
        seen: set[str] = set()
        for line in lines[:30]:
            line_text = str(line or "").strip()
            if not line_text:
                continue
            percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", line_text)
            if not percent_match:
                continue
            percent = self._coerce_float(percent_match.group(1))
            if percent is None or percent <= 0:
                continue

            lower = line_text.lower()
            offer_name = "Promotion"
            for keyword, normalized in PROMOTION_KEYWORDS.items():
                if keyword in lower:
                    offer_name = normalized
                    break

            parsed_offer = self._extract_promotion_offer(line_text, fallback_name=offer_name)
            start_date = self._coerce_optional_date(parsed_offer.get("start_date"))
            end_date = self._coerce_optional_date(parsed_offer.get("end_date"))
            booking_start_date = self._coerce_optional_date(parsed_offer.get("booking_start_date"))
            booking_end_date = self._coerce_optional_date(parsed_offer.get("booking_end_date"))
            arrival_start_date = self._coerce_optional_date(parsed_offer.get("arrival_start_date"))
            arrival_end_date = self._coerce_optional_date(parsed_offer.get("arrival_end_date"))

            key = (
                f"{offer_name.lower()}|{percent:.2f}|{start_date}|{end_date}|"
                f"{booking_start_date}|{booking_end_date}|{arrival_start_date}|{arrival_end_date}"
            )
            if key in seen:
                continue
            seen.add(key)

            offers.append(
                {
                    "name": offer_name,
                    "description": line_text,
                    "discount_percent": round(percent, 2),
                    "start_date": start_date.isoformat() if start_date else None,
                    "end_date": end_date.isoformat() if end_date else None,
                    "booking_start_date": booking_start_date.isoformat() if booking_start_date else None,
                    "booking_end_date": booking_end_date.isoformat() if booking_end_date else None,
                    "arrival_start_date": arrival_start_date.isoformat() if arrival_start_date else None,
                    "arrival_end_date": arrival_end_date.isoformat() if arrival_end_date else None,
                }
            )

        return offers[:15]

    def _build_database_mapping_template(self) -> list[dict]:
        return [
            {
                "target_entity": "ContractDocument",
                "source_path": "$.room_types",
                "destination_field": "extraction.room_types",
                "transform": "dedupe + normalize title case",
                "required": True,
                "note": "Used for search and contract context.",
            },
            {
                "target_entity": "ContractDocument",
                "source_path": "$.seasonal_pricing_periods",
                "destination_field": "extraction.seasonal_periods",
                "transform": "preserve labels and normalize dates",
                "required": True,
                "note": "Feeds contract filtering and matrix periods.",
            },
            {
                "target_entity": "ContractDocument",
                "source_path": "$.board_types",
                "destination_field": "extraction.board_types",
                "transform": "map to RO/BB/HB/FB/AI",
                "required": True,
                "note": "Ensures board consistency across rules and matrix.",
            },
            {
                "target_entity": "PriceListEntry",
                "source_path": "$.pricing_lines[*]",
                "destination_field": "room/board/period/date/price",
                "transform": "normalize board code + currency + dates",
                "required": True,
                "note": "Primary adult base-rate matrix source.",
            },
            {
                "target_entity": "PricingRule",
                "source_path": "$.extra_guest_rules[*]",
                "destination_field": "rule_type + expression + metadata",
                "transform": "child => child_discount, adult => extra_guest_adjustment",
                "required": False,
                "note": "Captures occupancy logic such as 2nd child / 3rd adult.",
            },
            {
                "target_entity": "PricingRule",
                "source_path": "$.discounts[*]",
                "destination_field": "metadata.discount narratives",
                "transform": "parse percentages where explicit",
                "required": False,
                "note": "Fallback for occupancy discounts when explicit extra rules are absent.",
            },
            {
                "target_entity": "PromotionOffer",
                "source_path": "$.promotional_offers[*]",
                "destination_field": "offer_name/description/discount/date range",
                "transform": "normalize dates and percentages",
                "required": False,
                "note": "Used in validation and promotion-aware pricing checks.",
            },
        ]

    def _build_recommended_schema_for_contract(self, complexity_score: int) -> tuple[dict, str]:
        schema = deepcopy(DEFAULT_AI_PRICING_SCHEMA)
        properties = schema.get("properties")
        if not isinstance(properties, dict):
            return schema, "Default schema was used because schema properties were not available."

        schema["required"] = [
            "room_types",
            "seasonal_pricing_periods",
            "board_types",
            "pricing_lines",
            "extra_guest_rules",
            "discounts",
            "supplements",
            "promotional_offers",
        ]

        pricing_lines = properties.get("pricing_lines")
        if isinstance(pricing_lines, dict):
            items = pricing_lines.get("items")
            if isinstance(items, dict):
                items["required"] = ["room_type", "board_type", "adult_price"]
                item_properties = items.get("properties")
                if isinstance(item_properties, dict):
                    if isinstance(item_properties.get("adult_price"), dict):
                        item_properties["adult_price"]["description"] = "Adult base rate per night in contract currency."
                    if isinstance(item_properties.get("period_label"), dict):
                        item_properties["period_label"]["description"] = "Period label from contract (e.g., Low Season A)."
                    if isinstance(item_properties.get("start_date"), dict):
                        item_properties["start_date"]["description"] = "Optional ISO date when explicit in document."
                    if isinstance(item_properties.get("end_date"), dict):
                        item_properties["end_date"]["description"] = "Optional ISO date when explicit in document."

        extra_guest_rules = properties.get("extra_guest_rules")
        if isinstance(extra_guest_rules, dict):
            extra_items = extra_guest_rules.get("items")
            if isinstance(extra_items, dict):
                extra_items["required"] = ["description", "percent_of_adult"]
                extra_properties = extra_items.get("properties")
                if isinstance(extra_properties, dict):
                    extra_properties["guest_type"] = {
                        "type": "string",
                        "enum": ["adult", "child", "infant", "unknown"],
                    }
                    extra_properties["guest_position"] = {
                        "type": "number",
                        "description": "Ordinal position such as 2 for 2nd child or 3 for 3rd adult.",
                    }
                    extra_properties["percent_of_adult"] = {
                        "type": "number",
                        "description": "Commercial pricing ratio versus adult base rate.",
                    }

        seasonal_periods = properties.get("seasonal_pricing_periods")
        if isinstance(seasonal_periods, dict):
            seasonal_items = seasonal_periods.get("items")
            if isinstance(seasonal_items, dict):
                seasonal_items["required"] = ["label"]

        promotional_offers = properties.get("promotional_offers")
        if isinstance(promotional_offers, dict):
            promo_items = promotional_offers.get("items")
            if isinstance(promo_items, dict):
                promo_items["required"] = ["name", "discount_percent"]
                promo_properties = promo_items.get("properties")
                if isinstance(promo_properties, dict):
                    promo_properties["booking_start_date"] = {
                        "type": ["string", "null"],
                        "description": "Booking window start for early-booking offers (ISO date).",
                    }
                    promo_properties["booking_end_date"] = {
                        "type": ["string", "null"],
                        "description": "Booking window end for early-booking offers (ISO date).",
                    }
                    promo_properties["arrival_start_date"] = {
                        "type": ["string", "null"],
                        "description": "Stay/arrival window start for offer applicability (ISO date).",
                    }
                    promo_properties["arrival_end_date"] = {
                        "type": ["string", "null"],
                        "description": "Stay/arrival window end for offer applicability (ISO date).",
                    }
                    promo_properties["scope"] = {
                        "type": ["string", "null"],
                        "description": "Offer scope: all, selected_rooms, selected_board_types.",
                    }
                    promo_properties["non_cumulative"] = {
                        "type": ["boolean", "null"],
                        "description": "True when offer cannot be combined with other offers.",
                    }
                    promo_properties["applicable_room_types"] = {
                        "type": "array",
                        "items": {"type": "string"},
                    }
                    promo_properties["applicable_board_types"] = {
                        "type": "array",
                        "items": {"type": "string"},
                    }

        if complexity_score >= 8:
            schema_rationale = (
                "Stricter recommended schema for complex contracts: essential pricing arrays are required and "
                "line-level occupancy fields are enforced to reduce rule-loss during extraction."
            )
        else:
            schema_rationale = (
                "Balanced recommended schema: keeps flexible structure while requiring key pricing entities "
                "needed for contract-to-rule persistence."
            )

        return schema, schema_rationale

    def _extract_price_period_ranges(self, ai_extracted_data: dict) -> list[dict]:
        periods: dict[str, dict] = {}
        raw_periods = ai_extracted_data.get("seasonal_pricing_periods")
        if not isinstance(raw_periods, list):
            return []

        for index, item in enumerate(raw_periods, start=1):
            if not isinstance(item, dict):
                continue

            label = str(item.get("label") or item.get("name") or f"Period {index}").strip()
            if not label:
                label = f"Period {index}"
            start_date = self._coerce_optional_date(item.get("start_date") or item.get("from_date"))
            end_date = self._coerce_optional_date(item.get("end_date") or item.get("to_date"))
            if start_date and end_date and end_date < start_date:
                start_date, end_date = end_date, start_date

            key = label.lower()
            existing = periods.get(key)
            if not existing:
                periods[key] = {
                    "label": label,
                    "start_date": start_date,
                    "end_date": end_date,
                }
                continue

            if start_date and (existing.get("start_date") is None or start_date < existing["start_date"]):
                existing["start_date"] = start_date
            if end_date and (existing.get("end_date") is None or end_date > existing["end_date"]):
                existing["end_date"] = end_date

        output = list(periods.values())
        output.sort(key=lambda row: (row.get("start_date") or date.max, row.get("label") or ""))
        return output

    def _resolve_price_period(
        self,
        period_label: str | None,
        period_lookup: dict[str, dict],
        fallback_start: date | None = None,
        fallback_end: date | None = None,
    ) -> tuple[date | None, date | None, str | None]:
        label_value = (period_label or "").strip()
        start_date = fallback_start
        end_date = fallback_end
        resolved_label = label_value or None

        if label_value:
            direct = period_lookup.get(label_value.lower())
            if not direct:
                label_key = label_value.lower()
                for key, value in period_lookup.items():
                    if label_key in key or key in label_key:
                        direct = value
                        break
            if direct:
                resolved_label = str(direct.get("label") or resolved_label or "").strip() or None
                start_date = direct.get("start_date") or start_date
                end_date = direct.get("end_date") or end_date

        if start_date and end_date and end_date < start_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date, resolved_label

    def _extract_base_price_entries(
        self,
        rules: list[dict],
        ai_extracted_data: dict,
        period_lookup: dict[str, dict],
    ) -> list[dict]:
        entries: list[dict] = []

        for rule in rules:
            if rule.get("rule_type") != "base_rate":
                continue
            metadata = rule.get("metadata")
            if not isinstance(metadata, dict):
                continue
            rate = self._coerce_float(metadata.get("base_rate"))
            if rate is None or rate <= 0:
                continue

            room_type = str(metadata.get("room_type") or "Standard Room").strip() or "Standard Room"
            board_type = self._normalize_board_type(str(metadata.get("board_type") or "RO"))
            period_label = str(metadata.get("period_label") or "").strip() or None
            fallback_start = self._coerce_optional_date(metadata.get("start_date"))
            fallback_end = self._coerce_optional_date(metadata.get("end_date"))
            start_date, end_date, resolved_period = self._resolve_price_period(
                period_label=period_label,
                period_lookup=period_lookup,
                fallback_start=fallback_start,
                fallback_end=fallback_end,
            )
            currency = str(metadata.get("currency") or ai_extracted_data.get("currency") or "EUR").strip().upper() or "EUR"

            entries.append(
                {
                    "room_type": room_type,
                    "board_type": board_type,
                    "age_bucket": "adult",
                    "age_label": "Adult",
                    "age_category": "adult",
                    "age_min": None,
                    "age_max": None,
                    "period_label": resolved_period,
                    "start_date": start_date,
                    "end_date": end_date,
                    "price": round(rate, 2),
                    "currency": currency,
                    "source_rule_type": "base_rate",
                    "source_rule_name": rule.get("name"),
                }
            )

        if entries:
            return entries

        pricing_lines = ai_extracted_data.get("pricing_lines")
        if not isinstance(pricing_lines, list):
            return entries

        for line in pricing_lines:
            if not isinstance(line, dict):
                continue
            rate = self._coerce_float(line.get("adult_price"))
            if rate is None or rate <= 0:
                continue

            room_type = str(line.get("room_type") or "Standard Room").strip() or "Standard Room"
            board_type = self._normalize_board_type(str(line.get("board_type") or "RO"))
            period_label = str(line.get("period_label") or "").strip() or None
            fallback_start = self._coerce_optional_date(line.get("start_date"))
            fallback_end = self._coerce_optional_date(line.get("end_date"))
            start_date, end_date, resolved_period = self._resolve_price_period(
                period_label=period_label,
                period_lookup=period_lookup,
                fallback_start=fallback_start,
                fallback_end=fallback_end,
            )
            currency = str(line.get("currency") or ai_extracted_data.get("currency") or "EUR").strip().upper() or "EUR"

            entries.append(
                {
                    "room_type": room_type,
                    "board_type": board_type,
                    "age_bucket": "adult",
                    "age_label": "Adult",
                    "age_category": "adult",
                    "age_min": None,
                    "age_max": None,
                    "period_label": resolved_period,
                    "start_date": start_date,
                    "end_date": end_date,
                    "price": round(rate, 2),
                    "currency": currency,
                    "source_rule_type": "base_rate",
                    "source_rule_name": "AI pricing line",
                }
            )

        return entries

    def _extract_age_adjustments(self, rules: list[dict], ai_extracted_data: dict) -> list[dict]:
        adjustments: list[dict] = []

        for rule in rules:
            rule_type = str(rule.get("rule_type") or "").strip()
            if rule_type not in {"child_discount", "extra_guest_adjustment", "base_rate"}:
                continue
            metadata = rule.get("metadata")
            if not isinstance(metadata, dict):
                continue

            if rule_type == "base_rate":
                base_rate = self._coerce_float(metadata.get("base_rate"))
                if base_rate is not None and base_rate > 0:
                    continue

            discount_percent = self._coerce_float(metadata.get("percent_of_adult"))
            if discount_percent is None or discount_percent <= 0:
                discount_percent = self._coerce_float(metadata.get("discount_percent"))
            if discount_percent is None or discount_percent <= 0:
                continue

            source_text = " ".join(
                [
                    str(rule.get("name") or ""),
                    str(rule.get("expression") or ""),
                    str(metadata.get("description") or ""),
                ]
            ).strip()
            min_children = self._coerce_int(metadata.get("min_children"), default=0)
            age_category = str(metadata.get("age_category") or "").strip().lower()
            age_bucket = str(metadata.get("age_bucket") or "").strip()
            age_label = str(metadata.get("age_label") or "").strip()
            if not age_category or not age_bucket or not age_label:
                age_category, age_bucket, age_label = self._extract_age_bucket_from_text(
                    source_text=source_text,
                    min_children=min_children,
                    rule_type=rule_type,
                    guest_type=str(metadata.get("guest_type") or "") or None,
                )

            age_min = self._coerce_int(metadata.get("age_min"), default=0) or None
            age_max = self._coerce_int(metadata.get("age_max"), default=0) or None
            if age_min is None and age_max is None:
                age_min, age_max = self._extract_age_range_from_text(source_text)

            adjustments.append(
                {
                    "age_category": age_category,
                    "age_bucket": age_bucket,
                    "age_label": age_label,
                    "age_min": age_min,
                    "age_max": age_max,
                    "multiplier": round(discount_percent / 100.0, 4),
                    "source_rule_type": rule_type,
                    "source_rule_name": rule.get("name"),
                }
            )

        if adjustments:
            return adjustments

        return self._extract_adjustments_from_extra_guest_rules(ai_extracted_data)

    def _extract_adjustments_from_extra_guest_rules(self, ai_extracted_data: dict) -> list[dict]:
        adjustments: list[dict] = []
        extra_rules = self._collect_extra_guest_rule_candidates(ai_extracted_data)
        if not extra_rules:
            return adjustments

        for item in extra_rules:
            if not isinstance(item, dict):
                continue
            description = str(item.get("description") or item.get("condition") or "").strip()
            percent = self._coerce_float(item.get("percent_of_adult"))
            if (percent is None or percent <= 0) and description:
                percent_match = re.search(r"(\d{1,2}(?:\.\d{1,2})?)\s*%", description)
                if percent_match:
                    percent = self._coerce_float(percent_match.group(1))
            if percent is None or percent <= 0:
                continue

            guest_type_hint = str(item.get("guest_type") or "").strip().lower() or None
            if guest_type_hint in {"children", "kid", "kids"}:
                guest_type_hint = "child"
            elif guest_type_hint in {"adults", "extra adult", "extra_adult", "adult guest"}:
                guest_type_hint = "adult"
            min_children = self._coerce_int(item.get("guest_position"), default=0)
            normalized_rule_type = "child_discount" if (guest_type_hint == "child" or "child" in description.lower()) else "extra_guest_adjustment"
            age_category, age_bucket, age_label = self._extract_age_bucket_from_text(
                source_text=description,
                min_children=min_children if normalized_rule_type == "child_discount" else 0,
                rule_type=normalized_rule_type,
                guest_type=guest_type_hint,
            )
            age_min_raw = self._coerce_int(item.get("age_min"), default=0)
            age_max_raw = self._coerce_int(item.get("age_max"), default=0)
            age_min = age_min_raw if age_min_raw > 0 else None
            age_max = age_max_raw if age_max_raw > 0 else None
            if age_min is None and age_max is None:
                age_min, age_max = self._extract_age_range_from_text(description)
            adjustments.append(
                {
                    "age_category": age_category,
                    "age_bucket": age_bucket,
                    "age_label": age_label,
                    "age_min": age_min,
                    "age_max": age_max,
                    "multiplier": round(percent / 100.0, 4),
                    "source_rule_type": "extra_guest_adjustment",
                    "source_rule_name": description or "Additional guest rule",
                }
            )

        return adjustments

    def _derive_adjusted_price_entries(self, base_entries: list[dict], adjustments: list[dict]) -> list[dict]:
        derived: list[dict] = []
        if not base_entries or not adjustments:
            return derived

        for base in base_entries:
            base_price = self._coerce_float(base.get("price"))
            if base_price is None or base_price <= 0:
                continue
            for adjustment in adjustments:
                multiplier = self._coerce_float(adjustment.get("multiplier"))
                if multiplier is None or multiplier <= 0:
                    continue
                age_bucket = str(adjustment.get("age_bucket") or "").strip()
                if not age_bucket:
                    continue

                derived.append(
                    {
                        "room_type": base.get("room_type"),
                        "board_type": base.get("board_type"),
                        "age_bucket": age_bucket,
                        "age_label": adjustment.get("age_label") or age_bucket,
                        "age_category": adjustment.get("age_category") or "unknown",
                        "age_min": adjustment.get("age_min"),
                        "age_max": adjustment.get("age_max"),
                        "period_label": base.get("period_label"),
                        "start_date": base.get("start_date"),
                        "end_date": base.get("end_date"),
                        "price": round(base_price * multiplier, 2),
                        "currency": base.get("currency"),
                        "source_rule_type": "derived_adjustment",
                        "source_rule_name": adjustment.get("source_rule_name"),
                    }
                )

        return derived

    def _apply_promotions_to_matrix_entries(
        self,
        entries: list[dict],
        promotions: list[dict],
        booking_date: date | None = None,
    ) -> tuple[list[dict], set[str], set[str]]:
        if not entries or not promotions:
            return entries, set(), set()

        output: list[dict] = []
        used_promotion_ids: set[str] = set()
        used_promotion_names: set[str] = set()

        for entry in entries:
            base_price = self._coerce_float(entry.get("price"))
            if base_price is None or base_price <= 0:
                output.append(dict(entry))
                continue

            applicable: list[dict] = []
            for promotion in promotions:
                if self._promotion_applies_to_matrix_entry(
                    promotion=promotion,
                    entry=entry,
                    booking_date=booking_date,
                ):
                    applicable.append(promotion)

            if not applicable:
                output.append(
                    {
                        **entry,
                        "base_price": round(base_price, 2),
                        "delta_amount": 0.0,
                        "delta_percent": 0.0,
                        "promotion_applied": False,
                        "applied_promotions": [],
                    }
                )
                continue

            non_cumulative_promos = [promo for promo in applicable if bool(promo.get("non_cumulative", False))]
            applied_promos: list[dict]
            final_multiplier = 1.0
            if non_cumulative_promos:
                best = max(
                    non_cumulative_promos,
                    key=lambda promo: self._coerce_float(promo.get("discount_percent")) or 0.0,
                )
                discount = self._coerce_float(best.get("discount_percent")) or 0.0
                final_multiplier = max(0.0, 1.0 - discount / 100.0)
                applied_promos = [best]
            else:
                applied_promos = []
                for promo in applicable:
                    discount = self._coerce_float(promo.get("discount_percent"))
                    if discount is None or discount <= 0:
                        continue
                    final_multiplier *= max(0.0, 1.0 - discount / 100.0)
                    applied_promos.append(promo)

            adjusted_price = round(base_price * final_multiplier, 2)
            delta_amount = round(adjusted_price - base_price, 2)
            delta_percent = round((delta_amount / base_price * 100.0), 2) if base_price > 0 else 0.0
            promotion_names = sorted(
                {
                    str(item.get("offer_name") or "Promotion").strip() or "Promotion"
                    for item in applied_promos
                }
            )
            for item in applied_promos:
                promotion_id = str(item.get("id") or "").strip()
                if promotion_id:
                    used_promotion_ids.add(promotion_id)
                offer_name = str(item.get("offer_name") or "").strip()
                if offer_name:
                    used_promotion_names.add(offer_name)

            output.append(
                {
                    **entry,
                    "price": adjusted_price,
                    "base_price": round(base_price, 2),
                    "delta_amount": delta_amount,
                    "delta_percent": delta_percent,
                    "promotion_applied": bool(promotion_names and abs(delta_amount) > 0.0001),
                    "applied_promotions": promotion_names,
                }
            )

        return output, used_promotion_ids, used_promotion_names

    def _promotion_applies_to_matrix_entry(self, promotion: dict, entry: dict, booking_date: date | None = None) -> bool:
        discount = self._coerce_float(promotion.get("discount_percent"))
        if discount is None or discount <= 0:
            return False

        room_filters = self._normalize_promotion_scope_values(promotion.get("applicable_room_types"))
        if room_filters:
            room_value = str(entry.get("room_type") or "").strip().lower()
            if room_value and not any(token in room_value for token in room_filters):
                return False

        board_filters = self._normalize_promotion_scope_values(promotion.get("applicable_board_types"))
        if board_filters:
            board_value = self._normalize_board_type(str(entry.get("board_type") or ""))
            board_lower = board_value.lower()
            if board_lower and not any(token in board_lower for token in board_filters):
                return False

        promo_start, promo_end, booking_start_date, booking_end_date = self._resolve_promotion_window_dates(promotion)
        entry_start = self._coerce_optional_date(entry.get("start_date"))
        entry_end = self._coerce_optional_date(entry.get("end_date"))

        if promo_start and promo_end and promo_end < promo_start:
            promo_start, promo_end = promo_end, promo_start
        if entry_start and entry_end and entry_end < entry_start:
            entry_start, entry_end = entry_end, entry_start

        if promo_start or promo_end:
            if entry_start and promo_end and entry_start > promo_end:
                return False
            if entry_end and promo_start and entry_end < promo_start:
                return False

        if booking_start_date or booking_end_date:
            if booking_date is None:
                return False
            if booking_start_date and booking_date < booking_start_date:
                return False
            if booking_end_date and booking_date > booking_end_date:
                return False

        return True

    def _normalize_promotion_scope_values(self, value: object) -> set[str]:
        if not isinstance(value, list):
            return set()
        normalized: set[str] = set()
        for item in value:
            if item is None:
                continue
            text = str(item).strip().lower()
            if not text:
                continue
            normalized.add(text)
            normalized.add(self._normalize_board_type(text).lower())
        return normalized

    def _extract_age_bucket_from_text(
        self,
        source_text: str,
        min_children: int,
        rule_type: str,
        guest_type: str | None = None,
    ) -> tuple[str, str, str]:
        lower = source_text.lower()
        guest_hint = (guest_type or "").strip().lower()
        if guest_hint in {"children", "kid", "kids"}:
            guest_hint = "child"
        elif guest_hint in {"adults", "extra adult", "extra_adult", "adult guest"}:
            guest_hint = "adult"

        if "infant" in lower or "baby" in lower:
            return "infant", "infant", "Infant"
        if "senior" in lower:
            return "senior", "senior", "Senior"

        ordinal_from_text = self._extract_ordinal_from_text(lower)
        if guest_hint == "child" or rule_type == "child_discount" or "child" in lower or "kid" in lower:
            child_index = min_children if min_children > 0 else ordinal_from_text
            if child_index <= 0:
                child_index = 1
            label = f"{self._format_ordinal(child_index)} Child"
            return "child", f"{child_index}_child", label

        if guest_hint == "adult" or rule_type == "extra_guest_adjustment" or "adult" in lower:
            if ordinal_from_text > 1:
                label = f"{self._format_ordinal(ordinal_from_text)} Adult"
                return "adult", f"{ordinal_from_text}_adult", label
            if rule_type == "extra_guest_adjustment":
                return "adult", "extra_adult", "Extra Adult"
            return "adult", "adult", "Adult"

        return "unknown", "other", "Other"

    def _extract_ordinal_from_text(self, text: str) -> int:
        match = re.search(r"\b(\d+)(?:st|nd|rd|th)\b", text, flags=re.IGNORECASE)
        if match:
            return self._coerce_int(match.group(1), default=0)

        word_map = {
            "first": 1,
            "second": 2,
            "third": 3,
            "fourth": 4,
            "fifth": 5,
        }
        for word, value in word_map.items():
            if re.search(rf"\b{word}\b", text, flags=re.IGNORECASE):
                return value
        return 0

    def _format_ordinal(self, value: int) -> str:
        if 10 <= value % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(value % 10, "th")
        return f"{value}{suffix}"

    def _extract_age_range_from_text(self, source_text: str) -> tuple[int | None, int | None]:
        if not source_text:
            return None, None
        match = re.search(r"\b(\d{1,2})\s*(?:-|to|–|—)\s*(\d{1,2})\b", source_text, flags=re.IGNORECASE)
        if match:
            left = self._coerce_int(match.group(1), default=0)
            right = self._coerce_int(match.group(2), default=0)
            if left <= 0 or right <= 0:
                return None, None
            return (left, right) if left <= right else (right, left)

        under_match = re.search(r"(?:under|up to)\s*(\d{1,2})", source_text, flags=re.IGNORECASE)
        if under_match:
            upper = self._coerce_int(under_match.group(1), default=0)
            if upper > 0:
                return 0, upper

        return None, None

    def _coerce_mixed_items_to_text(self, raw_value: object) -> list[str]:
        if not isinstance(raw_value, list):
            return []
        output: list[str] = []
        for item in raw_value:
            if item is None:
                continue
            if isinstance(item, dict):
                for key in ("name", "label", "description", "value", "text"):
                    candidate = item.get(key)
                    if candidate is None:
                        continue
                    text = str(candidate).strip()
                    if text:
                        output.append(text)
                        break
                else:
                    compact = " ".join(str(value).strip() for value in item.values() if value is not None)
                    if compact:
                        output.append(compact)
            else:
                text = str(item).strip()
                if text:
                    output.append(text)
        return output

    def _extract_rate_candidates(self, text: str) -> list[float]:
        raw_numbers = re.findall(r"\b\d{2,4}(?:\.\d{1,2})?\b", text)
        values: list[float] = []
        for raw in raw_numbers:
            value = float(raw)
            if 25 <= value <= 5000 and not (1900 <= value <= 2100):
                values.append(value)

        dedup: list[float] = []
        seen: set[float] = set()
        for value in values:
            rounded = round(value, 2)
            if rounded in seen:
                continue
            seen.add(rounded)
            dedup.append(rounded)
            if len(dedup) >= 20:
                break
        return dedup

    def _extract_child_discount_percent(self, extraction: dict) -> float:
        for line in extraction.get("discounts", []):
            match = re.search(r"(?:2nd|second)\s+child[^\d]*(\d{1,2})\s*%", line, flags=re.IGNORECASE)
            if match:
                return float(match.group(1))

        for line in extraction.get("discounts", []):
            match = re.search(r"(\d{1,2})\s*%", line)
            if match:
                return float(match.group(1))

        return 50.0

    def _build_rule_lookup(self, rules: list[dict]) -> tuple[dict[tuple[str, str], float], float]:
        base_lookup: dict[tuple[str, str], float] = {}
        child_discount_percent = 0.0

        for rule in rules:
            if not rule.get("is_active", True):
                continue
            if rule.get("rule_type") == "base_rate":
                metadata = rule.get("metadata", {})
                room = str(metadata.get("room_type", "")).strip().lower()
                board = str(metadata.get("board_type", "")).strip().lower()
                base_rate = float(metadata.get("base_rate") or 0)
                if room and board and base_rate > 0:
                    base_lookup[(room, board)] = base_rate
            elif rule.get("rule_type") == "child_discount":
                metadata = rule.get("metadata", {})
                child_discount_percent = float(metadata.get("discount_percent") or child_discount_percent)

        if child_discount_percent <= 0:
            child_discount_percent = 50.0

        return base_lookup, child_discount_percent

    def _build_validation_pricing_context(
        self,
        *,
        contract: dict,
        rules: list[dict],
        base_rate_lookup: dict[tuple[str, str], float],
    ) -> dict:
        ai_extracted_data = contract.get("ai_extracted_data")
        if not isinstance(ai_extracted_data, dict):
            ai_extracted_data = {}

        period_ranges = self._extract_price_period_ranges(ai_extracted_data)
        period_lookup = {
            str(period.get("label") or "").strip().lower(): period
            for period in period_ranges
            if str(period.get("label") or "").strip()
        }

        base_entries = self._extract_base_price_entries(
            rules=rules,
            ai_extracted_data=ai_extracted_data,
            period_lookup=period_lookup,
        )
        adjustments = self._extract_age_adjustments(rules=rules, ai_extracted_data=ai_extracted_data)
        derived_entries = self._derive_adjusted_price_entries(base_entries=base_entries, adjustments=adjustments)
        entries = [*base_entries, *derived_entries]
        board_supplement_entries = self._extract_board_supplement_entries(
            contract=contract,
            ai_extracted_data=ai_extracted_data,
            period_lookup=period_lookup,
        )

        entries_by_board: dict[str, list[dict]] = {"__all__": list(entries)}
        for entry in entries:
            board_value = self._normalize_board_type(str(entry.get("board_type") or "")).strip().lower()
            if board_value:
                entries_by_board.setdefault(board_value, []).append(entry)

        board_supplements_by_board: dict[str, list[dict]] = {}
        for entry in board_supplement_entries:
            target_board = self._normalize_board_type(str(entry.get("target_board_type") or "")).strip().lower()
            if not target_board:
                continue
            board_supplements_by_board.setdefault(target_board, []).append(entry)

        average_base_rate = self._average_rate_lookup(base_rate_lookup)
        if average_base_rate <= 0:
            adult_prices = [
                self._coerce_float(item.get("price")) or 0.0
                for item in base_entries
                if str(item.get("age_category") or "").strip().lower() == "adult"
            ]
            adult_prices = [value for value in adult_prices if value > 0]
            if adult_prices:
                average_base_rate = round(sum(adult_prices) / len(adult_prices), 2)

        return {
            "entries": entries,
            "entries_by_board": entries_by_board,
            "board_supplement_entries": board_supplement_entries,
            "board_supplements_by_board": board_supplements_by_board,
            "average_base_rate": average_base_rate,
        }

    def _extract_board_supplement_entries(
        self,
        *,
        contract: dict,
        ai_extracted_data: dict,
        period_lookup: dict[str, dict],
    ) -> list[dict]:
        entries: list[dict] = []
        seen_keys: set[tuple] = set()

        def add_entry(
            *,
            target_board_type: str | None,
            source_board_type: str | None = None,
            room_type: str | None = None,
            start_date: date | None = None,
            end_date: date | None = None,
            amount: float | None = None,
            adult_amount: float | None = None,
            additional_adult_amount: float | None = None,
            child_amount: float | None = None,
            baby_amount: float | None = None,
            per_person: bool = False,
            source_label: str | None = None,
            source_text: str | None = None,
        ) -> None:
            target_normalized = self._normalize_board_type(str(target_board_type or "")).strip().upper()
            if not target_normalized:
                return
            adult_value = self._coerce_float(adult_amount)
            if adult_value is None:
                adult_value = self._coerce_float(amount)
            if adult_value is None:
                return
            if abs(adult_value) > 500:
                return

            additional_value = self._coerce_float(additional_adult_amount)
            if additional_value is None:
                additional_value = adult_value
            if abs(additional_value) > 500:
                return

            child_value = self._coerce_float(child_amount)
            baby_value = self._coerce_float(baby_amount)

            if start_date and end_date and end_date < start_date:
                start_date, end_date = end_date, start_date

            source_board_normalized = self._normalize_board_type(str(source_board_type or "")).strip().upper() if source_board_type else None
            room_value = str(room_type or "").strip() or None
            key = (
                target_normalized,
                source_board_normalized or "",
                room_value or "",
                start_date.isoformat() if start_date else "",
                end_date.isoformat() if end_date else "",
                round(adult_value, 4),
                round(additional_value, 4),
                round(child_value, 4) if child_value is not None else None,
                round(baby_value, 4) if baby_value is not None else None,
                bool(per_person),
            )
            if key in seen_keys:
                return
            seen_keys.add(key)

            entries.append(
                {
                    "target_board_type": target_normalized,
                    "source_board_type": source_board_normalized,
                    "room_type": room_value,
                    "start_date": start_date,
                    "end_date": end_date,
                    "amount": round(adult_value, 2),
                    "adult_amount": round(adult_value, 2),
                    "additional_adult_amount": round(additional_value, 2),
                    "child_amount": round(child_value, 2) if child_value is not None else None,
                    "baby_amount": round(baby_value, 2) if baby_value is not None else None,
                    "per_person": bool(per_person),
                    "source_label": str(source_label or "").strip() or None,
                    "source_text": str(source_text or "").strip() or None,
                }
            )

        supplements = ai_extracted_data.get("supplements")
        if isinstance(supplements, list):
            for item in supplements[:120]:
                if isinstance(item, dict):
                    raw_name = str(item.get("name") or item.get("label") or "").strip()
                    raw_description = str(item.get("description") or item.get("text") or "").strip()
                    combined = " ".join(value for value in (raw_name, raw_description) if value).strip()
                    target_boards = self._extract_board_types_from_text(combined)
                    if not target_boards:
                        continue

                    period_label = str(item.get("period_label") or item.get("period") or "").strip() or None
                    fallback_start = self._coerce_optional_date(item.get("start_date") or item.get("from_date"))
                    fallback_end = self._coerce_optional_date(item.get("end_date") or item.get("to_date"))
                    start_date, end_date, _ = self._resolve_price_period(
                        period_label=period_label,
                        period_lookup=period_lookup,
                        fallback_start=fallback_start,
                        fallback_end=fallback_end,
                    )
                    source_board_candidates = self._extract_board_types_from_text(
                        str(item.get("base_board_type") or item.get("source_board_type") or "")
                    )
                    source_board = source_board_candidates[0] if source_board_candidates else None
                    add_entry(
                        target_board_type=target_boards[0],
                        source_board_type=source_board,
                        room_type=str(item.get("room_type") or "").strip() or None,
                        start_date=start_date,
                        end_date=end_date,
                        amount=self._coerce_float(item.get("amount")),
                        adult_amount=self._coerce_float(item.get("adult_amount") or item.get("adult_price")),
                        additional_adult_amount=self._coerce_float(item.get("additional_adult_amount") or item.get("additional_amount")),
                        child_amount=self._coerce_float(item.get("child_amount")),
                        baby_amount=self._coerce_float(item.get("baby_amount") or item.get("infant_amount")),
                        per_person=bool(item.get("per_person", False)),
                        source_label=f"{target_boards[0]} supplement",
                        source_text=combined,
                    )
                    continue

                if not isinstance(item, str):
                    continue
                text = item.strip()
                if not text:
                    continue
                target_boards = self._extract_board_types_from_text(text)
                if not target_boards:
                    continue
                fixed_values = self._extract_fixed_amounts_from_text(text)
                if not fixed_values:
                    continue
                add_entry(
                    target_board_type=target_boards[0],
                    amount=fixed_values[0],
                    adult_amount=fixed_values[0],
                    additional_adult_amount=fixed_values[1] if len(fixed_values) > 1 else fixed_values[0],
                    per_person=len(fixed_values) > 1,
                    source_label=f"{target_boards[0]} supplement",
                    source_text=text,
                )

        parsed_text = str(contract.get("parsed_text_preview") or "").strip()
        if parsed_text:
            lines = [line.strip() for line in parsed_text.splitlines() if line and line.strip()]
            active_target_board: str | None = None
            per_person_hint = False
            for line in lines:
                lower = line.lower()
                if "adult" in lower and "additional" in lower and "children" in lower:
                    per_person_hint = True
                boards_in_line = self._extract_board_types_from_text(line)
                line_has_numbers = bool(re.search(r"\d", line))
                line_has_fixed = "fixed" in lower

                if boards_in_line and not line_has_numbers and not line_has_fixed and "board types" not in lower and "applies" not in lower:
                    active_target_board = boards_in_line[0]
                    continue

                if not line_has_fixed:
                    continue

                fixed_values = self._extract_fixed_amounts_from_text(line)
                if not fixed_values:
                    continue

                target_board = active_target_board
                if not target_board and boards_in_line:
                    target_board = boards_in_line[0]
                if not target_board:
                    continue

                source_board = None
                for board_code in boards_in_line:
                    if board_code != target_board:
                        source_board = board_code
                        break

                start_date, end_date = self._extract_date_range_from_text(line)
                adult_amount = fixed_values[0]
                additional_amount = fixed_values[1] if len(fixed_values) > 1 else fixed_values[0]
                child_amount = None
                baby_amount = None
                if len(fixed_values) >= 4:
                    child_amount = fixed_values[2]
                    baby_amount = fixed_values[3]
                elif len(fixed_values) >= 3 and "babies" in lower:
                    baby_amount = fixed_values[2]
                elif len(fixed_values) >= 3:
                    child_amount = fixed_values[2]

                add_entry(
                    target_board_type=target_board,
                    source_board_type=source_board,
                    start_date=start_date,
                    end_date=end_date,
                    amount=adult_amount,
                    adult_amount=adult_amount,
                    additional_adult_amount=additional_amount,
                    child_amount=child_amount,
                    baby_amount=baby_amount,
                    per_person=per_person_hint or len(fixed_values) > 1,
                    source_label=f"{target_board} supplement",
                    source_text=line,
                )

        return entries

    def _extract_board_types_from_text(self, text: str) -> list[str]:
        if not text:
            return []
        normalized = str(text).strip().lower()
        if not normalized:
            return []

        patterns: list[tuple[str, str]] = [
            (r"\bfull\s+board(?:\s+plus)?\b|\bfb\+?\b", "FB"),
            (r"\bhalf\s+board(?:\s+plus)?\b|\bhb\+?\b", "HB"),
            (r"\bbed\s*(?:&|and)?\s*breakfast\b|\bbed\s+breakfast\b|\bbb\+?\b", "BB"),
            (r"\broom\s+only\b|\bro\+?\b", "RO"),
            (r"\ball\s+inclusive\b|\bai\+?\b", "AI"),
        ]

        detected: list[str] = []
        for pattern, code in patterns:
            if re.search(pattern, normalized, flags=re.IGNORECASE):
                detected.append(code)
        return detected

    def _extract_date_range_from_text(self, text: str) -> tuple[date | None, date | None]:
        if not text:
            return None, None
        match = re.search(
            r"(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\s*(?:-|–|—|to)\s*(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})",
            text,
            flags=re.IGNORECASE,
        )
        if not match:
            return None, None
        start_date = self._coerce_optional_date(match.group(1))
        end_date = self._coerce_optional_date(match.group(2))
        if start_date and end_date and end_date < start_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date

    def _extract_fixed_amounts_from_text(self, text: str) -> list[float]:
        if not text:
            return []

        values: list[float] = []
        for match in re.finditer(r"fixed\s*:\s*([+-]?\d+(?:[.,]\d{1,2})?)", text, flags=re.IGNORECASE):
            raw_value = str(match.group(1) or "").strip().replace(",", ".")
            try:
                values.append(float(raw_value))
            except ValueError:
                continue

        if values:
            return values

        for match in re.finditer(r"([+-]?\d+(?:[.,]\d{1,2})?)\s*€", text, flags=re.IGNORECASE):
            raw_value = str(match.group(1) or "").strip().replace(",", ".")
            try:
                values.append(float(raw_value))
            except ValueError:
                continue
        return values

    def _select_board_supplement_for_line(
        self,
        *,
        pricing_context: dict,
        requested_board_type: str,
        stay_date: date,
        room_type: str,
        base_board_type: str | None = None,
    ) -> dict | None:
        supplements_by_board = pricing_context.get("board_supplements_by_board")
        if not isinstance(supplements_by_board, dict):
            return None

        requested_board = self._normalize_board_type(str(requested_board_type or "")).strip().lower()
        if not requested_board:
            return None

        pool = supplements_by_board.get(requested_board)
        if not isinstance(pool, list) or not pool:
            return None

        normalized_base_board = self._normalize_board_type(str(base_board_type or "")).strip().lower()
        best: tuple[int, dict] | None = None

        for entry in pool:
            if not isinstance(entry, dict):
                continue
            amount = self._coerce_float(entry.get("adult_amount"))
            if amount is None:
                amount = self._coerce_float(entry.get("amount"))
            if amount is None:
                continue

            date_match, date_specificity = self._entry_matches_stay_date(entry, stay_date)
            if not date_match:
                continue

            room_score = self._room_match_score(room_type, entry.get("room_type"))
            source_board = self._normalize_board_type(str(entry.get("source_board_type") or "")).strip().lower()
            board_score = 0
            if normalized_base_board and source_board:
                board_score = 4 if source_board == normalized_base_board else -2
            elif normalized_base_board and not source_board:
                board_score = 1

            room_hint_bonus = 1 if str(entry.get("room_type") or "").strip() else 0
            score = (date_specificity * 5) + (room_score * 4) + board_score + room_hint_bonus
            if not best or score > best[0]:
                best = (score, entry)

        return best[1] if best else None

    def _compute_board_supplement_amount(
        self,
        *,
        supplement_entry: dict,
        pax_adults: int,
        pax_children: int,
    ) -> float | None:
        adult_amount = self._coerce_float(supplement_entry.get("adult_amount"))
        if adult_amount is None:
            adult_amount = self._coerce_float(supplement_entry.get("amount"))
        if adult_amount is None:
            return None

        additional_adult_amount = self._coerce_float(supplement_entry.get("additional_adult_amount"))
        if additional_adult_amount is None:
            additional_adult_amount = adult_amount

        child_amount = self._coerce_float(supplement_entry.get("child_amount"))
        per_person = bool(supplement_entry.get("per_person", False))
        adults = max(1, int(pax_adults or 1))
        children = max(0, int(pax_children or 0))

        if not per_person:
            return round(adult_amount, 2)

        included_adults = min(adults, 2)
        additional_adults = max(0, adults - 2)
        total = (adult_amount * included_adults) + (additional_adult_amount * additional_adults)
        if child_amount is not None and child_amount != 0:
            total += child_amount * children
        return round(total, 2)

    def _room_match_score(self, left: object, right: object) -> int:
        left_token = self._normalize_room_match_token(left)
        right_token = self._normalize_room_match_token(right)
        if not left_token or not right_token:
            return 0
        if left_token == right_token:
            return 10
        if left_token in right_token or right_token in left_token:
            return 8

        left_parts = set(left_token.split())
        right_parts = set(right_token.split())
        overlap = len(left_parts & right_parts)
        if overlap <= 0:
            return 0
        if overlap >= 3:
            return 7
        if overlap == 2:
            return 5
        return 3

    def _normalize_room_match_token(self, value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
        text = re.sub(r"([A-Z])([A-Z][a-z])", r"\1 \2", text)
        text = re.sub(r"([A-Za-z])(\d)", r"\1 \2", text)
        text = re.sub(r"(\d)([A-Za-z])", r"\1 \2", text)
        token = self._normalize_header_token(text)
        if not token:
            return ""

        for pattern, replacement in ROOM_MATCH_CANONICAL_REPLACEMENTS:
            token = pattern.sub(replacement, token)
        token = self._normalize_header_token(token)
        if not token:
            return ""

        compact_parts = [part for part in token.split() if part and part not in ROOM_MATCH_STOPWORDS]
        if compact_parts:
            token = " ".join(compact_parts)
        return token

    def _entry_matches_stay_date(self, entry: dict, stay_date: date) -> tuple[bool, int]:
        start_date = self._coerce_optional_date(entry.get("start_date"))
        end_date = self._coerce_optional_date(entry.get("end_date"))
        if start_date and end_date and end_date < start_date:
            start_date, end_date = end_date, start_date

        if start_date and stay_date < start_date:
            return False, 0
        if end_date and stay_date > end_date:
            return False, 0

        if start_date and end_date:
            return True, 2
        if start_date or end_date:
            return True, 1
        return True, 0

    def _select_pricing_entry_for_line(
        self,
        *,
        pricing_context: dict,
        room_type: str,
        board_type: str,
        stay_date: date,
        age_bucket_candidates: list[str],
        age_category: str | None = None,
    ) -> dict | None:
        entries_by_board = pricing_context.get("entries_by_board")
        if not isinstance(entries_by_board, dict):
            return None

        board_raw = str(board_type or "").strip()
        board_token = self._normalize_board_type(board_raw).strip().lower() if board_raw else ""
        bucket_rank = {
            str(bucket).strip().lower(): rank
            for rank, bucket in enumerate(age_bucket_candidates)
            if str(bucket).strip()
        }

        pool: list[dict] = []
        if board_token and isinstance(entries_by_board.get(board_token), list):
            pool.extend(entries_by_board[board_token])
        if isinstance(entries_by_board.get("__all__"), list):
            pool.extend(entries_by_board["__all__"])
        if not pool:
            return None

        deduped: list[dict] = []
        seen_ids: set[int] = set()
        for item in pool:
            marker = id(item)
            if marker in seen_ids:
                continue
            seen_ids.add(marker)
            deduped.append(item)

        best_exact: tuple[int, dict] | None = None
        best_fallback: tuple[int, dict] | None = None

        for entry in deduped:
            if not isinstance(entry, dict):
                continue

            entry_age_category = str(entry.get("age_category") or "").strip().lower()
            if age_category and entry_age_category and entry_age_category != age_category:
                continue

            entry_bucket = str(entry.get("age_bucket") or "").strip().lower()
            if bucket_rank:
                if entry_bucket not in bucket_rank:
                    continue
                rank_value = bucket_rank[entry_bucket]
            else:
                rank_value = 0

            entry_board = self._normalize_board_type(str(entry.get("board_type") or "")).strip().lower()
            board_penalty = 0
            if board_token and entry_board and entry_board != board_token:
                continue
            if board_token and not entry_board:
                board_penalty = 3

            date_match, date_specificity = self._entry_matches_stay_date(entry, stay_date)
            if not date_match:
                continue

            room_score = self._room_match_score(room_type, entry.get("room_type"))
            has_period_label = 1 if str(entry.get("period_label") or "").strip() else 0
            score = (
                (100 - (rank_value * 10))
                + (room_score * 6)
                + (date_specificity * 5)
                + has_period_label
                - board_penalty
            )

            if room_score > 0:
                if not best_exact or score > best_exact[0]:
                    best_exact = (score, entry)
            else:
                if not best_fallback or score > best_fallback[0]:
                    best_fallback = (score, entry)

        if best_exact:
            return best_exact[1]
        if best_fallback and not self._normalize_header_token(room_type):
            return best_fallback[1]
        return None

    def _is_early_booking_offer(
        self,
        *,
        offer_name: str | None = None,
        description: str | None = None,
        expression: str | None = None,
    ) -> bool:
        haystack = " ".join(
            value
            for value in [
                str(offer_name or "").strip().lower(),
                str(description or "").strip().lower(),
                str(expression or "").strip().lower(),
            ]
            if value
        )
        if not haystack:
            return False
        return any(
            marker in haystack
            for marker in (
                "early booking",
                "early-booking",
                "earlybooking",
                "early bird",
                "early-bird",
                "earlybird",
            )
        )

    def _has_promotion_booking_and_arrival_windows(self, promotion: dict | None) -> bool:
        if not isinstance(promotion, dict):
            return False
        metadata = promotion.get("metadata")
        metadata_map = metadata if isinstance(metadata, dict) else {}

        booking_start = self._coerce_optional_date(
            promotion.get("booking_start_date")
            or metadata_map.get("booking_start_date")
        )
        booking_end = self._coerce_optional_date(
            promotion.get("booking_end_date")
            or metadata_map.get("booking_end_date")
        )
        arrival_start = self._coerce_optional_date(
            promotion.get("arrival_start_date")
            or promotion.get("start_date")
            or metadata_map.get("arrival_start_date")
            or metadata_map.get("start_date")
        )
        arrival_end = self._coerce_optional_date(
            promotion.get("arrival_end_date")
            or promotion.get("end_date")
            or metadata_map.get("arrival_end_date")
            or metadata_map.get("end_date")
        )
        return bool((booking_start or booking_end) and (arrival_start or arrival_end))

    def _is_early_booking_promotion(self, promotion: dict | None) -> bool:
        if not isinstance(promotion, dict):
            return False
        metadata = promotion.get("metadata")
        metadata_map = metadata if isinstance(metadata, dict) else {}

        category = str(
            promotion.get("promotion_category")
            or metadata_map.get("promotion_category")
            or metadata_map.get("category")
            or ""
        ).strip().lower()
        if category in {"early_booking", "early-booking", "early booking", "earlybird", "early_bird"}:
            return True
        if self._has_promotion_booking_and_arrival_windows(promotion):
            return True

        return self._is_early_booking_offer(
            offer_name=str(promotion.get("offer_name") or promotion.get("name") or metadata_map.get("offer_name") or ""),
            description=str(promotion.get("description") or metadata_map.get("description") or ""),
            expression=str(promotion.get("expression") or ""),
        )

    def _resolve_promotion_window_dates(self, promotion: dict) -> tuple[date | None, date | None, date | None, date | None]:
        arrival_start_date = self._coerce_optional_date(promotion.get("arrival_start_date") or promotion.get("start_date"))
        arrival_end_date = self._coerce_optional_date(promotion.get("arrival_end_date") or promotion.get("end_date"))
        booking_start_date = self._coerce_optional_date(promotion.get("booking_start_date"))
        booking_end_date = self._coerce_optional_date(promotion.get("booking_end_date"))

        if arrival_start_date and arrival_end_date and arrival_end_date < arrival_start_date:
            arrival_start_date, arrival_end_date = arrival_end_date, arrival_start_date
        if booking_start_date and booking_end_date and booking_end_date < booking_start_date:
            booking_start_date, booking_end_date = booking_end_date, booking_start_date

        if self._is_early_booking_promotion(promotion):
            # Early-booking must be constrained by BOTH booking and stay windows.
            # If source data provides only one range, mirror it to preserve deterministic behavior.
            if not (booking_start_date or booking_end_date) and (arrival_start_date or arrival_end_date):
                booking_start_date = arrival_start_date
                booking_end_date = arrival_end_date
            if not (arrival_start_date or arrival_end_date) and (booking_start_date or booking_end_date):
                arrival_start_date = booking_start_date
                arrival_end_date = booking_end_date

        return arrival_start_date, arrival_end_date, booking_start_date, booking_end_date

    def _collect_effective_promotions(self, *, promotions: list[dict], rules: list[dict]) -> list[dict]:
        effective: list[dict] = []
        seen_ids: set[str] = set()
        seen_fingerprints: set[tuple] = set()

        def add(candidate: dict | None) -> None:
            normalized = self._normalize_effective_promotion(candidate)
            if not normalized:
                return

            promotion_id = str(normalized.get("id") or "").strip()
            fingerprint = self._promotion_fingerprint(normalized)
            if promotion_id and promotion_id in seen_ids:
                return
            if fingerprint in seen_fingerprints:
                return

            if promotion_id:
                seen_ids.add(promotion_id)
            seen_fingerprints.add(fingerprint)
            effective.append(normalized)

        for promotion in promotions:
            add(promotion)
        for rule in rules:
            add(self._promotion_from_rule(rule))

        return effective

    def _normalize_effective_promotion(self, promotion: dict | None) -> dict | None:
        if not isinstance(promotion, dict):
            return None
        discount = self._coerce_float(promotion.get("discount_percent"))
        if discount is None or discount <= 0:
            return None

        offer_name = str(promotion.get("offer_name") or promotion.get("name") or "").strip() or "Promotion"
        arrival_start_date, arrival_end_date, booking_start_date, booking_end_date = self._resolve_promotion_window_dates(promotion)
        is_early_booking = self._is_early_booking_promotion(promotion)
        return {
            **promotion,
            "offer_name": offer_name,
            "discount_percent": round(discount, 2),
            "description": str(promotion.get("description") or "").strip() or None,
            "start_date": arrival_start_date,
            "end_date": arrival_end_date,
            "booking_start_date": booking_start_date,
            "booking_end_date": booking_end_date,
            "arrival_start_date": arrival_start_date,
            "arrival_end_date": arrival_end_date,
            "applicable_room_types": promotion.get("applicable_room_types") or [],
            "applicable_board_types": promotion.get("applicable_board_types") or [],
            "scope": str(promotion.get("scope") or "all").strip() or "all",
            "non_cumulative": bool(promotion.get("non_cumulative", False)),
            "promotion_category": "early_booking" if is_early_booking else "general",
        }

    def _promotion_from_rule(self, rule: dict) -> dict | None:
        if not isinstance(rule, dict):
            return None
        if str(rule.get("rule_type") or "").strip().lower() != "promotion":
            return None
        if not bool(rule.get("is_active", True)):
            return None

        metadata = rule.get("metadata")
        if not isinstance(metadata, dict):
            metadata = {}

        discount = self._coerce_float(metadata.get("discount_percent"))
        if discount is None:
            expression = str(rule.get("expression") or "")
            match = re.search(r"(\d{1,3}(?:\.\d{1,2})?)\s*%", expression)
            if match:
                discount = self._coerce_float(match.group(1))
        if discount is None or discount <= 0:
            return None

        rule_id = str(rule.get("id") or rule.get("_id") or "").strip()
        promotion_id = str(metadata.get("promotion_id") or "").strip()
        offer_name = str(metadata.get("offer_name") or rule.get("name") or "").strip() or "Promotion"

        return {
            "id": promotion_id or (f"rule:{rule_id}" if rule_id else None),
            "offer_name": offer_name,
            "description": str(metadata.get("description") or "").strip() or None,
            "expression": str(rule.get("expression") or "").strip() or None,
            "discount_percent": round(discount, 2),
            "start_date": metadata.get("start_date"),
            "end_date": metadata.get("end_date"),
            "booking_start_date": metadata.get("booking_start_date"),
            "booking_end_date": metadata.get("booking_end_date"),
            "arrival_start_date": metadata.get("arrival_start_date"),
            "arrival_end_date": metadata.get("arrival_end_date"),
            "scope": metadata.get("scope"),
            "non_cumulative": bool(metadata.get("non_cumulative", False)),
            "applicable_room_types": metadata.get("applicable_room_types", []),
            "applicable_board_types": metadata.get("applicable_board_types", []),
            "promotion_category": metadata.get("promotion_category"),
        }

    def _promotion_fingerprint(self, promotion: dict) -> tuple:
        room_filters = tuple(sorted(self._normalize_promotion_scope_values(promotion.get("applicable_room_types"))))
        board_filters = tuple(sorted(self._normalize_promotion_scope_values(promotion.get("applicable_board_types"))))
        return (
            str(promotion.get("offer_name") or "").strip().lower(),
            round(self._coerce_float(promotion.get("discount_percent")) or 0.0, 4),
            str(promotion.get("booking_start_date") or ""),
            str(promotion.get("booking_end_date") or ""),
            str(promotion.get("arrival_start_date") or ""),
            str(promotion.get("arrival_end_date") or ""),
            str(promotion.get("start_date") or ""),
            str(promotion.get("end_date") or ""),
            bool(promotion.get("non_cumulative", False)),
            room_filters,
            board_filters,
        )

    def _validate_single_line(
        self,
        line_data: dict,
        base_rate_lookup: dict[tuple[str, str], float],
        pricing_context: dict,
        child_discount_percent: float,
        promotions: list[dict],
        tolerance_amount: float,
        tolerance_percent: float,
    ) -> dict:
        room_key = str(line_data.get("room_type", "")).strip().lower()
        board_key = str(line_data.get("board_type", "")).strip().lower()
        nights = max(1, int(line_data.get("nights") or 1))
        pax_children = int(line_data.get("pax_children") or 0)
        pax_adults = max(1, int(line_data.get("pax_adults") or 2))
        actual_price = float(line_data.get("actual_price") or 0)
        stay_date = self._coerce_date(line_data.get("stay_date"))
        booking_date = self._coerce_optional_date(
            line_data.get("booking_date")
            or line_data.get("reservation_date")
            or line_data.get("booking_created_at")
        )

        nightly_base_rates: list[dict] = []
        missing_base_dates: list[str] = []
        board_missing_dates: list[str] = []
        available_board_types: set[str] = set()
        expected_running = 0.0
        base_source_counts: dict[str, int] = {}
        board_supplements_agg: dict[str, dict] = {}
        applied_rules = []
        base_adult_units = max(1, min(pax_adults, 2))

        for offset in range(nights):
            current_date = stay_date + timedelta(days=offset)
            entry = self._select_pricing_entry_for_line(
                pricing_context=pricing_context,
                room_type=str(line_data.get("room_type") or ""),
                board_type=str(line_data.get("board_type") or ""),
                stay_date=current_date,
                age_bucket_candidates=["adult"],
                age_category="adult",
            )

            source_key = "contract_pricelist_rule"
            source_label = "Contract price list"
            period_label = None
            source_rule_name = None
            nightly_base_components: dict[str, object] = {}
            matched_pricelist_room_type: str | None = None
            matched_pricelist_board_type: str | None = None
            base_adult_unit_rate = self._coerce_float(entry.get("price")) if isinstance(entry, dict) else None
            supplement_amount_total = 0.0
            if isinstance(entry, dict):
                matched_pricelist_room_type = str(entry.get("room_type") or "").strip() or None
                matched_pricelist_board_type = (
                    self._normalize_board_type(str(entry.get("board_type") or "")).strip() or None
                )
            if base_adult_unit_rate is None or base_adult_unit_rate <= 0:
                board_fallback_entry = self._select_pricing_entry_for_line(
                    pricing_context=pricing_context,
                    room_type=str(line_data.get("room_type") or ""),
                    board_type="",
                    stay_date=current_date,
                    age_bucket_candidates=["adult"],
                    age_category="adult",
                )
                if isinstance(board_fallback_entry, dict):
                    matched_pricelist_room_type = str(board_fallback_entry.get("room_type") or "").strip() or None
                    matched_pricelist_board_type = (
                        self._normalize_board_type(str(board_fallback_entry.get("board_type") or "")).strip() or None
                    )
                    candidate_board = self._normalize_board_type(str(board_fallback_entry.get("board_type") or "")).strip()
                    if candidate_board:
                        available_board_types.add(candidate_board)
                    fallback_rate = self._coerce_float(board_fallback_entry.get("price"))
                    supplement_entry = self._select_board_supplement_for_line(
                        pricing_context=pricing_context,
                        requested_board_type=str(line_data.get("board_type") or ""),
                        stay_date=current_date,
                        room_type=str(line_data.get("room_type") or ""),
                        base_board_type=str(board_fallback_entry.get("board_type") or ""),
                    )
                    supplement_amount = (
                        self._compute_board_supplement_amount(
                            supplement_entry=supplement_entry,
                            pax_adults=pax_adults,
                            pax_children=pax_children,
                        )
                        if isinstance(supplement_entry, dict)
                        else None
                    )
                    if fallback_rate is not None and fallback_rate > 0 and supplement_amount is not None:
                        base_adult_unit_rate = fallback_rate
                        supplement_amount_total = supplement_amount
                        source_key = "contract_pricelist_with_board_supplement"
                        requested_board = self._normalize_board_type(str(line_data.get("board_type") or "")).strip() or "UNKNOWN"
                        source_label = f"Contract price list + {requested_board} supplement"
                        period_label = str(board_fallback_entry.get("period_label") or "").strip() or None
                        source_rule_name = str(board_fallback_entry.get("source_rule_name") or "").strip() or None

                        supplement_rule_name = str(supplement_entry.get("source_label") or source_label).strip() or source_label
                        source_board = self._normalize_board_type(str(board_fallback_entry.get("board_type") or "")).strip() or None
                        board_key = (
                            f"{requested_board}::{source_board or '-'}::{supplement_rule_name}::{round(supplement_amount, 2):.2f}"
                        )
                        agg = board_supplements_agg.get(board_key)
                        if not agg:
                            agg = {
                                "target_board_type": requested_board,
                                "source_board_type": source_board,
                                "rule_name": supplement_rule_name,
                                "unit_amount": round(supplement_amount, 2),
                                "nights_applied": 0,
                                "subtotal_amount": 0.0,
                                "dates": set(),
                            }
                            board_supplements_agg[board_key] = agg
                        agg["nights_applied"] += 1
                        agg["subtotal_amount"] += supplement_amount_total
                        agg["dates"].add(current_date.isoformat())

                        nightly_base_components = {
                            "board_supplement_amount": round(supplement_amount_total, 2),
                            "base_board_type": source_board,
                            "supplement_rule_name": supplement_rule_name,
                            "supplement_source_text": str(supplement_entry.get("source_text") or "").strip() or None,
                        }

                if base_adult_unit_rate is None or base_adult_unit_rate <= 0:
                    missing_base_dates.append(current_date.isoformat())
                    if isinstance(board_fallback_entry, dict):
                        board_missing_dates.append(current_date.isoformat())
                    nightly_base_rates.append(
                        {
                            "date": current_date.isoformat(),
                            "rate": 0.0,
                            "source": "Missing price-list rate",
                            "source_key": "missing_pricelist_rate",
                            "period_label": None,
                            "source_rule_name": None,
                            "pricelist_room_type": matched_pricelist_room_type,
                            "pricelist_board_type": matched_pricelist_board_type,
                        }
                    )
                    continue

            if isinstance(entry, dict) and (period_label is None and source_rule_name is None):
                period_label = str(entry.get("period_label") or "").strip() or None
                source_rule_name = str(entry.get("source_rule_name") or "").strip() or None

            nightly_rate = round((base_adult_unit_rate * base_adult_units) + supplement_amount_total, 2)
            nightly_base_components.setdefault("base_adult_unit_rate", round(base_adult_unit_rate, 2))
            nightly_base_components.setdefault("base_adult_units", base_adult_units)
            nightly_base_components.setdefault(
                "base_adult_subtotal",
                round(base_adult_unit_rate * base_adult_units, 2),
            )
            expected_running += nightly_rate
            base_source_counts[source_key] = base_source_counts.get(source_key, 0) + 1
            nightly_base_rates.append(
                {
                    "date": current_date.isoformat(),
                    "rate": round(nightly_rate, 2),
                    "source": source_label,
                    "source_key": source_key,
                    "period_label": period_label,
                    "source_rule_name": source_rule_name,
                    "pricelist_room_type": matched_pricelist_room_type,
                    "pricelist_board_type": matched_pricelist_board_type,
                    **nightly_base_components,
                }
            )

        base_subtotal = round(expected_running, 2)
        average_base_rate = round(base_subtotal / nights, 2) if nights > 0 else 0.0
        applied_rules.append(f"base_rate:{average_base_rate:.2f}*{nights}")

        guest_adjustments_agg: dict[str, dict] = {}
        for offset in range(nights):
            current_date = stay_date + timedelta(days=offset)

            for child_position in range(1, pax_children + 1):
                child_entry = self._select_pricing_entry_for_line(
                    pricing_context=pricing_context,
                    room_type=str(line_data.get("room_type") or ""),
                    board_type=str(line_data.get("board_type") or ""),
                    stay_date=current_date,
                    age_bucket_candidates=[f"{child_position}_child", "child"],
                    age_category="child",
                )
                child_rate = self._coerce_float(child_entry.get("price")) if isinstance(child_entry, dict) else None
                if child_rate is None or child_rate <= 0:
                    continue

                expected_running += child_rate
                age_bucket = str(child_entry.get("age_bucket") or "child").strip() or "child"
                rule_name = str(child_entry.get("source_rule_name") or "Child pricing rule").strip() or "Child pricing rule"
                key = f"child::{age_bucket}::{rule_name}"
                bucket = guest_adjustments_agg.get(key)
                if not bucket:
                    bucket = {
                        "guest_type": "child",
                        "age_bucket": age_bucket,
                        "rule_name": rule_name,
                        "unit_rate": round(child_rate, 2),
                        "units": 0,
                        "subtotal_amount": 0.0,
                        "guest_positions": set(),
                        "dates": set(),
                    }
                    guest_adjustments_agg[key] = bucket
                bucket["units"] += 1
                bucket["subtotal_amount"] += child_rate
                bucket["guest_positions"].add(child_position)
                bucket["dates"].add(current_date.isoformat())

            for adult_position in range(3, pax_adults + 1):
                adult_entry = self._select_pricing_entry_for_line(
                    pricing_context=pricing_context,
                    room_type=str(line_data.get("room_type") or ""),
                    board_type=str(line_data.get("board_type") or ""),
                    stay_date=current_date,
                    age_bucket_candidates=[f"{adult_position}_adult", "extra_adult"],
                    age_category="adult",
                )
                adult_rate = self._coerce_float(adult_entry.get("price")) if isinstance(adult_entry, dict) else None
                if adult_rate is None or adult_rate <= 0:
                    continue

                expected_running += adult_rate
                age_bucket = str(adult_entry.get("age_bucket") or "extra_adult").strip() or "extra_adult"
                rule_name = str(adult_entry.get("source_rule_name") or "Extra adult rule").strip() or "Extra adult rule"
                key = f"adult::{age_bucket}::{rule_name}"
                bucket = guest_adjustments_agg.get(key)
                if not bucket:
                    bucket = {
                        "guest_type": "adult_extra",
                        "age_bucket": age_bucket,
                        "rule_name": rule_name,
                        "unit_rate": round(adult_rate, 2),
                        "units": 0,
                        "subtotal_amount": 0.0,
                        "guest_positions": set(),
                        "dates": set(),
                    }
                    guest_adjustments_agg[key] = bucket
                bucket["units"] += 1
                bucket["subtotal_amount"] += adult_rate
                bucket["guest_positions"].add(adult_position)
                bucket["dates"].add(current_date.isoformat())

        guest_adjustments: list[dict] = []
        for item in guest_adjustments_agg.values():
            subtotal_amount = round(float(item["subtotal_amount"]), 2)
            guest_adjustments.append(
                {
                    "guest_type": item["guest_type"],
                    "age_bucket": item["age_bucket"],
                    "rule_name": item["rule_name"],
                    "unit_rate": item["unit_rate"],
                    "units": int(item["units"]),
                    "subtotal_amount": subtotal_amount,
                    "guest_positions": sorted(int(value) for value in item["guest_positions"]),
                    "dates_count": len(item["dates"]),
                }
            )

        guest_adjustments.sort(
            key=lambda row: (
                str(row.get("guest_type") or ""),
                str(row.get("age_bucket") or ""),
                str(row.get("rule_name") or ""),
            )
        )
        for row in guest_adjustments:
            applied_rules.append(
                f"{row['guest_type']}:{row['age_bucket']}:{row['unit_rate']:.2f}x{row['units']}"
            )

        board_supplement_adjustments: list[dict] = []
        for item in board_supplements_agg.values():
            nights_applied = int(item.get("nights_applied") or 0)
            subtotal_amount = round(float(item.get("subtotal_amount") or 0.0), 2)
            board_supplement_adjustments.append(
                {
                    "target_board_type": str(item.get("target_board_type") or ""),
                    "source_board_type": str(item.get("source_board_type") or "") or None,
                    "rule_name": str(item.get("rule_name") or "Board supplement"),
                    "unit_amount": round(float(item.get("unit_amount") or 0.0), 2),
                    "nights_applied": nights_applied,
                    "subtotal_amount": subtotal_amount,
                    "dates_count": len(item.get("dates") or []),
                }
            )
        board_supplement_adjustments.sort(
            key=lambda row: (
                str(row.get("target_board_type") or ""),
                str(row.get("source_board_type") or ""),
                str(row.get("rule_name") or ""),
            )
        )
        for row in board_supplement_adjustments:
            source_board = str(row.get("source_board_type") or "-")
            applied_rules.append(
                f"board_supplement:{source_board}->{row['target_board_type']}:{row['unit_amount']:.2f}x{row['nights_applied']}"
            )

        child_adjustment_total = round(
            sum(
                float(row.get("subtotal_amount") or 0.0)
                for row in guest_adjustments
                if str(row.get("guest_type") or "").strip() == "child"
            ),
            2,
        )
        child_adjustment_percent = round((child_adjustment_total / base_subtotal * 100.0), 2) if base_subtotal > 0 else 0.0
        child_discount_breakdown = {
            "applied": child_adjustment_total > 0,
            "children_count": pax_children,
            "discount_percent": child_adjustment_percent if child_adjustment_total > 0 else round(child_discount_percent, 2),
            "multiplier": round((1 + child_adjustment_total / base_subtotal), 4) if base_subtotal > 0 else 1.0,
            "subtotal_before": base_subtotal,
            "adjustment_amount": child_adjustment_total,
            "subtotal_after": round(base_subtotal + child_adjustment_total, 2),
        }

        applied_promotions: list[str] = []
        promotion_breakdown: list[dict] = []
        applicable_promotions: list[dict] = []
        promotions_require_booking_date = False
        booking_date_missing_for_promotion_check = False
        for promo in promotions:
            discount = float(promo.get("discount_percent") or 0)
            if discount <= 0:
                continue
            _, _, booking_start_date, booking_end_date = self._resolve_promotion_window_dates(promo)
            requires_booking_date = bool(booking_start_date or booking_end_date)
            if requires_booking_date:
                promotions_require_booking_date = True
                if booking_date is None:
                    booking_date_missing_for_promotion_check = True
            if self._promotion_applies(
                promo,
                stay_date,
                line_data.get("promo_code"),
                booking_date=booking_date,
                room_type=line_data.get("room_type"),
                board_type=line_data.get("board_type"),
            ):
                applicable_promotions.append(promo)
        if missing_base_dates:
            applicable_promotions = []

        promotions_to_apply: list[dict] = []
        if applicable_promotions:
            non_cumulative_promos = [promo for promo in applicable_promotions if bool(promo.get("non_cumulative", False))]
            if non_cumulative_promos:
                best = max(
                    non_cumulative_promos,
                    key=lambda promo: self._coerce_float(promo.get("discount_percent")) or 0.0,
                )
                promotions_to_apply = [best]
            else:
                promotions_to_apply = applicable_promotions

        for promo in promotions_to_apply:
            discount = self._coerce_float(promo.get("discount_percent")) or 0.0
            if discount <= 0:
                continue
            promo_multiplier = max(0.0, 1 - discount / 100)
            promo_before = expected_running
            expected_running = promo_before * promo_multiplier
            offer_name = str(promo.get("offer_name") or "Promotion").strip() or "Promotion"
            applied_promotions.append(offer_name)
            applied_rules.append(f"promotion:{offer_name}:{discount:.2f}%")
            promotion_breakdown.append(
                {
                    "offer_name": offer_name,
                    "discount_percent": round(discount, 2),
                    "multiplier": round(promo_multiplier, 4),
                    "subtotal_before": round(promo_before, 2),
                    "adjustment_amount": round(promo_before - expected_running, 2),
                    "subtotal_after": round(expected_running, 2),
                }
            )

        expected_price = round(expected_running, 2)
        variance_amount = round(actual_price - expected_price, 2)
        variance_percent = round((variance_amount / expected_price * 100.0), 2) if expected_price > 0 else 0.0
        promo_code_value = str(line_data.get("promo_code") or "").strip() or None

        is_match = abs(variance_amount) <= tolerance_amount or abs(variance_percent) <= tolerance_percent
        reason = (
            "Pricing aligns with contract rules"
            if is_match
            else "Variance detected. Likely manual configuration error or missing promotion in PMS"
        )
        if missing_base_dates:
            is_match = False
            preview = ", ".join(missing_base_dates[:5])
            if len(missing_base_dates) > 5:
                preview = f"{preview}, +{len(missing_base_dates) - 5} more"
            requested_board = self._normalize_board_type(str(line_data.get("board_type") or "")).strip() or "UNKNOWN"
            if board_missing_dates:
                available = ", ".join(sorted(available_board_types)) or "N/A"
                reason = (
                    f"Missing daily rates for board {requested_board} in uploaded price list for this period "
                    f"({preview}). Available board(s) for same room/date: {available}."
                )
            else:
                reason = (
                    "Missing daily rates in uploaded price list for this reservation period "
                    f"({preview})."
                )
        elif promotions_require_booking_date and booking_date_missing_for_promotion_check and not applied_promotions:
            reason = (
                f"{reason}. Booking date is missing, so early-booking windows could not be evaluated."
            )

        sorted_sources = sorted(base_source_counts.items(), key=lambda item: item[0])
        if missing_base_dates:
            base_rate_source = "incomplete_pricelist_match"
        elif len(sorted_sources) == 1:
            base_rate_source = sorted_sources[0][0]
        elif sorted_sources:
            base_rate_source = "mixed"
        else:
            base_rate_source = "unknown"
        source_details = (
            ", ".join(f"{source}:{count} night(s)" for source, count in sorted_sources)
            if sorted_sources
            else "No pricing source detected."
        )
        if missing_base_dates:
            source_details = (
                f"{source_details} Missing nights in price-list coverage: {len(missing_base_dates)}."
            )
        if board_missing_dates:
            available = ", ".join(sorted(available_board_types)) or "N/A"
            source_details = (
                f"{source_details} No rates for requested board {self._normalize_board_type(str(line_data.get('board_type') or '')).strip() or 'UNKNOWN'} "
                f"on {len(board_missing_dates)} night(s); available board(s): {available}."
            )
        if board_supplement_adjustments:
            supplements_summary = ", ".join(
                (
                    f"{item['source_board_type'] or '-'}->{item['target_board_type']}:"
                    f"{item['unit_amount']:.2f}/night x {item['nights_applied']}"
                )
                for item in board_supplement_adjustments
            )
            source_details = f"{source_details} Applied board supplement(s): {supplements_summary}."
        if promotions_require_booking_date and booking_date_missing_for_promotion_check and not applied_promotions:
            source_details = f"{source_details} Missing booking_date prevented booking-window promotion checks."

        return {
            "reservation_id": str(line_data.get("reservation_id")),
            "hotel_code": str(line_data.get("hotel_code")),
            "operator_code": str(line_data.get("operator_code")),
            "room_type": str(line_data.get("room_type")),
            "board_type": str(line_data.get("board_type")),
            "booking_date": booking_date.isoformat() if booking_date else None,
            "stay_date": stay_date.isoformat(),
            "nights": nights,
            "pax_adults": pax_adults,
            "pax_children": pax_children,
            "promo_code": promo_code_value,
            "expected_price": expected_price,
            "actual_price": actual_price,
            "variance_amount": variance_amount,
            "variance_percent": variance_percent,
            "status": "match" if is_match else "mismatch",
            "reason": reason,
            "applied_promotions": applied_promotions,
            "applied_rules": applied_rules,
            "expected_calculation": {
                "base_rate": average_base_rate,
                "base_rate_source": base_rate_source,
                "source_details": source_details,
                "room_board_key": f"{room_key or '-'}::{board_key or '-'}",
                "nights": nights,
                "base_adult_units": base_adult_units,
                "base_subtotal": base_subtotal,
                "nightly_base_rates": nightly_base_rates,
                "missing_base_rate_dates": missing_base_dates,
                "available_board_types": sorted(available_board_types),
                "board_supplement_adjustments": board_supplement_adjustments,
                "guest_adjustments": guest_adjustments,
                "child_discount": child_discount_breakdown,
                "promotion_adjustments": promotion_breakdown,
                "final_expected_price": expected_price,
            },
        }

    def _promotion_applies(
        self,
        promotion: dict,
        stay_date: date,
        promo_code: str | None,
        booking_date: date | None = None,
        room_type: str | None = None,
        board_type: str | None = None,
    ) -> bool:
        start_date, end_date, booking_start_date, booking_end_date = self._resolve_promotion_window_dates(promotion)
        is_early_booking = self._is_early_booking_promotion(promotion)

        # Early-booking must always evaluate both booking and stay windows.
        if is_early_booking:
            if not (booking_start_date or booking_end_date):
                return False
            if not (start_date or end_date):
                return False

        if booking_start_date or booking_end_date:
            if booking_date is None:
                return False
            if booking_start_date and booking_date < booking_start_date:
                return False
            if booking_end_date and booking_date > booking_end_date:
                return False

        if start_date and stay_date < start_date:
            return False
        if end_date and stay_date > end_date:
            return False

        room_filters = self._normalize_promotion_scope_values(promotion.get("applicable_room_types"))
        if room_filters and room_type:
            room_value = str(room_type).strip().lower()
            if room_value and not any(token in room_value for token in room_filters):
                return False

        board_filters = self._normalize_promotion_scope_values(promotion.get("applicable_board_types"))
        if board_filters and board_type:
            board_value = self._normalize_board_type(str(board_type))
            board_lower = board_value.lower()
            if board_lower and not any(token in board_lower for token in board_filters):
                return False

        # For now, do not gate early-booking promotions by promo_code token matching.
        # They are eligibility-based (booking/stay windows), not code-based.
        if promo_code and not is_early_booking:
            promo_normalized = promo_code.strip().lower()
            offer_name = str(promotion.get("offer_name") or "").lower()
            return promo_normalized in offer_name

        return True

    def _load_reconciliation_workbook(self, file_name: str, content: bytes) -> object:
        suffix = Path(file_name).suffix.lower()
        if suffix not in ALLOWED_RECONCILIATION_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Reconciliation upload must be an Excel file (.xlsx/.xlsm/.xls).",
            )

        try:
            from openpyxl import load_workbook

            return load_workbook(filename=BytesIO(content), data_only=True, read_only=True, keep_vba=True)
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Could not parse reconciliation file.",
                event="reconciliation.workbook_parse_failed",
                exc=exc,
            )

    def _resolve_reconciliation_sheet(self, workbook: object, sheet_name: str) -> object:
        for sheet in workbook.worksheets:
            if str(sheet.title).strip() == sheet_name:
                return sheet
        requested_lower = sheet_name.strip().lower()
        for sheet in workbook.worksheets:
            if str(sheet.title).strip().lower() == requested_lower:
                return sheet
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Sheet '{sheet_name}' was not found in the uploaded workbook.",
        )

    def _build_reconciliation_sheet_preview(self, sheet: object, sample_rows: int) -> dict:
        rows = list(sheet.iter_rows(min_row=1, max_row=1800, values_only=True))
        if not rows:
            return {
                "sheet_name": str(sheet.title),
                "total_rows": 0,
                "non_empty_rows": 0,
                "column_count": 0,
                "detected_header_row": None,
                "detected_fields": [],
                "confidence": 0.0,
                "sample_headers": [],
                "sample_rows": [],
            }

        header_row_index, header_map = self._detect_reconciliation_header(rows)
        non_empty_rows = 0
        max_columns = 0
        for row in rows:
            if row is None:
                continue
            if any(value is not None and str(value).strip() for value in row):
                non_empty_rows += 1
                max_columns = max(max_columns, len(row))

        header_values: list[object] = []
        if header_row_index is not None and 0 <= header_row_index < len(rows):
            header_values = list(rows[header_row_index] or [])
        if not any(self._stringify_sheet_cell(value) for value in header_values):
            for candidate in rows[:12]:
                candidate_values = list(candidate or [])
                if any(self._stringify_sheet_cell(value) for value in candidate_values):
                    header_values = candidate_values
                    break

        sample_headers = [self._stringify_sheet_cell(value) for value in header_values[: max(max_columns, 1)]]
        sample_headers = [value if value else f"col_{index + 1}" for index, value in enumerate(sample_headers)]

        start_row = (header_row_index + 1) if header_row_index is not None else 0
        preview_rows: list[list[str]] = []
        for row in rows[start_row:]:
            values = [self._stringify_sheet_cell(value) for value in list(row or [])[: max(max_columns, 1)]]
            if not any(values):
                continue
            preview_rows.append(values)
            if len(preview_rows) >= sample_rows:
                break

        confidence = 0.0
        if max_columns > 0:
            header_ratio = min(1.0, len(header_map) / 8.0)
            density_ratio = min(1.0, non_empty_rows / 120.0)
            confidence = round((header_ratio * 0.7) + (density_ratio * 0.3), 2)

        return {
            "sheet_name": str(sheet.title),
            "total_rows": len(rows),
            "non_empty_rows": non_empty_rows,
            "column_count": max_columns,
            "detected_header_row": (header_row_index + 1) if header_row_index is not None else None,
            "detected_fields": sorted(header_map.keys()),
            "confidence": confidence,
            "sample_headers": sample_headers[:24],
            "sample_rows": [row[:24] for row in preview_rows[:sample_rows]],
        }

    def _build_reconciliation_ai_payload(
        self,
        file_name: str,
        content: bytes,
        sheet_name: str,
        max_rows: int,
    ) -> dict:
        workbook = self._load_reconciliation_workbook(file_name=file_name, content=content)
        try:
            sheet = self._resolve_reconciliation_sheet(workbook=workbook, sheet_name=sheet_name)
            rows = list(sheet.iter_rows(min_row=1, max_row=3000, values_only=True))
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        if not rows:
            return {"sheet_name": sheet_name, "header_row": None, "detected_fields": [], "columns": [], "column_labels": [], "rows": []}

        header_row_index, header_map = self._detect_reconciliation_header(rows)
        start_row = (header_row_index + 1) if header_row_index is not None else 0
        max_columns = min(36, max((len(row or []) for row in rows), default=0))

        header_values: list[object] = []
        if header_row_index is not None and 0 <= header_row_index < len(rows):
            header_values = list(rows[header_row_index] or [])
        column_names: list[str] = []
        column_labels: list[str] = []
        used_names: dict[str, int] = {}
        for index in range(max_columns):
            raw_header = header_values[index] if index < len(header_values) else None
            header_label = self._stringify_sheet_cell(raw_header) or f"col_{index + 1}"
            normalized_header = self._normalize_header_token(raw_header)
            if normalized_header:
                base_name = normalized_header.replace(" ", "_")
            else:
                base_name = f"col_{index + 1}"
            suffix = used_names.get(base_name, 0)
            used_names[base_name] = suffix + 1
            column_name = base_name if suffix == 0 else f"{base_name}_{suffix + 1}"
            column_names.append(column_name)
            column_labels.append(header_label)

        date_like_columns: set[int] = set()
        for index, column_name in enumerate(column_names):
            label = column_labels[index] if index < len(column_labels) else ""
            if self._reconciliation_column_looks_like_date(column_name=column_name, column_label=label):
                date_like_columns.add(index)

        payload_rows: list[dict] = []
        empty_streak = 0
        for row_number, row in enumerate(rows[start_row:], start=start_row + 1):
            if len(payload_rows) >= max_rows:
                break
            values = list(row or [])
            if not any(value is not None and str(value).strip() for value in values):
                empty_streak += 1
                if payload_rows and empty_streak >= 50:
                    break
                continue
            empty_streak = 0

            mapped_values: dict[str, str] = {}
            for index in range(min(max_columns, len(values))):
                raw_value = values[index]
                if index in date_like_columns:
                    coerced_date = self._coerce_optional_date(raw_value)
                    if coerced_date:
                        cell_text = coerced_date.isoformat()
                    else:
                        cell_text = self._stringify_sheet_cell(raw_value)
                else:
                    cell_text = self._stringify_sheet_cell(raw_value)
                if cell_text:
                    mapped_values[column_names[index]] = cell_text
            if not mapped_values:
                continue
            payload_rows.append({"row_number": row_number, "values": mapped_values})

        return {
            "sheet_name": sheet_name,
            "header_row": (header_row_index + 1) if header_row_index is not None else None,
            "detected_fields": sorted(header_map.keys()),
            "columns": column_names,
            "column_labels": column_labels,
            "rows": payload_rows,
        }

    def _reconciliation_column_looks_like_date(self, *, column_name: str, column_label: str) -> bool:
        token_name = self._normalize_header_token(str(column_name).replace("_", " "))
        token_label = self._normalize_header_token(column_label)
        token = f"{token_name} {token_label}".strip()
        if not token:
            return False

        if any(
            marker in token
            for marker in (
                "check in",
                "check out",
                "arrival",
                "departure",
                "stay date",
                "booking date",
            )
        ):
            return True
        return bool(re.search(r"(^|\\s)date($|\\s)", token))

    def _resolve_reconciliation_id_column_key(self, *, ai_payload: dict, selected_column: str | None) -> str | None:
        raw_selected = str(selected_column or "").strip()
        if not raw_selected:
            return None

        columns = [str(item).strip() for item in ai_payload.get("columns", []) if str(item).strip()]
        labels = [str(item).strip() for item in ai_payload.get("column_labels", [])]
        if not columns:
            return None

        selected_token = self._normalize_header_token(raw_selected).replace(" ", "_")

        for column in columns:
            if column == raw_selected or column.lower() == raw_selected.lower() or column == selected_token:
                return column

        if labels and len(labels) == len(columns):
            normalized_labels = [self._normalize_header_token(label).replace(" ", "_") for label in labels]
            for index, label_token in enumerate(normalized_labels):
                if label_token and (label_token == selected_token or label_token == raw_selected.lower()):
                    return columns[index]

        for column in columns:
            if selected_token and (column.startswith(selected_token) or selected_token.startswith(column)):
                return column

        return None

    def _build_reservation_id_value_lookup(self, *, ai_payload: dict, reservation_id_column_key: str | None) -> dict[int, str]:
        if not reservation_id_column_key:
            return {}

        lookup: dict[int, str] = {}
        raw_rows = ai_payload.get("rows")
        if not isinstance(raw_rows, list):
            return lookup

        for item in raw_rows:
            if not isinstance(item, dict):
                continue
            row_number = self._coerce_int(item.get("row_number"), default=0)
            if row_number <= 0:
                continue
            values = item.get("values")
            if not isinstance(values, dict):
                continue
            value = str(values.get(reservation_id_column_key) or "").strip()
            if value:
                lookup[row_number] = value

        return lookup

    def _resolve_reconciliation_header_mapping(
        self,
        *,
        ai_payload: dict,
        raw_header_mapping: object,
        reservation_id_column_key: str | None,
    ) -> dict[str, str]:
        columns = [str(item).strip() for item in ai_payload.get("columns", []) if str(item).strip()]
        if not columns:
            return {}

        column_labels_raw = ai_payload.get("column_labels")
        column_labels = (
            [str(item).strip() for item in column_labels_raw]
            if isinstance(column_labels_raw, list)
            else []
        )
        labels_aligned = [
            (column_labels[index] if index < len(column_labels) and column_labels[index] else column)
            for index, column in enumerate(columns)
        ]

        def resolve_field_key(token: object) -> str | None:
            text = str(token or "").strip()
            if not text:
                return None
            direct = text.lower().replace(" ", "_")
            if direct in RECONCILIATION_HEADER_SYNONYMS:
                return direct
            normalized = self._normalize_header_token(text)
            normalized_key = normalized.replace(" ", "_")
            if normalized_key in RECONCILIATION_HEADER_SYNONYMS:
                return normalized_key
            for field_name, aliases in RECONCILIATION_HEADER_SYNONYMS.items():
                field_token = field_name.replace("_", " ")
                if normalized == field_token or field_token in normalized:
                    return field_name
                if any(alias in normalized for alias in aliases):
                    return field_name
            return None

        def resolve_source_column(token: object) -> str | None:
            text = str(token or "").strip()
            if not text:
                return None
            return self._resolve_reconciliation_id_column_key(ai_payload=ai_payload, selected_column=text)

        resolved: dict[str, str] = {}

        # Baseline mapping from detected headers keeps the flow deterministic if AI mapping is sparse.
        auto_header_map = self._map_reconciliation_headers(labels_aligned)
        for field_name, column_index in auto_header_map.items():
            if 0 <= column_index < len(columns):
                resolved[field_name] = columns[column_index]

        # AI mapping may be either "target -> source" or "source -> target". Support both forms.
        if isinstance(raw_header_mapping, dict):
            for raw_key, raw_value in raw_header_mapping.items():
                key_field = resolve_field_key(raw_key)
                value_field = resolve_field_key(raw_value)
                key_column = resolve_source_column(raw_key)
                value_column = resolve_source_column(raw_value)

                if key_field and value_column:
                    resolved[key_field] = value_column
                    continue
                if value_field and key_column:
                    resolved[value_field] = key_column
                    continue
                if key_field and not value_column:
                    # Some models may return normalized payload keys directly.
                    normalized_value = str(raw_value or "").strip()
                    if normalized_value in columns:
                        resolved[key_field] = normalized_value
                        continue
                if value_field and not key_column:
                    normalized_key = str(raw_key or "").strip()
                    if normalized_key in columns:
                        resolved[value_field] = normalized_key

        if reservation_id_column_key and reservation_id_column_key in columns:
            resolved["reservation_id"] = reservation_id_column_key

        return {
            field_name: column_key
            for field_name, column_key in resolved.items()
            if field_name in RECONCILIATION_HEADER_SYNONYMS and column_key in columns
        }

    def _pick_reconciliation_payload_value(
        self,
        *,
        row_values: dict[str, object],
        resolved_header_mapping: dict[str, str],
        field_name: str,
        column_label_by_key: dict[str, str],
    ) -> object | None:
        mapped_column = resolved_header_mapping.get(field_name)
        if mapped_column:
            value = row_values.get(mapped_column)
            if value is not None and str(value).strip():
                return value

        aliases = RECONCILIATION_HEADER_SYNONYMS.get(field_name, ())
        best_score = -1
        best_value: object | None = None
        field_token = field_name.replace("_", " ")

        for column_key, value in row_values.items():
            if value is None or not str(value).strip():
                continue
            header_text = column_label_by_key.get(column_key, column_key)
            key_token = self._normalize_header_token(column_key)
            label_token = self._normalize_header_token(header_text)

            score = 0
            if key_token == field_token or label_token == field_token:
                score = max(score, 120)
            if key_token == field_name or key_token == field_name.replace("_", ""):
                score = max(score, 110)
            if any(alias == key_token or alias == label_token for alias in aliases):
                score = max(score, 100)
            if any(alias in key_token or alias in label_token for alias in aliases):
                score = max(score, 70)
            if field_token in key_token or field_token in label_token:
                score = max(score, 60)

            if score > best_score:
                best_score = score
                best_value = value

        if best_score >= 60:
            return best_value
        return None

    def _reconciliation_column_looks_like_status(self, *, column_name: str, column_label: str) -> bool:
        token_name = self._normalize_header_token(str(column_name).replace("_", " "))
        token_label = self._normalize_header_token(column_label)
        token = f"{token_name} {token_label}".strip()
        if not token:
            return False
        if "policy" in token and "status" not in token and "state" not in token:
            return False
        return any(marker in token for marker in RECONCILIATION_STATUS_COLUMN_HINTS)

    def _is_cancelled_reconciliation_status_value(self, value: object) -> bool:
        token = self._normalize_header_token(value)
        if not token:
            return False
        compact = token.replace(" ", "")

        # Guard against explicit negation such as "not cancelled".
        if any(negation in token for negation in ("not cancelled", "not canceled", "non cancelled", "non canceled")):
            return False
        if any(
            phrase in token
            for phrase in (
                "cancellation policy",
                "free cancellation",
                "cancellation fee",
                "cancellation charge",
                "cancel fee",
                "cancel charge",
            )
        ):
            return False

        if token in RECONCILIATION_CANCELLED_STATUS_TOKENS:
            return True
        if compact in RECONCILIATION_CANCELLED_STATUS_COMPACT_TOKENS:
            return True

        if token.startswith(("cancel", "cxl", "void", "annul")):
            return True
        if "no show" in token or "noshow" in compact:
            return True
        return False

    def _is_cancelled_reconciliation_row_values(
        self,
        *,
        row_values: dict[str, object],
        resolved_header_mapping: dict[str, str],
        column_label_by_key: dict[str, str],
    ) -> bool:
        explicit_status = self._pick_reconciliation_payload_value(
            row_values=row_values,
            resolved_header_mapping=resolved_header_mapping,
            field_name="status",
            column_label_by_key=column_label_by_key,
        )
        if self._is_cancelled_reconciliation_status_value(explicit_status):
            return True

        for column_key, raw_value in row_values.items():
            if raw_value is None or not str(raw_value).strip():
                continue
            column_label = column_label_by_key.get(column_key, column_key)
            if not self._reconciliation_column_looks_like_status(
                column_name=column_key,
                column_label=column_label,
            ):
                continue
            if self._is_cancelled_reconciliation_status_value(raw_value):
                return True
        return False

    def _is_cancelled_ai_reconciliation_row(self, item: dict) -> bool:
        if not isinstance(item, dict):
            return False

        for bool_key in ("is_cancelled", "cancelled", "canceled", "is_voided", "voided"):
            bool_value = self._coerce_optional_bool(item.get(bool_key))
            if bool_value is True:
                return True

        for status_key in (
            "status",
            "reservation_status",
            "booking_status",
            "reservation_state",
            "booking_state",
            "status_code",
        ):
            if self._is_cancelled_reconciliation_status_value(item.get(status_key)):
                return True

        for key, value in item.items():
            if value is None or not str(value).strip():
                continue
            if not self._reconciliation_column_looks_like_status(column_name=str(key), column_label=str(key)):
                continue
            if self._is_cancelled_reconciliation_status_value(value):
                return True

        return False

    def _reconciliation_header_looks_like_nights(
        self,
        *,
        column_key: str | None,
        column_label_by_key: dict[str, str],
    ) -> bool:
        if not column_key:
            return False

        header_candidates = [column_key, column_label_by_key.get(column_key, column_key)]
        nights_hints = ("night", "nights", "los", "length of stay", "stay length")
        exclusion_hints = ("price", "amount", "cost", "rate", "total", "revenue")

        for candidate in header_candidates:
            token = self._normalize_header_token(candidate)
            if not token:
                continue
            if any(hint in token for hint in exclusion_hints):
                return False
            if any(hint in token for hint in nights_hints):
                return True
        return False

    def _calculate_reconciliation_nights_from_dates(
        self,
        *,
        check_in_date: date | None,
        check_out_date: date | None,
        stay_date: date | None,
    ) -> int | None:
        if check_in_date and check_out_date and check_out_date > check_in_date:
            nights = (check_out_date - check_in_date).days
            if 1 <= nights <= 60:
                return nights

        anchor_date = stay_date or check_in_date
        if anchor_date and check_out_date and check_out_date > anchor_date:
            nights = (check_out_date - anchor_date).days
            if 1 <= nights <= 60:
                return nights

        return None

    def _infer_reconciliation_nights_from_row_dates(
        self,
        *,
        row_values: dict[str, object],
        resolved_header_mapping: dict[str, str] | None,
        column_label_by_key: dict[str, str],
        stay_date: date,
    ) -> int | None:
        start_hints = ("check in", "check-in", "arrival", "start", "from")
        end_hints = ("check out", "check-out", "departure", "end", "to")
        booking_hints = ("booking", "reserved", "created", "issued")

        explicit_check_in: date | None = None
        explicit_check_out: date | None = None
        if resolved_header_mapping:
            mapped_check_in_column = resolved_header_mapping.get("check_in_date")
            mapped_check_out_column = resolved_header_mapping.get("check_out_date")
            explicit_check_in = self._coerce_optional_date(
                row_values.get(mapped_check_in_column) if mapped_check_in_column else None
            )
            explicit_check_out = self._coerce_optional_date(
                row_values.get(mapped_check_out_column) if mapped_check_out_column else None
            )

        explicit_nights = self._calculate_reconciliation_nights_from_dates(
            check_in_date=explicit_check_in,
            check_out_date=explicit_check_out,
            stay_date=stay_date,
        )
        if explicit_nights is not None:
            return explicit_nights

        start_dates: list[date] = []
        end_dates: list[date] = []
        all_dates: list[date] = []

        for column_key, raw_value in row_values.items():
            parsed_date = self._coerce_optional_date(raw_value)
            if not parsed_date:
                continue

            header_text = column_label_by_key.get(column_key, column_key)
            header_token = self._normalize_header_token(header_text)
            key_token = self._normalize_header_token(column_key)
            combined_token = f"{header_token} {key_token}".strip()

            all_dates.append(parsed_date)

            if any(hint in combined_token for hint in booking_hints):
                continue

            if any(hint in combined_token for hint in start_hints):
                start_dates.append(parsed_date)
            if any(hint in combined_token for hint in end_hints):
                end_dates.append(parsed_date)

        if start_dates and end_dates:
            anchor_start = min(start_dates, key=lambda candidate: abs((candidate - stay_date).days))
            best_end = min((candidate for candidate in end_dates if candidate > anchor_start), default=None)
            derived = self._calculate_reconciliation_nights_from_dates(
                check_in_date=anchor_start,
                check_out_date=best_end,
                stay_date=stay_date,
            )
            if derived is not None:
                return derived

        if end_dates:
            best_end = min((candidate for candidate in end_dates if candidate > stay_date), default=None)
            derived = self._calculate_reconciliation_nights_from_dates(
                check_in_date=None,
                check_out_date=best_end,
                stay_date=stay_date,
            )
            if derived is not None:
                return derived

        if all_dates:
            future_distances = sorted(
                {(candidate - stay_date).days for candidate in all_dates if candidate > stay_date and (candidate - stay_date).days <= 60}
            )
            if future_distances:
                return future_distances[0]

        return None

    def _infer_reconciliation_booking_date_from_row_dates(
        self,
        *,
        row_values: dict[str, object],
        resolved_header_mapping: dict[str, str] | None,
        column_label_by_key: dict[str, str],
        stay_date: date,
    ) -> date | None:
        if resolved_header_mapping:
            mapped_booking_column = resolved_header_mapping.get("booking_date")
            explicit_booking = self._coerce_optional_date(
                row_values.get(mapped_booking_column) if mapped_booking_column else None
            )
            if explicit_booking:
                return explicit_booking

        booking_hints = (
            "booking",
            "booked",
            "book date",
            "date booked",
            "reservation created",
            "created",
            "issued",
            "issue",
        )
        stay_hints = (
            "check in",
            "check-in",
            "check out",
            "check-out",
            "arrival",
            "departure",
            "stay",
            "start",
            "end",
            "from",
            "to",
        )

        scored_candidates: list[tuple[int, date]] = []
        generic_candidates: list[tuple[date, str]] = []

        for column_key, raw_value in row_values.items():
            parsed_date = self._coerce_optional_date(raw_value)
            if not parsed_date:
                continue

            header_text = column_label_by_key.get(column_key, column_key)
            header_token = self._normalize_header_token(header_text)
            key_token = self._normalize_header_token(column_key)
            combined_token = f"{header_token} {key_token}".strip()

            generic_candidates.append((parsed_date, combined_token))

            score = 0
            if any(hint in combined_token for hint in booking_hints):
                score += 100
            if "date" in combined_token or "time" in combined_token:
                score += 10
            if "created" in combined_token or "issued" in combined_token:
                score += 10
            if any(hint in combined_token for hint in stay_hints):
                score -= 40
            if score >= 60:
                scored_candidates.append((score, parsed_date))

        if scored_candidates:
            max_score = max(item[0] for item in scored_candidates)
            top_dates = [candidate_date for score, candidate_date in scored_candidates if score == max_score]
            not_after_stay = [candidate_date for candidate_date in top_dates if candidate_date <= stay_date]
            if not_after_stay:
                return max(not_after_stay)
            return min(top_dates)

        not_after_stay = [
            candidate_date
            for candidate_date, token in generic_candidates
            if candidate_date <= stay_date and not any(hint in token for hint in stay_hints)
        ]
        if not_after_stay:
            return max(not_after_stay)

        prior_dates = [candidate_date for candidate_date, _ in generic_candidates if candidate_date < stay_date]
        if prior_dates:
            return max(prior_dates)

        same_or_before = [candidate_date for candidate_date, _ in generic_candidates if candidate_date <= stay_date]
        if same_or_before:
            return max(same_or_before)

        return None

    def _build_reconciliation_lines_from_header_mapping(
        self,
        *,
        sheet_payload: dict,
        resolved_header_mapping: dict[str, str],
        contract_id: str,
        default_hotel: str,
        default_operator: str,
        sheet_title: str,
        max_lines: int,
        reservation_id_values: dict[int, str] | None = None,
    ) -> list[dict]:
        raw_rows = sheet_payload.get("rows")
        if not isinstance(raw_rows, list):
            return []

        column_keys = [str(item).strip() for item in sheet_payload.get("columns", []) if str(item).strip()]
        column_labels_raw = sheet_payload.get("column_labels")
        column_labels = (
            [str(item).strip() for item in column_labels_raw]
            if isinstance(column_labels_raw, list)
            else []
        )
        column_label_by_key = {
            key: (column_labels[index] if index < len(column_labels) and column_labels[index] else key)
            for index, key in enumerate(column_keys)
        }

        normalized: list[dict] = []
        for index, item in enumerate(raw_rows, start=1):
            if len(normalized) >= max_lines:
                break
            if not isinstance(item, dict):
                continue

            row_number = self._coerce_int(item.get("row_number"), default=index)
            values = item.get("values")
            if not isinstance(values, dict):
                continue
            row_values = {str(key): value for key, value in values.items()}
            if not any(value is not None and str(value).strip() for value in row_values.values()):
                continue
            if self._is_cancelled_reconciliation_row_values(
                row_values=row_values,
                resolved_header_mapping=resolved_header_mapping,
                column_label_by_key=column_label_by_key,
            ):
                continue

            actual_price = self._coerce_float(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="actual_price",
                    column_label_by_key=column_label_by_key,
                )
            )
            if actual_price is None:
                continue

            stay_date = self._coerce_optional_date(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="stay_date",
                    column_label_by_key=column_label_by_key,
                )
            ) or datetime.now(timezone.utc).date()
            booking_date = self._coerce_optional_date(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="booking_date",
                    column_label_by_key=column_label_by_key,
                )
            )
            if booking_date is None:
                booking_date = self._infer_reconciliation_booking_date_from_row_dates(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    column_label_by_key=column_label_by_key,
                    stay_date=stay_date,
                )

            reservation = ""
            if reservation_id_values:
                reservation = str(reservation_id_values.get(row_number) or "").strip()
            if not reservation:
                reservation = str(
                    self._pick_reconciliation_payload_value(
                        row_values=row_values,
                        resolved_header_mapping=resolved_header_mapping,
                        field_name="reservation_id",
                        column_label_by_key=column_label_by_key,
                    )
                    or ""
                ).strip()
            if not reservation:
                reservation = f"AUTO-{sheet_title}-{row_number}"

            hotel_code = str(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="hotel_code",
                    column_label_by_key=column_label_by_key,
                )
                or ""
            ).strip().upper() or default_hotel

            operator_code = str(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="operator_code",
                    column_label_by_key=column_label_by_key,
                )
                or ""
            ).strip().upper() or default_operator

            room_type_raw = str(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="room_type",
                    column_label_by_key=column_label_by_key,
                )
                or ""
            ).strip()
            if not room_type_raw:
                room_type_raw = "Standard Room"

            board_type_raw = str(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="board_type",
                    column_label_by_key=column_label_by_key,
                )
                or ""
            ).strip()

            raw_nights_value = self._pick_reconciliation_payload_value(
                row_values=row_values,
                resolved_header_mapping=resolved_header_mapping,
                field_name="nights",
                column_label_by_key=column_label_by_key,
            )
            raw_nights = self._coerce_int(raw_nights_value, default=0)
            mapped_nights_column = resolved_header_mapping.get("nights")
            mapped_nights_is_semantic = self._reconciliation_header_looks_like_nights(
                column_key=mapped_nights_column,
                column_label_by_key=column_label_by_key,
            )
            inferred_nights = self._infer_reconciliation_nights_from_row_dates(
                row_values=row_values,
                resolved_header_mapping=resolved_header_mapping,
                column_label_by_key=column_label_by_key,
                stay_date=stay_date,
            )

            if not mapped_nights_is_semantic and inferred_nights is not None:
                nights = inferred_nights
            elif raw_nights <= 0:
                nights = inferred_nights if inferred_nights is not None else 1
            elif raw_nights > 60 and inferred_nights is not None:
                nights = inferred_nights
            elif raw_nights >= 45 and inferred_nights is not None and 1 <= inferred_nights <= 31:
                nights = inferred_nights
            else:
                nights = raw_nights

            nights = max(1, min(60, nights))
            pax_adults = max(
                1,
                min(
                    10,
                    self._coerce_int(
                        self._pick_reconciliation_payload_value(
                            row_values=row_values,
                            resolved_header_mapping=resolved_header_mapping,
                            field_name="pax_adults",
                            column_label_by_key=column_label_by_key,
                        ),
                        default=2,
                    ),
                ),
            )
            pax_children = max(
                0,
                min(
                    10,
                    self._coerce_int(
                        self._pick_reconciliation_payload_value(
                            row_values=row_values,
                            resolved_header_mapping=resolved_header_mapping,
                            field_name="pax_children",
                            column_label_by_key=column_label_by_key,
                        ),
                        default=0,
                    ),
                ),
            )
            contract_rate = self._coerce_float(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="contract_rate",
                    column_label_by_key=column_label_by_key,
                )
            )
            promo_code_raw = str(
                self._pick_reconciliation_payload_value(
                    row_values=row_values,
                    resolved_header_mapping=resolved_header_mapping,
                    field_name="promo_code",
                    column_label_by_key=column_label_by_key,
                )
                or ""
            ).strip()

            normalized.append(
                {
                    "reservation_id": reservation,
                    "hotel_code": hotel_code,
                    "operator_code": operator_code,
                    "contract_id": contract_id,
                    "room_type": self._to_title_case(room_type_raw),
                    "board_type": self._normalize_board_type(board_type_raw),
                    "booking_date": booking_date,
                    "stay_date": stay_date,
                    "nights": nights,
                    "pax_adults": pax_adults,
                    "pax_children": pax_children,
                    "actual_price": round(actual_price, 2),
                    "contract_rate": round(contract_rate, 2) if contract_rate is not None else None,
                    "promo_code": promo_code_raw or None,
                }
            )

        return normalized

    async def _run_openai_reconciliation_mapping(
        self,
        *,
        model: str | None,
        source_system: str | None,
        contract: dict,
        rules: list[dict],
        promotions: list[dict],
        sheet_payload: dict,
        reservation_id_column_key: str | None,
        mapping_instructions: str | None,
    ) -> tuple[dict, dict, str]:
        settings = get_settings()
        if not settings.openai_api_key:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="OPENAI_API_KEY is not configured.",
            )

        try:
            from openai import AsyncOpenAI
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="OpenAI SDK is not installed. Add `openai` to backend requirements.",
            ) from exc

        selected_model = model.strip() if model and model.strip() else settings.openai_pricing_model
        configured_base_url = (settings.openai_base_url or "").strip()
        client = AsyncOpenAI(
            api_key=settings.openai_api_key,
            base_url=configured_base_url or "https://api.openai.com/v1",
            timeout=None,
        )

        contract_context = {
            "contract_id": contract.get("id"),
            "hotel_code": contract.get("hotel_code"),
            "operator_code": contract.get("operator_code"),
            "season_label": contract.get("season_label"),
            "room_types": list(contract.get("extraction", {}).get("room_types", []))[:40],
            "board_types": list(contract.get("extraction", {}).get("board_types", []))[:20],
            "rules_summary": [
                {
                    "name": item.get("name"),
                    "type": item.get("rule_type"),
                    "metadata": item.get("metadata", {}),
                }
                for item in rules[:80]
            ],
            "promotions": [
                {
                    "offer_name": promo.get("offer_name"),
                    "discount_percent": promo.get("discount_percent"),
                    "booking_start_date": promo.get("booking_start_date"),
                    "booking_end_date": promo.get("booking_end_date"),
                    "arrival_start_date": promo.get("arrival_start_date") or promo.get("start_date"),
                    "arrival_end_date": promo.get("arrival_end_date") or promo.get("end_date"),
                    "applicable_room_types": promo.get("applicable_room_types", []),
                    "applicable_board_types": promo.get("applicable_board_types", []),
                    "promotion_category": promo.get("promotion_category"),
                }
                for promo in promotions[:30]
            ],
        }

        payload_json = json.dumps(sheet_payload, ensure_ascii=False, default=str)[:180000]
        context_json = json.dumps(contract_context, ensure_ascii=False, default=str)[:50000]
        source_system_text = str(source_system or "").strip() or "unknown"
        additional_instructions = str(mapping_instructions or "").strip()

        prompt = (
            "Map reservation rows from an Excel reconciliation sheet into a strict hospitality validation shape.\n"
            "The target shape represents one reservation line used to validate PMS/accounting prices against contract rules and promotions.\n"
            "Prioritize accurate header_mapping and include only a small representative sample of mapped lines (max 3).\n"
            "Do not attempt to map the full dataset in this response; downstream code applies the mapping to all rows.\n"
            "Keep numeric amounts precise.\n"
            "Normalize board_type to common codes where possible (RO/BB/HB/FB/AI).\n"
            "Use contract context to normalize room and board naming. Keep uncertainty low and avoid inventing values.\n"
            "Do not include rows without a usable actual_price.\n"
            "If a reservation status column exists, map it to status and exclude cancelled/canceled/cxl/void/no-show rows.\n"
            "If reservation_id is missing, generate a stable synthetic value based on row number.\n"
            "stay_date must be ISO date (YYYY-MM-DD).\n\n"
            "If a dedicated nights/LOS column exists, map it to nights.\n"
            "If nights is not explicitly available, map check_in_date and check_out_date to allow deriving nights.\n\n"
            "If booking/reservation creation date is available, map it to booking_date (ISO YYYY-MM-DD).\n\n"
            f"Source system hint: {source_system_text}\n\n"
            f"Contract context:\n{context_json}\n\n"
            f"Sheet payload:\n{payload_json}\n"
        )
        if reservation_id_column_key:
            prompt += (
                "\nReservation ID mapping requirement:\n"
                f"- Use column `{reservation_id_column_key}` as reservation_id whenever present.\n"
                "- Preserve exact source value (trim surrounding spaces only).\n"
                "- Treat this as the canonical unique reservation identifier.\n"
            )
        if additional_instructions:
            prompt += f"\nModel-driven mapping instructions:\n{additional_instructions}\n"

        system_prompt = (
            "You are a senior hospitality revenue-control analyst. "
            "Map reservation exports from tour operators, accounting, or PMS systems to pricing validation lines. "
            "Return strict JSON only."
        )

        try:
            completion = await client.chat.completions.create(
                model=selected_model,
                temperature=0,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "reconciliation_sheet_mapping",
                        "schema": AI_RECONCILIATION_MAPPING_SCHEMA,
                    },
                },
            )
        except Exception as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider request failed.",
                event="openai.reconciliation_mapping.request_failed",
                exc=exc,
            )

        if not completion.choices:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI reconciliation mapping returned no choices.",
            )

        message = completion.choices[0].message
        refusal = getattr(message, "refusal", None)
        if refusal:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                user_message="Upstream AI provider refused reconciliation mapping.",
                event="openai.reconciliation_mapping.refused",
                context={"refusal": str(refusal)[:400]},
            )

        content = message.content or ""
        if not isinstance(content, str) or not content.strip():
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="OpenAI reconciliation mapping returned empty content.",
            )

        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as exc:
            self._raise_sanitized_http_error(
                status_code=status.HTTP_502_BAD_GATEWAY,
                user_message="Upstream AI provider returned invalid JSON.",
                event="openai.reconciliation_mapping.invalid_json",
                exc=exc,
            )

        usage_obj = getattr(completion, "usage", None)
        usage = {}
        if usage_obj:
            usage = {
                "prompt_tokens": int(getattr(usage_obj, "prompt_tokens", 0) or 0),
                "completion_tokens": int(getattr(usage_obj, "completion_tokens", 0) or 0),
                "total_tokens": int(getattr(usage_obj, "total_tokens", 0) or 0),
            }

        return parsed if isinstance(parsed, dict) else {}, usage, selected_model

    def _normalize_ai_reconciliation_lines(
        self,
        *,
        ai_lines: object,
        contract_id: str,
        default_hotel: str,
        default_operator: str,
        sheet_title: str,
        max_lines: int,
        reservation_id_values: dict[int, str] | None = None,
    ) -> list[dict]:
        if not isinstance(ai_lines, list):
            return []

        normalized: list[dict] = []
        for index, item in enumerate(ai_lines, start=1):
            if len(normalized) >= max_lines:
                break
            if not isinstance(item, dict):
                continue
            if self._is_cancelled_ai_reconciliation_row(item):
                continue

            actual_price = self._coerce_float(item.get("actual_price") or item.get("amount") or item.get("price"))
            if actual_price is None:
                continue

            stay_date = self._coerce_optional_date(
                item.get("stay_date")
                or item.get("arrival_date")
                or item.get("check_in")
                or item.get("check_in_date")
                or item.get("date")
            ) or datetime.now(timezone.utc).date()
            booking_date = self._coerce_optional_date(
                item.get("booking_date")
                or item.get("reservation_date")
                or item.get("booked_on")
                or item.get("booking_created_at")
                or item.get("booking_datetime")
                or item.get("date_booked")
                or item.get("reservation_created_at")
                or item.get("created_at")
                or item.get("issue_date")
            )

            row_hint = self._coerce_int(item.get("row_number"), default=index)
            reservation = ""
            if reservation_id_values:
                reservation = str(reservation_id_values.get(row_hint) or "").strip()
            if not reservation:
                reservation = str(
                    item.get("reservation_id")
                    or item.get("booking_reference")
                    or item.get("reservation")
                    or item.get("reference")
                    or ""
                ).strip()
            if not reservation:
                reservation = f"AUTO-{sheet_title}-{row_hint}"

            room_type_raw = str(item.get("room_type") or item.get("room") or "").strip()
            board_type_raw = str(item.get("board_type") or item.get("meal_plan") or item.get("board") or "").strip()
            if not room_type_raw:
                room_type_raw = "Standard Room"

            hotel_code = str(item.get("hotel_code") or "").strip().upper() or default_hotel
            operator_code = str(item.get("operator_code") or "").strip().upper() or default_operator

            nights = max(1, min(60, self._coerce_int(item.get("nights"), default=1)))
            pax_adults = max(1, min(10, self._coerce_int(item.get("pax_adults") or item.get("adults"), default=2)))
            pax_children = max(0, min(10, self._coerce_int(item.get("pax_children") or item.get("children"), default=0)))
            contract_rate = self._coerce_float(item.get("contract_rate") or item.get("expected_rate"))
            promo_code_raw = str(item.get("promo_code") or item.get("promotion_code") or "").strip()

            normalized.append(
                {
                    "reservation_id": reservation,
                    "hotel_code": hotel_code,
                    "operator_code": operator_code,
                    "contract_id": contract_id,
                    "room_type": self._to_title_case(room_type_raw),
                    "board_type": self._normalize_board_type(board_type_raw),
                    "booking_date": booking_date,
                    "stay_date": stay_date,
                    "nights": nights,
                    "pax_adults": pax_adults,
                    "pax_children": pax_children,
                    "actual_price": round(actual_price, 2),
                    "contract_rate": round(contract_rate, 2) if contract_rate is not None else None,
                    "promo_code": promo_code_raw or None,
                }
            )

        return normalized

    async def _extract_reconciliation_lines(
        self,
        file_name: str,
        content: bytes,
        contract_id: str,
        sheet_name: str | None,
        max_lines: int,
    ) -> list[dict]:
        upload_limits = await self._get_effective_upload_limits_bytes()
        file_name = self._ensure_file_payload(
            file_name=file_name,
            content=content,
            allowed_extensions=ALLOWED_RECONCILIATION_EXTENSIONS,
            max_bytes=upload_limits["reconciliation"],
            label="reconciliation",
            empty_message="Reconciliation file is empty.",
        )

        contract = await self.get_contract(contract_id)
        workbook = self._load_reconciliation_workbook(file_name=file_name, content=content)

        lines: list[dict] = []
        try:
            if sheet_name and sheet_name.strip():
                sheet = self._resolve_reconciliation_sheet(workbook=workbook, sheet_name=sheet_name.strip())
                extracted = self._extract_reconciliation_from_sheet(
                    sheet=sheet,
                    contract_id=contract_id,
                    default_hotel=contract["hotel_code"],
                    default_operator=contract["operator_code"],
                    max_lines=max_lines,
                )
                lines.extend(extracted)
            else:
                for sheet in workbook.worksheets[:8]:
                    remaining = max_lines - len(lines)
                    if remaining <= 0:
                        break
                    extracted = self._extract_reconciliation_from_sheet(
                        sheet=sheet,
                        contract_id=contract_id,
                        default_hotel=contract["hotel_code"],
                        default_operator=contract["operator_code"],
                        max_lines=remaining,
                    )
                    lines.extend(extracted)
        finally:
            try:
                workbook.close()
            except Exception:
                pass
        return lines

    def _extract_reconciliation_from_sheet(
        self,
        sheet: object,
        contract_id: str,
        default_hotel: str,
        default_operator: str,
        max_lines: int,
    ) -> list[dict]:
        rows = list(sheet.iter_rows(min_row=1, max_row=1500, values_only=True))
        if not rows:
            return []

        header_row_index, header_map = self._detect_reconciliation_header(rows)
        start_row = (header_row_index + 1) if header_row_index is not None else 0
        fallback_map = {
            "reservation_id": 0,
            "hotel_code": 1,
            "operator_code": 2,
            "room_type": 3,
            "board_type": 4,
            "stay_date": 5,
            "nights": 6,
            "pax_adults": 7,
            "pax_children": 8,
            "actual_price": 9,
            "contract_rate": 10,
            "promo_code": 11,
            "booking_date": 12,
        }

        extracted: list[dict] = []
        empty_streak = 0
        for offset, row in enumerate(rows[start_row:], start=start_row + 1):
            if len(extracted) >= max_lines:
                break
            if row is None:
                continue

            raw_values = [cell for cell in row]
            if not any(value is not None and str(value).strip() for value in raw_values):
                empty_streak += 1
                if extracted and empty_streak >= 35:
                    break
                continue
            empty_streak = 0

            line = self._build_reconciliation_line(
                row=raw_values,
                header_map=header_map,
                fallback_map=fallback_map,
                contract_id=contract_id,
                default_hotel=default_hotel,
                default_operator=default_operator,
                sheet_title=str(sheet.title),
                row_number=offset,
            )
            if line:
                extracted.append(line)

        return extracted

    def _detect_reconciliation_header(self, rows: list[tuple | list]) -> tuple[int | None, dict[str, int]]:
        best_index: int | None = None
        best_score = -1
        best_map: dict[str, int] = {}

        for index, row in enumerate(rows[:18]):
            if row is None:
                continue
            header_map = self._map_reconciliation_headers(list(row))
            score = len(header_map)
            if score > best_score:
                best_score = score
                best_index = index
                best_map = header_map

        if best_score >= 3 and best_index is not None:
            return best_index, best_map
        return None, {}

    def _map_reconciliation_headers(self, header_row: list[object]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for index, cell in enumerate(header_row):
            token = self._normalize_header_token(cell)
            if not token:
                continue

            for field_name, aliases in RECONCILIATION_HEADER_SYNONYMS.items():
                if field_name in mapped:
                    continue
                if any(alias in token for alias in aliases):
                    mapped[field_name] = index
                    break
        return mapped

    def _normalize_header_token(self, value: object) -> str:
        if value is None:
            return ""
        token = str(value).strip().lower()
        token = re.sub(r"[^a-z0-9]+", " ", token)
        token = re.sub(r"\s+", " ", token).strip()
        return token

    def _stringify_sheet_cell(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.4f}".rstrip("0").rstrip(".")
        return str(value).strip()

    def _build_reconciliation_line(
        self,
        row: list[object],
        header_map: dict[str, int],
        fallback_map: dict[str, int],
        contract_id: str,
        default_hotel: str,
        default_operator: str,
        sheet_title: str,
        row_number: int,
    ) -> dict | None:
        def pick(field_name: str) -> object:
            if field_name in header_map and header_map[field_name] < len(row):
                return row[header_map[field_name]]
            fallback_index = fallback_map.get(field_name)
            if fallback_index is not None and fallback_index < len(row):
                return row[fallback_index]
            return None

        if self._is_cancelled_reconciliation_status_value(pick("status")):
            return None

        actual_price = self._coerce_float(pick("actual_price"))
        if actual_price is None:
            return None

        stay_date = self._coerce_optional_date(pick("stay_date")) or datetime.now(timezone.utc).date()
        booking_date = self._coerce_optional_date(pick("booking_date"))
        room_type = str(pick("room_type") or "").strip()
        board_type = self._normalize_board_type(str(pick("board_type") or "").strip())

        reservation = str(pick("reservation_id") or "").strip()
        if not reservation:
            reservation = f"AUTO-{sheet_title}-{row_number}"

        hotel_code = str(pick("hotel_code") or "").strip().upper() or default_hotel
        operator_code = str(pick("operator_code") or "").strip().upper() or default_operator

        raw_nights = self._coerce_int(pick("nights"), default=0)
        check_in_date = self._coerce_optional_date(pick("check_in_date"))
        check_out_date = self._coerce_optional_date(pick("check_out_date"))
        inferred_nights = self._calculate_reconciliation_nights_from_dates(
            check_in_date=check_in_date,
            check_out_date=check_out_date,
            stay_date=stay_date,
        )
        if raw_nights <= 0 and inferred_nights is not None:
            nights = inferred_nights
        elif raw_nights > 60 and inferred_nights is not None:
            nights = inferred_nights
        elif raw_nights >= 45 and inferred_nights is not None and 1 <= inferred_nights <= 31:
            nights = inferred_nights
        else:
            nights = raw_nights if raw_nights > 0 else 1
        nights = max(1, min(60, nights))
        pax_adults = max(1, min(10, self._coerce_int(pick("pax_adults"), default=2)))
        pax_children = max(0, min(10, self._coerce_int(pick("pax_children"), default=0)))
        contract_rate = self._coerce_float(pick("contract_rate"))
        promo_code_raw = str(pick("promo_code") or "").strip()

        return {
            "reservation_id": reservation,
            "hotel_code": hotel_code,
            "operator_code": operator_code,
            "contract_id": contract_id,
            "room_type": self._to_title_case(room_type) if room_type else "Standard Room",
            "board_type": board_type,
            "booking_date": booking_date,
            "stay_date": stay_date,
            "nights": nights,
            "pax_adults": pax_adults,
            "pax_children": pax_children,
            "actual_price": round(actual_price, 2),
            "contract_rate": round(contract_rate, 2) if contract_rate is not None else None,
            "promo_code": promo_code_raw or None,
        }

    def _extract_keyword_lines(self, text: str, keywords: list[str]) -> list[str]:
        lines: list[str] = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            lower = line.lower()
            if any(keyword in lower for keyword in keywords):
                lines.append(line)
        return lines

    def _normalize_board_type(self, raw_value: str) -> str:
        normalized = raw_value.strip().lower()
        if not normalized:
            return "RO"
        for key, board_code in BOARD_TYPE_LOOKUP.items():
            if key in normalized:
                return board_code
        return raw_value.strip().upper()[:10]

    def _normalize_reservation_unique_key(self, reservation_id: str, sheet_name: str, row_number: int) -> str:
        raw = str(reservation_id or "").strip()
        if not raw:
            raw = f"AUTO-{sheet_name}-{row_number}"
        normalized = re.sub(r"\s+", "", raw).upper()
        return normalized[:220]

    def _normalize_reservation_group_key(self, reservation_id: str, sheet_name: str, row_number: int) -> str:
        raw = str(reservation_id or "").strip()
        if not raw:
            raw = f"AUTO-{sheet_name}-{row_number}"
        normalized = re.sub(r"[^A-Za-z0-9]+", "", raw).upper()
        if not normalized:
            normalized = re.sub(r"\s+", "", raw).upper()
        return normalized[:220]

    def _coerce_int(self, value: object, default: int) -> int:
        if value is None:
            return default
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            match = re.search(r"-?\d+", value.replace(",", ""))
            if match:
                try:
                    return int(match.group(0))
                except ValueError:
                    return default
        return default

    def _guess_content_type(self, *, file_name: str, fallback: str | None = None) -> str:
        guessed, _ = mimetypes.guess_type(file_name)
        normalized_guess = (guessed or "").split(";")[0].strip().lower()
        normalized_fallback = (fallback or "").split(";")[0].strip().lower()

        if normalized_guess and normalized_guess not in BLOCKED_RESPONSE_CONTENT_TYPES:
            return normalized_guess
        if normalized_fallback and normalized_fallback not in BLOCKED_RESPONSE_CONTENT_TYPES:
            return normalized_fallback
        return "application/octet-stream"

    def _coerce_binary_content(self, value: object) -> bytes | None:
        if value is None:
            return None
        if isinstance(value, bytes):
            return value
        if isinstance(value, bytearray):
            return bytes(value)
        if isinstance(value, memoryview):
            return value.tobytes()
        try:
            if hasattr(value, "__bytes__"):
                coerced = bytes(value)
                if coerced:
                    return coerced
        except Exception:
            return None
        return None

    def _to_bson_compatible(self, value: object) -> object:
        if isinstance(value, datetime):
            if value.tzinfo is None:
                return value.replace(tzinfo=timezone.utc)
            return value.astimezone(timezone.utc)
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
        if isinstance(value, dict):
            return {str(key): self._to_bson_compatible(item) for key, item in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [self._to_bson_compatible(item) for item in value]
        return value

    def _coerce_float(self, value: object) -> float | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            normalized = value.strip().replace(",", "")
            if not normalized:
                return None
            match = re.search(r"-?\d+(?:\.\d+)?", normalized)
            if match:
                try:
                    return float(match.group(0))
                except ValueError:
                    return None
        return None

    def _coerce_optional_bool(self, value: object) -> bool | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            token = value.strip().lower()
            if token in {"true", "yes", "y", "1"}:
                return True
            if token in {"false", "no", "n", "0"}:
                return False
        return None

    def _coerce_optional_date(self, value: object) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # Excel serial date numbers are usually > 10_000 in modern files.
            if value > 10000:
                try:
                    return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
                except Exception:
                    return None
        if isinstance(value, str):
            parsed = self._parse_date_string(value)
            return parsed
        return None

    def _coerce_date(self, value: object) -> date:
        maybe = self._coerce_optional_date(value)
        if maybe:
            return maybe
        return datetime.now(timezone.utc).date()

    def _parse_date_string(self, value: str) -> date | None:
        normalized = value.strip()
        if not normalized:
            return None

        numeric_token = normalized.replace(",", ".")
        # Handle Excel serial dates that arrive as strings (e.g. "45757" or "45757.0").
        if re.fullmatch(r"\d+(?:\.\d+)?", numeric_token):
            try:
                serial_value = float(numeric_token)
            except ValueError:
                serial_value = 0.0
            if 10_000 < serial_value < 80_000:
                try:
                    return (datetime(1899, 12, 30) + timedelta(days=serial_value)).date()
                except Exception:
                    pass

        for candidate in (normalized, normalized.replace("Z", "+00:00")):
            try:
                return datetime.fromisoformat(candidate).date()
            except ValueError:
                continue

        for fmt in (
            "%d/%m/%Y",
            "%d/%m/%y",
            "%d-%m-%Y",
            "%d-%m-%y",
            "%d.%m.%Y",
            "%d.%m.%y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%d %b %Y",
            "%d %B %Y",
            "%d %b %y",
            "%d %B %y",
            "%d/%m/%Y %H:%M:%S",
            "%d-%m-%Y %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%Y.%m.%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(normalized, fmt).date()
            except ValueError:
                continue
        return None

    def _to_title_case(self, value: str) -> str:
        return re.sub(r"\s+", " ", value).strip().title()

    def _average_rate_lookup(self, lookup: dict[tuple[str, str], float]) -> float:
        if not lookup:
            return 0.0
        values = list(lookup.values())
        return round(sum(values) / len(values), 2)
